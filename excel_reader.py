import time
from pathlib import Path
import openpyxl
import config


def _open_ledger():
    path = config.LEDGER_PATH
    if not Path(path).exists():
        raise FileNotFoundError(f"Ledger not found: {path}")
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def read_portfolio_data():
    try:
        wb = _open_ledger()
    except FileNotFoundError as e:
        return {"error": str(e)}
    ws = wb["PORTFOLIO"]
    holdings = []
    current_section = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(v is not None for v in row):
            continue
        section_raw = str(row[0] or "").strip()
        ticker_raw  = str(row[1] or "").strip() if len(row) > 1 else ""
        name        = str(row[2] or "").strip() if len(row) > 2 else ""
        if section_raw:
            current_section = section_raw
        if (not ticker_raw or "Checking" in ticker_raw
                or name.upper().startswith("TOTAL") or name.upper().startswith("SUMMARY")
                or not name):
            continue
        try:
            shares = float(row[3]) if isinstance(row[3], (int, float)) else (float(row[3]) if row[3] else 0.0)
        except (ValueError, TypeError):
            shares = 0.0
        avg_cost     = float(row[4]) if len(row) > 4 and isinstance(row[4], (int, float)) else None
        cached_price = float(row[5]) if len(row) > 5 and isinstance(row[5], (int, float)) else None
        is_crypto = ticker_raw.startswith("CURRENCY:")
        is_proxy  = ("proxy" in current_section.lower() or "401" in current_section.lower()
                     or "voya" in current_section.lower())
        av_symbol = ticker_raw.replace("MUTF:", "").replace("CURRENCY:", "").split("USD")[0]
        holdings.append({
            "section":      current_section,
            "ticker":       av_symbol,
            "name":         name,
            "shares":       shares,
            "avg_cost":     avg_cost,
            "cached_price": cached_price,
            "is_crypto":    is_crypto,
            "is_proxy":     is_proxy,
        })
    wb.close()
    return holdings


def fetch_av_price(symbol, api_key, is_crypto=False):
    import requests
    if is_crypto:
        url = (f"https://www.alphavantage.co/query?function=DIGITAL_CURRENCY_DAILY"
               f"&symbol={symbol}&market=USD&apikey={api_key}")
        data = requests.get(url, timeout=10).json()
        ts = data.get("Time Series (Digital Currency Daily)", {})
        if ts:
            latest = sorted(ts.keys())[-1]
            return float(ts[latest]["4a. close (USD)"])
    else:
        url = (f"https://www.alphavantage.co/query?function=GLOBAL_QUOTE"
               f"&symbol={symbol}&apikey={api_key}")
        data = requests.get(url, timeout=10).json()
        price = data.get("Global Quote", {}).get("05. price")
        if price:
            return float(price)
    return None


def read_dashboard_data():
    try:
        wb = _open_ledger()
    except FileNotFoundError as e:
        return {"error": str(e)}

    ws_dash = wb["DASHBOARD"]
    kv = {}
    freedom_levels = []
    allocation = {}
    cashflow = {}
    in_levels = in_alloc = in_cashflow = False

    for row in ws_dash.iter_rows(min_row=2, values_only=True):
        label = row[0] if len(row) > 0 else None
        value = row[1] if len(row) > 1 else None
        if label is None:
            continue
        if label == "FINANCIAL FREEDOM LEVELS":
            in_levels = True; in_alloc = False; in_cashflow = False; continue
        if label == "ASSET ALLOCATION (For Pie Chart)":
            in_alloc = True; in_levels = False; in_cashflow = False; continue
        if label == "CASH FLOW (For Bar Chart)":
            in_cashflow = True; in_alloc = False; in_levels = False; continue
        if label == "QUICK ACTIONS":
            in_levels = False; in_alloc = False; in_cashflow = False; continue

        if in_levels and value is not None:
            status = row[2] if len(row) > 2 else None
            freedom_levels.append({
                "name":      label,
                "goal":      value if isinstance(value, (int, float)) else None,
                "goal_text": value if isinstance(value, str) else None,
                "status":    status if isinstance(status, str) else None,
                "progress":  float(status) if isinstance(status, float) else None,
            })
        elif in_alloc and value is not None:
            allocation[label] = float(value)
        elif in_cashflow and value is not None:
            cashflow[label] = float(value)
        else:
            kv[label] = value

    ws_spend = wb["SPENDING"]
    months = []
    spending = {}
    for row in ws_spend.iter_rows(min_row=1, max_row=10, values_only=True):
        if row[0] in ("TYPE", None):
            if row[0] == "TYPE":
                months = [str(m) for m in row[1:] if m is not None]
            continue
        if row[0] and any(v is not None for v in row[1:]):
            spending[row[0]] = [v for v in row[1:] if v is not None]

    nw = {
        "ss_monthly_62": 0.0, "ss_monthly_67": 0.0, "ss_monthly_70": 0.0,
        "checking_balance": 0.0, "sgov_balance": 0.0, "total_invested": 0.0,
        "monthly_burn": 0.0, "net_monthly_income": 0.0,
    }
    ws_nw = wb["NET_WORTH"]
    for row in ws_nw.iter_rows(min_row=1, values_only=True):
        if not row or not any(v is not None for v in row):
            continue
        c0 = str(row[0] or "").strip()
        c1 = str(row[1] or "").strip() if len(row) > 1 else ""
        v1 = row[1] if len(row) > 1 and isinstance(row[1], (int, float)) else None
        v2 = row[2] if len(row) > 2 and isinstance(row[2], (int, float)) else None
        if   "SS Benefit @ 62"    in c0 and v1 is not None: nw["ss_monthly_62"]      = float(v1)
        elif "SS Benefit @ 67"    in c0 and v1 is not None: nw["ss_monthly_67"]      = float(v1)
        elif "SS Benefit @ 70"    in c0 and v1 is not None: nw["ss_monthly_70"]      = float(v1)
        elif "Monthly Burn"       in c0 and v1 is not None: nw["monthly_burn"]        = float(v1)
        elif "Net Monthly Income" in c0 and v1 is not None: nw["net_monthly_income"]  = float(v1)
        elif "Checking" in c1 and "Ops" in c1 and v2 is not None: nw["checking_balance"] = float(v2)
        elif "SGOV"         in c1 and v2 is not None: nw["sgov_balance"]   = float(v2)
        elif "TOTAL INVESTED" in c1 and v2 is not None: nw["total_invested"] = float(v2)

    wb.close()

    engine_bal = max(0.0, nw["total_invested"] - nw["sgov_balance"])
    mc_prefill = {
        "current_age":        None,
        "engine_balance":     round(engine_bal),
        "sgov_balance":       round(nw["sgov_balance"]),
        "checking_balance":   round(nw["checking_balance"]),
        "full_ss_annual":     round(nw["ss_monthly_67"] * 12),
        "ss_monthly_67":      nw["ss_monthly_67"],
        "ss_monthly_62":      nw["ss_monthly_62"],
        "ss_monthly_70":      nw["ss_monthly_70"],
        "monthly_burn":       round(nw["monthly_burn"]),
        "annual_floor_cost":  round(nw["monthly_burn"] * 12),
        "net_monthly_income": round(nw["net_monthly_income"]),
    }

    return {
        "metrics":        kv,
        "allocation":     allocation,
        "cashflow":       cashflow,
        "freedom_levels": freedom_levels,
        "spending_months": months,
        "spending":       spending,
        "mc_prefill":     mc_prefill,
    }


def read_roadmap_data():
    try:
        wb = _open_ledger()
    except FileNotFoundError as e:
        return {"error": str(e)}
    ws = wb["ROADMAP"]
    cfg = {}
    rows = []
    in_data = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not any(v is not None for v in row):
            continue
        c0 = str(row[0] or "").strip()
        if c0 == "Year":
            in_data = True
            continue
        if not in_data:
            v = row[1] if len(row) > 1 else None
            if c0 and v is not None:
                cfg[c0.rstrip(":")] = v
        else:
            if row[0] is None:
                continue
            try:
                year = int(float(row[0]))
            except Exception:
                continue
            rows.append({
                "year":     year,
                "age":      int(row[1]) if row[1] is not None else None,
                "phase":    str(row[2] or ""),
                "sgov":     round(float(row[3] or 0), 2),
                "schwab":   round(float(row[4] or 0), 2),
                "roth":     round(float(row[5] or 0), 2),
                "liquid_nw": round(float(row[6] or 0), 2),
                "k401":     round(float(row[7] or 0), 2),
                "total_nw": round(float(row[8] or 0), 2),
            })
    wb.close()
    return {"config": cfg, "rows": rows}


def read_transactions_data(page=1, limit=50, month_filter=None, type_filter=None):
    try:
        wb = _open_ledger()
    except FileNotFoundError as e:
        return {"error": str(e)}
    ws = wb["TRANSACTIONS"]
    all_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        date = row[1]
        if date is None:
            continue
        month = row[0]
        month_str = month.strftime("%Y-%m") if hasattr(month, "strftime") else str(month or "")[:7]
        if month_filter and month_str != month_filter:
            continue
        txtype = str(row[2] or "").strip()
        if type_filter and txtype.lower() != type_filter.lower():
            continue
        all_rows.append({
            "month":    month_str,
            "date":     date.strftime("%Y-%m-%d") if hasattr(date, "strftime") else str(date),
            "type":     txtype,
            "category": str(row[3] or "").strip(),
            "amount":   round(float(row[4] or 0), 2),
            "account":  str(row[5] or "").strip(),
            "memo":     str(row[6] or "").strip(),
            "signed":   round(float(row[7] or 0), 2),
        })
    wb.close()
    all_rows.sort(key=lambda r: r["date"], reverse=True)
    months = sorted({r["month"] for r in all_rows}, reverse=True)
    types  = sorted({r["type"]  for r in all_rows if r["type"]})
    total  = len(all_rows)
    start  = (page - 1) * limit
    return {
        "total": total, "page": page, "limit": limit,
        "pages": max(1, (total + limit - 1) // limit),
        "rows":  all_rows[start:start + limit],
        "months": months, "types": types,
    }


def read_forecast_data():
    try:
        wb = _open_ledger()
    except FileNotFoundError as e:
        return {"error": str(e)}
    ws = wb["FORECAST_V3"]
    calib = None
    rows = []
    in_data = False
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if not any(v is not None for v in row):
            continue
        c0 = row[0]
        if i == 2:
            calib = {
                "current_date":       c0.strftime("%Y-%m-%d") if hasattr(c0, "strftime") else str(c0),
                "checking_balance":   round(float(row[1] or 0), 2),
                "savings_balance":    round(float(row[2] or 0), 2),
                "projected_checking": round(float(row[5] or 0), 2) if len(row) > 5 else 0,
                "projected_savings":  round(float(row[6] or 0), 2) if len(row) > 6 else 0,
                "projected_total":    round(float(row[7] or 0), 2) if len(row) > 7 else 0,
            }
        if str(c0 or "") == "DATE":
            in_data = True
            continue
        if in_data and hasattr(c0, "strftime"):
            rows.append({
                "date":     c0.strftime("%Y-%m-%d"),
                "day":      str(row[1] or ""),
                "income":   round(float(row[2] or 0), 2),
                "expense":  round(float(row[3] or 0), 2),
                "invest":   round(float(row[4] or 0), 2),
                "checking": round(float(row[5] or 0), 2),
                "savings":  round(float(row[6] or 0), 2),
                "total":    round(float(row[7] or 0), 2),
            })
    wb.close()
    return {"calibration": calib, "rows": rows}


def read_tax_loss_data():
    try:
        wb = _open_ledger()
    except FileNotFoundError as e:
        return {"error": str(e)}
    ws = wb["TAX-LOSS"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row[:5]):
            continue
        date_val = row[0]
        date_str = date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else str(date_val or "")
        rows.append({
            "date":   date_str,
            "action": str(row[1] or "").strip(),
            "amount": round(float(row[2] or 0), 2),
            "notes":  str(row[3] or "").strip(),
            "signed": round(float(row[4] or 0), 2),
        })
    wb.close()
    rows.sort(key=lambda r: r["date"], reverse=True)
    net_carryover    = round(sum(r["signed"] for r in rows), 2)
    total_harvested  = round(sum(r["signed"] for r in rows if r["signed"] < 0), 2)
    total_realized   = round(sum(r["signed"] for r in rows if r["signed"] > 0), 2)
    return {
        "rows":            rows,
        "net_carryover":   net_carryover,
        "total_harvested": total_harvested,
        "total_realized":  total_realized,
        "entry_count":     len(rows),
    }
