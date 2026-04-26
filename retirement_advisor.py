from mcp.server.fastmcp import FastMCP
import requests
import openpyxl
import os
import math
import random
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import numpy as np
from starlette.applications import Starlette
from starlette.routing import Route, Mount
from starlette.responses import JSONResponse, HTMLResponse
from starlette.requests import Request
import uvicorn
import asyncio
from concurrent.futures import ProcessPoolExecutor
import itertools

LEDGER_PATH = os.environ.get('LEDGER_PATH', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Road To FI.xlsx'))
# Set OLLAMA_URL env var to override if Ollama container uses bridge networking (not host)
OLLAMA_URL  = os.environ.get('OLLAMA_URL', 'http://172.17.0.1:11434')

SYSTEM_PROMPT = (
    "You are a sharp, direct fiduciary retirement advisor for Kurtis — "
    "a 44-year-old Best Buy employee on track to retire at 62. You know his plan inside out. "
    "Speak like a trusted advisor, not a chatbot. Be specific, reference exact numbers, "
    "skip generic disclaimers. If a number isn't in the data, say so — never invent figures.\n\n"
    "PERSONAL RETIREMENT STRATEGY — Money Machine V2.6:\n"
    "- Retire at 62 (target). Work identity kept until then; no early exit pressure.\n"
    "- SGOV Bridge Moat: $300,000 carved into SGOV at retirement to cover the 62–67 gap. "
    "  This funds a strict floor draw without touching the equity engine during sequence risk.\n"
    "- Social Security: Claim at 67. Projected $36,697/yr. This covers the entire annual floor.\n"
    "- Protocol Alpha: At 67, SS income = floor cost → 0% withdrawal rate from investments. "
    "  The portfolio becomes a pure wealth-compounding machine post-67.\n"
    "- Lifestyle Ratchet: Discretionary spending unlocks only when portfolio hits 1.5× the "
    "  retirement-entry balance. Prevents lifestyle creep in early retirement.\n"
    "- Spending Smile: Go-go (62–75) active travel spending, slow-go (75–85) declining, "
    "  no-go (85+) minimal. Skim rates adjust by phase.\n"
    "- Freedom Levels: A tiered system of financial milestones tracked in the ledger. "
    "  Reference achieved vs pending levels when relevant.\n"
    "- Equity engine: VTI-heavy brokerage (Schwab). Not touched during bridge period.\n"
    "- Key risk: Sequence-of-returns in first 5 years of retirement. The SGOV moat is the "
    "  primary defense. Secondary: Ripcord — claim SS early at 62 if moat runway < 5 years.\n\n"
    "RESPONSE STYLE:\n"
    "- 2–4 paragraphs for analysis; bullet points for action items\n"
    "- Always lead with the number that matters most\n"
    "- Name risks plainly (e.g., 'moat breach rate is X% — that's elevated')\n"
    "- No filler phrases like 'Great question!' or 'It's important to note that'\n"
    "- Today is {today}"
)

# ── MCP Server ────────────────────────────────────────────────────────────────

mcp = FastMCP("RetirementAuditor", host="0.0.0.0", port=8000)

@mcp.resource("finance://2026_rules")
def get_2026_rules() -> str:
    return """
    2026 RETIREMENT LIMITS (GROUND TRUTH):
    - 401k/403b Limit: $24,500
    - Standard Catch-up (Age 50+): $8,000
    - 'Super' Catch-up (Age 60-63): $11,250
    - IRA/Roth Limit: $7,500 ($8,600 if 50+)
    - ROTH-IFICATION: If 2025 income >$145k, catch-ups MUST be Roth.
    """

@mcp.tool()
def get_stock_price(ticker: str, api_key: str) -> str:
    """Gets the current price for any stock ticker."""
    url = f'https://www.alphavantage.co/query?function=GLOBAL_QUOTE&symbol={ticker}&apikey={api_key}'
    r = requests.get(url, timeout=10)
    data = r.json()
    return f"Current price for {ticker}: ${data['Global Quote']['05. price']}"

@mcp.tool()
def get_fi_dashboard() -> str:
    """Returns a snapshot of the Road To FI dashboard — net worth, FI progress, freedom levels."""
    wb = None
    try:
        wb = openpyxl.load_workbook(LEDGER_PATH, read_only=True, data_only=True)
        ws = wb['DASHBOARD']
        rows = {r[0]: r[1] for r in ws.iter_rows(min_row=2, values_only=True) if r[0] and len(r) > 1 and r[1] is not None}
        pct = float(rows.get('PROGRESS TO FI', 0)) * 100
        return (
            f"FREEDOM LEDGER SNAPSHOT\n"
            f"  Liquid Net Worth : ${rows.get('LIQUID NET WORTH', 0):,.0f}\n"
            f"  Total Net Worth  : ${rows.get('TOTAL NET WORTH', 0):,.0f}\n"
            f"  Liquid Cash      : ${rows.get('LIQUID CASH', 0):,.0f}\n"
            f"  Survival Runway  : {rows.get('SURVIVAL RUNWAY', 'N/A')}\n"
            f"  FI Target (62)   : ${rows.get('FI TARGET (Age 62)', 0):,.0f}\n"
            f"  Progress to FI   : {pct:.2f}%"
        )
    except Exception as e:
        return f"Error reading ledger: {e}"
    finally:
        if wb:
            wb.close()

# ── Portfolio helpers ─────────────────────────────────────────────────────────

def read_portfolio_data():
    wb = openpyxl.load_workbook(LEDGER_PATH, read_only=True, data_only=True)
    ws = wb['PORTFOLIO']
    holdings = []
    current_section = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(v is not None for v in row):
            continue
        section_raw  = str(row[0] or '').strip()
        ticker_raw   = str(row[1] or '').strip() if len(row) > 1 else ''
        name         = str(row[2] or '').strip() if len(row) > 2 else ''
        if section_raw:
            current_section = section_raw
        # Skip non-holding rows
        if (not ticker_raw or 'Checking' in ticker_raw
                or name.upper().startswith('TOTAL') or name.upper().startswith('SUMMARY')
                or not name):
            continue
        try:
            shares = float(row[3]) if isinstance(row[3], (int, float)) else (float(row[3]) if row[3] else 0.0)
        except (ValueError, TypeError):
            shares = 0.0
        avg_cost     = float(row[4]) if len(row) > 4 and isinstance(row[4], (int, float)) else None
        cached_price = float(row[5]) if len(row) > 5 and isinstance(row[5], (int, float)) else None
        is_crypto = ticker_raw.startswith('CURRENCY:')
        is_proxy  = ('proxy' in current_section.lower() or '401' in current_section.lower()
                     or 'voya' in current_section.lower())
        # Strip exchange prefixes for Alpha Vantage
        av_symbol = ticker_raw.replace('MUTF:', '').replace('CURRENCY:', '').split('USD')[0]
        holdings.append({
            "section":       current_section,
            "ticker":        av_symbol,
            "name":          name,
            "shares":        shares,
            "avg_cost":      avg_cost,
            "cached_price":  cached_price,
            "is_crypto":     is_crypto,
            "is_proxy":      is_proxy,
        })
    wb.close()
    return holdings

def fetch_av_price(symbol, api_key, is_crypto=False):
    if is_crypto:
        url = (f'https://www.alphavantage.co/query?function=DIGITAL_CURRENCY_DAILY'
               f'&symbol={symbol}&market=USD&apikey={api_key}')
        data = requests.get(url, timeout=10).json()
        ts = data.get('Time Series (Digital Currency Daily)', {})
        if ts:
            latest = sorted(ts.keys())[-1]
            return float(ts[latest]['4a. close (USD)'])
    else:
        url = (f'https://www.alphavantage.co/query?function=GLOBAL_QUOTE'
               f'&symbol={symbol}&apikey={api_key}')
        data = requests.get(url, timeout=10).json()
        price = data.get('Global Quote', {}).get('05. price')
        if price:
            return float(price)
    return None

# ── Ledger reader helper ───────────────────────────────────────────────────────

def read_dashboard_data():
    wb = openpyxl.load_workbook(LEDGER_PATH, read_only=True, data_only=True)

    # DASHBOARD tab — key/value pairs
    ws_dash = wb['DASHBOARD']
    kv = {}
    freedom_levels = []
    allocation = {}
    cashflow = {}
    in_levels = False
    in_alloc = False
    in_cashflow = False

    for row in ws_dash.iter_rows(min_row=2, values_only=True):
        label = row[0] if len(row) > 0 else None
        value = row[1] if len(row) > 1 else None
        if label is None:
            continue
        if label == 'FINANCIAL FREEDOM LEVELS':
            in_levels = True; in_alloc = False; in_cashflow = False; continue
        if label == 'ASSET ALLOCATION (For Pie Chart)':
            in_alloc = True; in_levels = False; in_cashflow = False; continue
        if label == 'CASH FLOW (For Bar Chart)':
            in_cashflow = True; in_alloc = False; in_levels = False; continue
        if label == 'QUICK ACTIONS':
            in_levels = False; in_alloc = False; in_cashflow = False; continue

        if in_levels and value is not None:
            status = row[2] if len(row) > 2 else None
            freedom_levels.append({
                "name": label,
                "goal": value if isinstance(value, (int, float)) else None,
                "goal_text": value if isinstance(value, str) else None,
                "status": status if isinstance(status, str) else None,
                "progress": float(status) if isinstance(status, float) else None,
            })
        elif in_alloc and value is not None:
            allocation[label] = float(value)
        elif in_cashflow and value is not None:
            cashflow[label] = float(value)
        else:
            kv[label] = value

    # SPENDING tab — last 5 months
    ws_spend = wb['SPENDING']
    months = []
    spending = {}
    for row in ws_spend.iter_rows(min_row=1, max_row=10, values_only=True):
        if row[0] in ('TYPE', None):
            if row[0] == 'TYPE':
                months = [str(m) for m in row[1:] if m is not None]
            continue
        if row[0] and any(v is not None for v in row[1:]):
            spending[row[0]] = [v for v in row[1:] if v is not None]

    # NET_WORTH tab — extract SS benefits and balances for MC pre-fill
    nw = {
        "ss_monthly_62": 0.0, "ss_monthly_67": 0.0, "ss_monthly_70": 0.0,
        "checking_balance": 0.0, "sgov_balance": 0.0, "total_invested": 0.0,
        "monthly_burn": 0.0, "net_monthly_income": 0.0,
    }
    ws_nw = wb['NET_WORTH']
    for row in ws_nw.iter_rows(min_row=1, values_only=True):
        if not row or not any(v is not None for v in row):
            continue
        c0 = str(row[0] or '').strip()
        c1 = str(row[1] or '').strip() if len(row) > 1 else ''
        v1 = row[1] if len(row) > 1 and isinstance(row[1], (int, float)) else None
        v2 = row[2] if len(row) > 2 and isinstance(row[2], (int, float)) else None
        # VITALS rows: label in col0, numeric value in col1
        if 'SS Benefit @ 62' in c0 and v1 is not None: nw['ss_monthly_62'] = float(v1)
        elif 'SS Benefit @ 67' in c0 and v1 is not None: nw['ss_monthly_67'] = float(v1)
        elif 'SS Benefit @ 70' in c0 and v1 is not None: nw['ss_monthly_70'] = float(v1)
        elif 'Monthly Burn' in c0 and v1 is not None: nw['monthly_burn'] = float(v1)
        elif 'Net Monthly Income' in c0 and v1 is not None: nw['net_monthly_income'] = float(v1)
        # ASSET rows: account name in col1, balance in col2
        elif 'Checking' in c1 and 'Ops' in c1 and v2 is not None: nw['checking_balance'] = float(v2)
        elif 'SGOV' in c1 and v2 is not None: nw['sgov_balance'] = float(v2)
        elif 'TOTAL INVESTED' in c1 and v2 is not None: nw['total_invested'] = float(v2)

    wb.close()

    # Derived MC pre-fill values
    engine_bal = max(0.0, nw['total_invested'] - nw['sgov_balance'])
    mc_prefill = {
        "current_age":       None,   # not in ledger — user enters
        "engine_balance":    round(engine_bal),
        "sgov_balance":      round(nw['sgov_balance']),
        "checking_balance":  round(nw['checking_balance']),
        "full_ss_annual":    round(nw['ss_monthly_67'] * 12),
        "ss_monthly_67":     nw['ss_monthly_67'],
        "monthly_burn":      round(nw['monthly_burn']),
        "annual_floor_cost": round(nw['monthly_burn'] * 12),
    }

    return {
        "metrics": kv,
        "allocation": allocation,
        "cashflow": cashflow,
        "freedom_levels": freedom_levels,
        "spending_months": months,
        "spending": spending,
        "mc_prefill": mc_prefill,
    }

def read_roadmap_data():
    wb = openpyxl.load_workbook(LEDGER_PATH, read_only=True, data_only=True)
    ws = wb['ROADMAP']
    config = {}
    rows = []
    in_data = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not any(v is not None for v in row):
            continue
        c0 = str(row[0] or '').strip()
        if c0 == 'Year':
            in_data = True
            continue
        if not in_data:
            v = row[1] if len(row) > 1 else None
            if c0 and v is not None:
                config[c0.rstrip(':')] = v
        else:
            if row[0] is None:
                continue
            try:
                year = int(float(row[0]))
            except Exception:
                continue
            rows.append({
                'year':     year,
                'age':      int(row[1]) if row[1] is not None else None,
                'phase':    str(row[2] or ''),
                'sgov':     round(float(row[3] or 0), 2),
                'schwab':   round(float(row[4] or 0), 2),
                'roth':     round(float(row[5] or 0), 2),
                'liquid_nw': round(float(row[6] or 0), 2),
                'k401':     round(float(row[7] or 0), 2),
                'total_nw': round(float(row[8] or 0), 2),
            })
    wb.close()
    return {'config': config, 'rows': rows}

def read_transactions_data(page=1, limit=50, month_filter=None, type_filter=None):
    wb = openpyxl.load_workbook(LEDGER_PATH, read_only=True, data_only=True)
    ws = wb['TRANSACTIONS']
    all_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        date = row[1]
        if date is None:
            continue
        month = row[0]
        month_str = month.strftime('%Y-%m') if hasattr(month, 'strftime') else str(month or '')[:7]
        if month_filter and month_str != month_filter:
            continue
        txtype = str(row[2] or '').strip()
        if type_filter and txtype.lower() != type_filter.lower():
            continue
        all_rows.append({
            'month':    month_str,
            'date':     date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else str(date),
            'type':     txtype,
            'category': str(row[3] or '').strip(),
            'amount':   round(float(row[4] or 0), 2),
            'account':  str(row[5] or '').strip(),
            'memo':     str(row[6] or '').strip(),
            'signed':   round(float(row[7] or 0), 2),
        })
    wb.close()
    all_rows.sort(key=lambda r: r['date'], reverse=True)
    months = sorted({r['month'] for r in all_rows}, reverse=True)
    types  = sorted({r['type']  for r in all_rows if r['type']})
    total  = len(all_rows)
    start  = (page - 1) * limit
    return {
        'total': total, 'page': page, 'limit': limit,
        'pages': max(1, (total + limit - 1) // limit),
        'rows':  all_rows[start:start + limit],
        'months': months, 'types': types,
    }

def read_forecast_data():
    wb = openpyxl.load_workbook(LEDGER_PATH, read_only=True, data_only=True)
    ws = wb['FORECAST_V3']
    calib = None
    rows = []
    in_data = False
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if not any(v is not None for v in row):
            continue
        c0 = row[0]
        if i == 2:
            calib = {
                'current_date':      c0.strftime('%Y-%m-%d') if hasattr(c0, 'strftime') else str(c0),
                'checking_balance':  round(float(row[1] or 0), 2),
                'savings_balance':   round(float(row[2] or 0), 2),
                'projected_checking': round(float(row[5] or 0), 2) if len(row) > 5 else 0,
                'projected_savings':  round(float(row[6] or 0), 2) if len(row) > 6 else 0,
                'projected_total':    round(float(row[7] or 0), 2) if len(row) > 7 else 0,
            }
        if str(c0 or '') == 'DATE':
            in_data = True
            continue
        if in_data and hasattr(c0, 'strftime'):
            rows.append({
                'date':     c0.strftime('%Y-%m-%d'),
                'day':      str(row[1] or ''),
                'income':   round(float(row[2] or 0), 2),
                'expense':  round(float(row[3] or 0), 2),
                'invest':   round(float(row[4] or 0), 2),
                'checking': round(float(row[5] or 0), 2),
                'savings':  round(float(row[6] or 0), 2),
                'total':    round(float(row[7] or 0), 2),
            })
    wb.close()
    return {'calibration': calib, 'rows': rows}

def read_tax_loss_data():
    wb = openpyxl.load_workbook(LEDGER_PATH, read_only=True, data_only=True)
    ws = wb['TAX-LOSS']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row[:5]):
            continue
        date_val = row[0]
        date_str = date_val.strftime('%Y-%m-%d') if hasattr(date_val, 'strftime') else str(date_val or '')
        action   = str(row[1] or '').strip()
        amount   = round(float(row[2] or 0), 2)
        notes    = str(row[3] or '').strip()
        signed   = round(float(row[4] or 0), 2)
        rows.append({'date': date_str, 'action': action, 'amount': amount, 'notes': notes, 'signed': signed})
    wb.close()
    rows.sort(key=lambda r: r['date'], reverse=True)
    net_carryover = round(sum(r['signed'] for r in rows), 2)
    total_harvested = round(sum(r['signed'] for r in rows if r['signed'] < 0), 2)
    total_realized  = round(sum(r['signed'] for r in rows if r['signed'] > 0), 2)
    return {
        'rows': rows,
        'net_carryover':   net_carryover,
        'total_harvested': total_harvested,
        'total_realized':  total_realized,
        'entry_count':     len(rows),
    }

# ── ProcessPoolExecutor helpers (module-level so they pickle cleanly) ─────────

def _run_mc_worker(params):
    """Top-level wrapper so ProcessPoolExecutor can pickle run_monte_carlo."""
    return run_monte_carlo(params)

def _run_grid_sync(params_list):
    """Runs all grid combinations in parallel across CPU cores."""
    workers = min(os.cpu_count() or 4, len(params_list))
    with ProcessPoolExecutor(max_workers=workers) as ex:
        return list(ex.map(_run_mc_worker, params_list))

# ── Monte Carlo Engine ────────────────────────────────────────────────────────

RMD_TABLE = {
    75:22.9,76:22.0,77:21.1,78:20.2,79:19.4,80:18.7,81:17.9,82:17.1,
    83:16.3,84:15.5,85:14.8,86:14.1,87:13.4,88:12.7,89:12.0,90:11.4,
    91:10.8,92:10.2,93:9.6,94:9.1,95:8.6
}

def compute_ss_benefit(claimed_age, full_benefit):
    if claimed_age >= 70:
        return full_benefit * 1.24
    elif claimed_age > 67:
        return full_benefit * (1 + 0.08 * (claimed_age - 67))
    elif claimed_age == 67:
        return full_benefit
    else:
        months_early = (67 - claimed_age) * 12
        if months_early <= 36:
            reduction = months_early * (5/9) / 100
        else:
            reduction = 36*(5/9)/100 + (months_early-36)*(5/12)/100
        return full_benefit * (1 - reduction)

def compute_federal_tax(gross, year_off, infl, filing, reverted=False):
    if gross <= 0: return 0.0
    f = (1 + infl) ** year_off
    if reverted:
        std = (12700 if filing=='mfj' else 6350) * f
        if filing == 'mfj':
            bkts = [(18650*f,.10),(75900*f,.15),(153100*f,.25),(233350*f,.28),(416700*f,.33),(470700*f,.35),(math.inf,.396)]
        else:
            bkts = [(9325*f,.10),(37950*f,.15),(91900*f,.25),(191650*f,.28),(416700*f,.33),(418400*f,.35),(math.inf,.396)]
    else:
        std = (31500 if filing=='mfj' else 15750) * f
        if filing == 'mfj':
            bkts = [(23850*f,.10),(96950*f,.12),(206700*f,.22),(394600*f,.24),(501050*f,.32),(751600*f,.35),(math.inf,.37)]
        else:
            bkts = [(11925*f,.10),(48475*f,.12),(103350*f,.22),(197300*f,.24),(250525*f,.32),(626350*f,.35),(math.inf,.37)]
    taxable = max(0.0, gross - std)
    if taxable <= 0: return 0.0
    tax, prev = 0.0, 0.0
    for limit, rate in bkts:
        if taxable <= prev: break
        tax += (min(taxable, limit) - prev) * rate
        prev = limit
    return max(0.0, tax)

def taxable_ss_amount(ss_annual, other_income, filing):
    if ss_annual <= 0: return 0.0
    pi = other_income + ss_annual * 0.5
    lower = 32000 if filing == 'mfj' else 25000
    upper = 44000 if filing == 'mfj' else 34000
    if pi <= lower: return 0.0
    if pi <= upper: return min(0.5*ss_annual, 0.5*(pi-lower))
    return min(0.85*ss_annual, 0.5*(upper-lower) + 0.85*(pi-upper))

def compute_conversion_amount(trad, ss, rate, year_off, infl, filing):
    if trad <= 0: return 0.0
    f = (1 + infl) ** year_off
    std = (31500 if filing=='mfj' else 15750) * f
    if filing == 'mfj':
        bkts = [(23850*f,.10),(96950*f,.12),(206700*f,.22),(394600*f,.24),(501050*f,.32),(751600*f,.35)]
    else:
        bkts = [(11925*f,.10),(48475*f,.12),(103350*f,.22),(197300*f,.24),(250525*f,.32),(626350*f,.35)]
    top = bkts[-1][0]
    for lim, r in bkts:
        if r >= rate: top = lim; break
    target = std + top
    lo, hi = 0.0, min(trad, target)
    for _ in range(30):
        mid = (lo + hi) / 2
        if mid + taxable_ss_amount(ss, mid, filing) >= target: hi = mid
        else: lo = mid
    return min((lo+hi)/2, trad)

def mortality_mult(age):
    if age <= 70: return 1.0
    if age <= 80: return 1.0 - 0.03*(age-70)
    return max(0.15, 0.70 - (0.55/15)*(age-80))

def run_monte_carlo(params):
    t0 = time.time()
    g = lambda k, d: params.get(k, d)

    current_age  = int(g('current_age', 45))
    target_age   = int(g('target_age', 62))
    filing       = g('filing_status', 'single')
    end_age      = 95
    n_years      = end_age - current_age + 1

    engine0  = float(g('start_engine', 0))
    sgov0    = float(g('start_sgov', 0))
    chk0     = float(g('start_checking', 0))
    contrib  = float(g('annual_contribution', 0))
    wage_gr  = float(g('wage_growth', 0.02))

    moat_target = float(g('moat_target', 360000))     # dollars carved from engine into SGOV at retirement
    strict_moat = float(g('strict_moat_cost', 54292))  # annual floor draw from moat (already annual)
    full_moat   = moat_target                           # tent-skim accumulation target = carve target

    full_ss      = float(g('full_ss', 25620))
    ss_age_tgt   = int(g('ss_age', 67))
    use_haircut  = bool(g('use_ss_haircut', False))
    haircut_pct  = float(g('ss_haircut_pct', 0.21))

    mu     = float(g('mean_return', 0.09))
    sigma  = float(g('volatility', 0.16))
    syld   = float(g('sgov_yield', 0.04))
    divyld = float(g('dividend_yield', 0.015))

    infl         = float(g('inflation_rate', 0.03))
    use_si       = bool(g('use_stochastic_inflation', False))
    infl_vol     = float(g('inflation_volatility', 0.01))
    infl_min     = float(g('inflation_min', 0.01))
    infl_max     = float(g('inflation_max', 0.08))
    stag_corr    = float(g('stagflation_corr', 0.30))

    use_aca_shock  = bool(g('use_aca_shock', False))
    aca_shock_prob = float(g('aca_shock_prob', 0.30))
    aca_shock_mag  = float(g('aca_shock_mag', 15000))

    use_tax_rev    = bool(g('use_tax_reversion', False))
    tax_risk_near  = float(g('tax_risk_near', 0.20))
    tax_risk_mid   = float(g('tax_risk_mid', 0.40))
    tax_risk_late  = float(g('tax_risk_late', 0.60))

    gogo_e  = float(g('gogo_e', 0.25));  gogo_n  = float(g('gogo_n', 0.15))
    slgo_e  = float(g('slowgo_e', 0.20)); slgo_n = float(g('slowgo_n', 0.10))
    nogo_e  = float(g('nogo_e', 0.10));  nogo_n  = float(g('nogo_n', 0.05))
    euph_off = float(g('euphoric_offset', 0.03))
    euph_trig = mu + euph_off

    use_gf       = bool(g('use_gogo_floor', False))
    gogo_fl_ann  = float(g('gogo_floor_monthly', 1000)) * 12

    use_conv     = bool(g('use_conversion', False))
    trad0        = float(g('trad_balance', 0))
    ann_match    = float(g('annual_match', 0))
    tgt_bkt      = float(g('target_bracket', 0.12))
    cust_conv    = float(g('custom_conv_amt', 0))
    state_tx     = float(g('state_tax_rate', 0.0399))

    use_mc       = bool(g('use_medicare_surcharge', False))
    mc_ann       = float(g('medicare_monthly', 500)) * 12
    hc_infl      = float(g('healthcare_inflation_rate', 0.05))

    use_tail     = bool(g('use_tail_shock', False))
    tail_ret     = float(g('tail_shock_return', -0.25))
    tail_cnt     = int(g('tail_shock_count', 1))
    use_mr       = bool(g('use_mean_reversion', False))
    mr_str       = float(g('mean_reversion_strength', 0.15))

    gk_trig      = float(g('gk_trigger', 0.20))
    gk_cut       = float(g('gk_cut_rate', 0.50))
    bear_yrs     = int(g('bear_streak_years', 3))
    bear_cut     = float(g('bear_streak_cut', 0.25))

    p_cap        = float(g('portfolio_cap', 5_000_000))
    cap_infl     = float(g('cap_inflation', 0.03))
    cap_gg       = float(g('cap_gogo', 0.10))
    cap_sg       = float(g('cap_slowgo', 0.05))
    cap_ng       = float(g('cap_nogo', 0.02))

    use_rat      = bool(g('use_ratchet', False))
    rat_ann      = float(g('ratchet_boost_monthly', 1000)) * 12

    use_wf       = bool(g('use_wealth_floor', False))
    wf_rate      = float(g('wealth_floor_rate', 0.03))
    use_eb       = bool(g('use_euphoric_bonus', False))
    eb_rate      = float(g('euphoric_bonus_rate', 0.02))
    use_mort     = bool(g('use_mortality_weighting', True))
    use_res      = bool(g('use_residual_draw', False))
    res_ann      = float(g('residual_draw_monthly', 500)) * 12
    use_ph       = bool(g('use_prime_harvest', False))
    ph_yrs       = float(g('phase3_moat_years', 2))
    tent_rate    = float(g('tent_skim_rate', 0.50))

    SIZE_MAP = {"1k": 1_000, "10k": 10_000, "100k": 100_000, "1m": 1_000_000}
    sim_size = g('sim_size', None)
    if sim_size:
        n_trials = SIZE_MAP.get(str(sim_size).lower(), 1_000)
    else:
        n_trials = min(int(g('trials', 1000)), 1_000_000)

    return_model = str(g('return_model', 'normal')).lower()

    seq_yr   = int(g('seq_shock_year', 0))   # 0=off, 1/3/5 = years into retirement

    bridge_years  = max(1, ss_age_tgt - target_age)
    base_draw_ann = strict_moat  # already an annual figure

    rng = np.random.default_rng()

    if return_model == 'fat_tail':
        fat_df = int(g('fat_tail_df', 5))
        raw = rng.standard_t(fat_df, size=(n_trials, n_years))
        rets_all = raw * (sigma / np.sqrt(fat_df / (fat_df - 2))) + mu
    elif return_model == 'regime_switch':
        mu_bull  = float(g('mu_bull', 0.12));   sigma_bull = float(g('sigma_bull', 0.14))
        mu_bear  = float(g('mu_bear', -0.05));  sigma_bear = float(g('sigma_bear', 0.22))
        p_to_bear = float(g('p_bull_to_bear', 0.15))
        p_to_bull = float(g('p_bear_to_bull', 0.35))
        states = np.zeros((n_trials, n_years), dtype=np.int8)
        trans  = rng.random((n_trials, n_years))
        for t in range(1, n_years):
            flip_b = (states[:, t-1] == 0) & (trans[:, t] < p_to_bear)
            flip_u = (states[:, t-1] == 1) & (trans[:, t] < p_to_bull)
            states[:, t] = np.where(flip_b, 1, np.where(flip_u, 0, states[:, t-1]))
        bull_mask = (states == 0)
        rets_all = np.where(bull_mask,
            rng.normal(mu_bull, sigma_bull, (n_trials, n_years)),
            rng.normal(mu_bear, sigma_bear, (n_trials, n_years)))
    elif return_model == 'garch':
        g_omega = float(g('garch_omega', 0.0001))
        g_alpha = float(g('garch_alpha', 0.15))
        g_beta  = float(g('garch_beta',  0.80))
        var_t   = np.full(n_trials, sigma ** 2)
        rets_all = np.zeros((n_trials, n_years))
        z = rng.standard_normal((n_trials, n_years))
        for t in range(n_years):
            rets_all[:, t] = mu + np.sqrt(np.maximum(var_t, 1e-8)) * z[:, t]
            var_t = g_omega + g_alpha * rets_all[:, t] ** 2 + g_beta * var_t
    else:
        rets_all = rng.normal(mu, sigma, (n_trials, n_years))

    infl_all = rng.normal(infl, infl_vol, (n_trials, n_years)) if use_si else None

    all_paths      = np.zeros((n_trials, n_years))
    moat_paths     = np.zeros((n_trials, n_years))
    term_vals      = np.zeros(n_trials)
    arrival_arr    = np.zeros(n_trials)
    ss_age_arr     = np.zeros(n_trials)
    ripcord_arr    = np.zeros(n_trials, dtype=bool)
    breach_arr     = np.zeros(n_trials, dtype=bool)
    gogo_arr       = np.zeros(n_trials)
    slgo_arr       = np.zeros(n_trials)
    nogo_arr       = np.zeros(n_trials)
    conv_tx_arr    = np.zeros(n_trials)
    shadow_tx_arr  = np.zeros(n_trials)
    dd_arr         = np.zeros(n_trials)
    rat_t1_age_arr = np.zeros(n_trials)
    rat_t2_age_arr = np.zeros(n_trials)
    rat_t3_age_arr = np.zeros(n_trials)
    ph_peak_arr    = np.zeros(n_trials)
    ph_harv_arr    = np.zeros(n_trials)
    ph_drawn_arr   = np.zeros(n_trials)
    ph_refill_arr  = np.zeros(n_trials)
    ph_funded_arr  = np.zeros(n_trials)

    # ── Vectorized trial loop ─────────────────────────────────────────────────
    eng      = np.full(n_trials, engine0)
    sg       = np.full(n_trials, sgov0)
    chk      = np.full(n_trials, chk0)
    trad     = np.full(n_trials, trad0)
    sh_trad  = np.full(n_trials, trad0)
    ph3_moat = np.zeros(n_trials)
    infl_acc = np.ones(n_trials)
    cum_dev  = np.zeros(n_trials)
    cdn      = np.zeros(n_trials, dtype=np.int32)
    ath      = np.full(n_trials, engine0)
    eng_ret  = np.zeros(n_trials)
    br_moat  = np.zeros(n_trials)
    cl_ss    = np.full(n_trials, float(ss_age_tgt))
    ss_ben   = np.zeros(n_trials)
    ripcord  = np.zeros(n_trials, dtype=bool)
    breached = np.zeros(n_trials, dtype=bool)
    rat_tiers = np.zeros(n_trials, dtype=np.int32)
    rat_t1_fired = np.zeros(n_trials, dtype=bool)
    rat_t2_fired = np.zeros(n_trials, dtype=bool)
    rat_t3_fired = np.zeros(n_trials, dtype=bool)
    ph3_peak     = np.zeros(n_trials)
    ph_total_harv = np.zeros(n_trials)
    ph_total_drawn = np.zeros(n_trials)
    ph_refills   = np.zeros(n_trials)
    ph_was_drawn = np.zeros(n_trials, dtype=bool)
    ph_funded_age = np.zeros(n_trials)
    tot_gg = np.zeros(n_trials)
    tot_sg = np.zeros(n_trials)
    tot_ng = np.zeros(n_trials)
    tot_ctx = np.zeros(n_trials)
    tot_shx = np.zeros(n_trials)
    max_dd  = np.zeros(n_trials)
    yp      = 0

    # Pre-generate tail shock mask
    shock_mask = np.zeros((n_trials, n_years), dtype=bool)
    if use_tail:
        for ti in range(n_trials):
            cands = list(range(target_age, min(target_age + 10, end_age + 1)))
            chosen = random.sample(cands, min(tail_cnt, len(cands)))
            for age_c in chosen:
                shock_mask[ti, age_c - current_age] = True

    # Pre-generate tax reversion ages
    if use_tax_rev:
        rv = rng.random(n_trials)
        tax_rev_age_arr = np.where(
            rv < tax_risk_near,
            target_age + rng.integers(0, 5, n_trials),
            np.where(
                rv < tax_risk_near + tax_risk_mid,
                target_age + 5 + rng.integers(0, 5, n_trials),
                np.where(
                    rv < tax_risk_near + tax_risk_mid + tax_risk_late,
                    target_age + 10 + rng.integers(0, 10, n_trials),
                    np.full(n_trials, end_age + 1)
                )
            )
        )
    else:
        tax_rev_age_arr = np.full(n_trials, end_age + 1)
    tax_rev_mult = np.ones(n_trials)

    for ai in range(n_years):
        age = current_age + ai

        # Tax reversion
        tax_rev_mult = np.where(use_tax_rev & (age >= tax_rev_age_arr), 1.13, tax_rev_mult)

        # Returns
        raw = rets_all[:, ai].copy()
        if use_mr and yp > 0:
            raw += -mr_str * cum_dev
        # Sequence shock (forced, all trials)
        if seq_yr > 0 and age == target_age + seq_yr - 1:
            ret = np.full(n_trials, tail_ret)
        else:
            ret = np.where(shock_mask[:, ai], tail_ret, raw)
        cum_dev = (cum_dev + ret - mu) * 0.9
        cdn = np.where(ret < 0, cdn + 1, np.zeros(n_trials, dtype=np.int32))

        # Inflation
        if use_si:
            sh = np.where(ret < 0, np.abs(ret) * stag_corr, 0.0)
            ai_infl = np.clip(infl_all[:, ai] + sh, infl_min, infl_max)
        else:
            ai_infl = np.full(n_trials, infl)
        infl_acc *= (1 + ai_infl)

        # ── Phase 1: Accumulation ─────────────────────────────────────────────
        if age < target_age:
            c = contrib * ((1 + wage_gr) ** yp)
            # Tent skim in last 4 years before retirement
            tent_eligible = (age >= target_age - 4) & (ret > mu) & (sg < full_moat)
            excess = eng * ret - eng * mu
            sk = np.where(tent_eligible & (excess > 0),
                          np.minimum(excess * tent_rate, full_moat - sg), 0.0)
            eng = eng * (1 + ret) + c - sk
            sg  = sg + sk
            sg  = sg * (1 + syld)
            if use_conv:
                trad    = trad * (1 + ret) + ann_match * ((1 + wage_gr) ** yp)
                sh_trad = sh_trad * (1 + ret) + ann_match * ((1 + wage_gr) ** yp)
            all_paths[:, ai] = eng + sg + chk
            yp += 1
            continue

        # ── Transition: Enter retirement ──────────────────────────────────────
        entering = (age == target_age) & (eng_ret == 0.0)
        if np.any(entering):
            port = eng + sg + chk
            arrival_arr = np.where(entering, port, arrival_arr)
            dy = np.where(entering, np.minimum(port, moat_target), 0.0)
            new_eng = np.where(entering, np.maximum(0.0, port - dy - chk), eng)
            eng_ret = np.where(entering, new_eng, eng_ret)
            br_moat = np.where(entering, dy, br_moat)
            eng     = new_eng
            sg      = np.where(entering, 0.0, sg)
            runway  = np.where(base_draw_ann > 0, dy / base_draw_ann, float(bridge_years))
            rip     = runway < bridge_years
            ripcord = np.where(entering, rip, ripcord)
            new_cl  = np.where(rip,
                               np.round(np.minimum(70.0, np.maximum(62.0, target_age + runway))),
                               float(ss_age_tgt))
            cl_ss   = np.where(entering, new_cl, cl_ss)
            ripcord_arr = np.where(entering, rip, ripcord_arr)
            ss_age_arr  = np.where(entering, new_cl, ss_age_arr)
            raw_ss = np.array([compute_ss_benefit(float(a), full_ss) for a in cl_ss])
            raw_ss = np.where(entering, raw_ss, ss_ben)
            ss_ben = np.where(entering,
                              raw_ss * (1 - haircut_pct) if use_haircut else raw_ss,
                              ss_ben)
            ath = np.where(entering & (eng > ath), eng, ath)

        # ── Phase 2: Bridge ───────────────────────────────────────────────────
        in_bridge = age < cl_ss
        if np.any(in_bridge):
            eng_new = np.where(in_bridge, eng * (1 + ret), eng)
            ath = np.where(in_bridge & (eng_new > ath), eng_new, ath)
            draw = base_draw_ann * infl_acc
            if use_mc and age < 65:
                draw = draw + mc_ann * ((1 + hc_infl) ** (age - target_age))
            if use_aca_shock and target_age <= age < min(target_age + 3, int(np.min(cl_ss)) + 1):
                aca_hit = rng.random(n_trials) < aca_shock_prob
                draw = np.where(in_bridge & aca_hit, draw + aca_shock_mag * infl_acc, draw)
            br_moat_new = np.where(in_bridge, br_moat * (1 + syld) - draw, br_moat)
            overflow = br_moat_new < 0
            eng_new = np.where(in_bridge & overflow, np.maximum(0.0, eng_new + br_moat_new), eng_new)
            br_moat_new = np.where(in_bridge & overflow, 0.0, br_moat_new)
            breached = np.where(in_bridge & overflow, True, breached)
            breach_arr = np.where(in_bridge & overflow, True, breach_arr)
            if use_conv:
                conv_eligible = in_bridge & (age <= 74) & (trad > 0)
                if np.any(conv_eligible):
                    trad    = np.where(conv_eligible, trad * (1 + ret), trad)
                    sh_trad = np.where(conv_eligible, sh_trad * (1 + ret), sh_trad)
                    yo = age - current_age
                    cv_arr = np.zeros(n_trials)
                    for ti in np.where(conv_eligible)[0]:
                        if tgt_bkt == 'custom':
                            cv_arr[ti] = min(cust_conv, trad[ti])
                        else:
                            cv_arr[ti] = compute_conversion_amount(trad[ti], 0, tgt_bkt, yo, infl, filing)
                    has_conv = conv_eligible & (cv_arr > 0)
                    if np.any(has_conv):
                        fed_arr = np.array([compute_federal_tax(cv_arr[ti], yo, infl, filing)
                                            if has_conv[ti] else 0.0 for ti in range(n_trials)])
                        ttx = fed_arr * tax_rev_mult + cv_arr * state_tx
                        trad = np.where(has_conv, trad - cv_arr, trad)
                        eng_new = np.where(has_conv, eng_new + cv_arr, eng_new)
                        fc = np.minimum(chk, ttx)
                        chk = np.where(has_conv, chk - fc, chk)
                        eng_new = np.where(has_conv, np.maximum(0.0, eng_new - (ttx - fc)), eng_new)
                        tot_ctx = np.where(has_conv, tot_ctx + ttx, tot_ctx)
                    trad = np.where(conv_eligible & (trad < 1), 0.0, trad)
            dd = np.where(ath > 0, (ath - eng_new) / ath, 0.0)
            max_dd = np.maximum(max_dd, dd)
            all_paths[:, ai] = np.where(in_bridge, eng_new + br_moat_new, all_paths[:, ai])
            moat_paths[:, ai] = np.where(in_bridge, br_moat_new, moat_paths[:, ai])
            eng    = np.where(in_bridge, eng_new, eng)
            br_moat = np.where(in_bridge, br_moat_new, br_moat)
            yp += 1
            continue

        # ── Phase 3 entry (merge moat into engine) ────────────────────────────
        at_ss = (age == cl_ss.astype(int)) & (br_moat > 0)
        eng = np.where(at_ss, eng + br_moat, eng)
        br_moat = np.where(at_ss, 0.0, br_moat)

        # ── Phase 3: Smile ────────────────────────────────────────────────────
        s_eng = eng.copy()
        if use_rat and np.any(eng_ret > 0):
            ratio = np.where(eng_ret > 0, eng / eng_ret, 0.0)
            fire1 = ~rat_t1_fired & (rat_tiers == 0) & (ratio >= 1.5)
            fire2 = ~rat_t2_fired & (rat_tiers == 1) & (ratio >= 2.0)
            fire3 = ~rat_t3_fired & (rat_tiers == 2) & (ratio >= 2.5)
            rat_tiers = np.where(fire1, 1, np.where(fire2, 2, np.where(fire3, 3, rat_tiers)))
            rat_t1_fired = rat_t1_fired | fire1
            rat_t2_fired = rat_t2_fired | fire2
            rat_t3_fired = rat_t3_fired | fire3
            rat_t1_age_arr = np.where(fire1, age, rat_t1_age_arr)
            rat_t2_age_arr = np.where(fire2, age, rat_t2_age_arr)
            rat_t3_age_arr = np.where(fire3, age, rat_t3_age_arr)

        divs = s_eng * divyld
        eng  = eng * (1 + ret)
        ath  = np.maximum(ath, eng)
        mgain = eng - s_eng

        if use_ph:
            ph3_moat *= (1 + syld)
            euph = ret >= euph_trig
            tgt_ph = base_draw_ann * infl_acc * ph_yrs
            can_harv = euph & (mgain > 0) & (ph3_moat < tgt_ph)
            hv = np.where(can_harv, np.minimum(tgt_ph - ph3_moat, mgain), 0.0)
            eng -= hv; ph3_moat += hv; mgain -= hv
            ph_total_harv += hv
            newly_funded = can_harv & (ph3_moat >= tgt_ph)
            ph_funded_age = np.where(newly_funded & (ph_funded_age == 0), age, ph_funded_age)
            refilling = can_harv & ph_was_drawn
            ph_refills = np.where(refilling, ph_refills + 1, ph_refills)
            ph_was_drawn = np.where(refilling, False, ph_was_drawn)
            ph3_peak = np.maximum(ph3_peak, ph3_moat)

        ss_now    = ss_ben * infl_acc
        floor_now = base_draw_ann * infl_acc
        if use_mc and age < 65:
            floor_now = floor_now + mc_ann * ((1 + hc_infl) ** (age - target_age))
        gap = np.maximum(0.0, floor_now - ss_now)
        fd = np.minimum(divs, gap); divs -= fd; eng -= fd; gap -= fd
        if use_ph:
            ph_d = np.where(gap > 0, np.minimum(ph3_moat, gap), 0.0)
            ph3_moat -= ph_d; gap -= ph_d
            ph_total_drawn += ph_d
            ph_was_drawn = ph_was_drawn | (ph_d > 0)
        eng = np.maximum(0.0, eng - gap)

        if use_res and not ripcord.all():
            rgap = np.maximum(0.0, res_ann * infl_acc - ss_now)
            rgap = np.where(ripcord, 0.0, rgap)
            fd2 = np.minimum(divs, rgap); divs -= fd2; eng -= fd2; rgap -= fd2
            if use_ph:
                ph_d2 = np.where(rgap > 0, np.minimum(ph3_moat, rgap), 0.0)
                ph3_moat -= ph_d2; rgap -= ph_d2
                ph_total_drawn += ph_d2
                ph_was_drawn = ph_was_drawn | (ph_d2 > 0)
            eng = np.maximum(0.0, eng - rgap)

        # Spend rate by age band
        er = np.where(age <= 75, gogo_e, np.where(age <= 85, slgo_e, nogo_e))
        nr = np.where(age <= 75, gogo_n, np.where(age <= 85, slgo_n, nogo_n))
        if use_mort:
            mm = mortality_mult(age)
            er = er * mm; nr = nr * mm

        euph_vec = ret >= euph_trig
        skim = np.where(euph_vec & (mgain > 0), mgain * er,
               np.where((ret > 0) & (mgain > 0), mgain * nr, 0.0))
        skim = np.minimum(skim, eng)

        dd = np.where(ath > 0, (ath - eng) / ath, 0.0)
        max_dd = np.maximum(max_dd, dd)
        gk_active  = dd > gk_trig
        bear_active = cdn >= bear_yrs
        skim = np.where(gk_active & (skim > 0), skim * (1 - gk_cut), skim)
        skim = np.where(bear_active & (skim > 0), skim * (1 - bear_cut), skim)

        if use_rat:
            rf = rat_tiers * rat_ann * ((1 + infl) ** (age - target_age))
            skim = np.where((rat_tiers > 0) & (ret > 0), np.maximum(skim, np.minimum(rf, eng)), skim)
        if use_wf:
            wf = s_eng * wf_rate
            wf = np.where(gk_active, wf * (1 - gk_cut), wf)
            wf = np.where(bear_active, wf * (1 - bear_cut), wf)
            skim = np.where(ret > 0, np.maximum(skim, np.minimum(wf, eng)), skim)
        if use_eb:
            bn = s_eng * eb_rate
            bn = np.where(gk_active, bn * (1 - gk_cut), bn)
            bn = np.where(bear_active, bn * (1 - bear_cut), bn)
            bn = np.minimum(bn, np.maximum(0.0, eng - skim))
            skim = np.where(euph_vec, skim + bn, skim)

        eng -= skim
        tot_gg = np.where(age <= 75, tot_gg + skim, tot_gg)
        tot_sg = np.where((age > 75) & (age <= 85), tot_sg + skim, tot_sg)
        tot_ng = np.where(age > 85, tot_ng + skim, tot_ng)

        if use_gf and age <= 75:
            top = np.maximum(0.0, gogo_fl_ann * infl_acc - skim)
            a = np.minimum(top, np.maximum(0.0, eng))
            eng -= a; tot_gg += a

        cr = np.where(age <= 75, cap_gg, np.where(age <= 85, cap_sg, cap_ng))
        nc = p_cap * ((1 + cap_infl) ** (age - target_age))
        over_cap = eng > nc
        hc = np.where(over_cap, (eng - nc) * cr, 0.0)
        hc = np.where(gk_active, hc * (1 - gk_cut), hc)
        hc = np.where(bear_active, hc * (1 - bear_cut), hc)
        eng -= hc
        tot_gg = np.where(over_cap & (age <= 75), tot_gg + hc, tot_gg)
        tot_sg = np.where(over_cap & (age > 75) & (age <= 85), tot_sg + hc, tot_sg)
        tot_ng = np.where(over_cap & (age > 85), tot_ng + hc, tot_ng)

        yo = age - current_age
        if use_conv and age <= 74:
            conv_ok = trad > 0
            if np.any(conv_ok):
                trad    = np.where(conv_ok, trad * (1 + ret), trad)
                sh_trad = np.where(conv_ok, sh_trad * (1 + ret), sh_trad)
                cv_arr = np.zeros(n_trials)
                for ti in np.where(conv_ok)[0]:
                    if tgt_bkt == 'custom':
                        cv_arr[ti] = min(cust_conv, trad[ti])
                    else:
                        cv_arr[ti] = compute_conversion_amount(trad[ti], ss_ben[ti]*infl_acc[ti], tgt_bkt, yo, infl, filing)
                has_cv = conv_ok & (cv_arr > 0)
                if np.any(has_cv):
                    for ti in np.where(has_cv)[0]:
                        ss_n = ss_ben[ti] * infl_acc[ti]
                        twc = taxable_ss_amount(ss_n, cv_arr[ti], filing)
                        tnc = taxable_ss_amount(ss_n, 0, filing)
                        fed = max(0.0,
                            compute_federal_tax(cv_arr[ti]+twc, yo, infl, filing) -
                            compute_federal_tax(tnc, yo, infl, filing))
                        ttx = fed * tax_rev_mult[ti] + cv_arr[ti] * state_tx
                        trad[ti] -= cv_arr[ti]; eng[ti] += cv_arr[ti]
                        fc = min(chk[ti], ttx); chk[ti] -= fc
                        eng[ti] = max(0.0, eng[ti] - (ttx - fc))
                        tot_ctx[ti] += ttx
                trad = np.where(conv_ok & (trad < 1), 0.0, trad)

        if use_conv and age >= 75:
            rmd_ok = trad > 0
            if np.any(rmd_ok):
                trad = np.where(rmd_ok, trad * (1 + ret), trad)
                rf_val = RMD_TABLE.get(min(age, 95), 8.6)
                rmd = np.where(rmd_ok, trad / rf_val, 0.0)
                for ti in np.where(rmd_ok)[0]:
                    ss_n = ss_ben[ti] * infl_acc[ti]
                    tss = taxable_ss_amount(ss_n, rmd[ti], filing)
                    fed = max(0.0,
                        compute_federal_tax(rmd[ti]+tss, yo, infl, filing) -
                        compute_federal_tax(tss, yo, infl, filing))
                    st = rmd[ti] * state_tx
                    eng[ti] += max(0.0, rmd[ti] - fed - st)
                    trad[ti] -= rmd[ti]
                    tot_ctx[ti] += fed + st
                    if sh_trad[ti] > 0:
                        sh_trad[ti] *= (1 + ret[ti])
                        sr = sh_trad[ti] / rf_val
                        sts = taxable_ss_amount(ss_n, sr, filing)
                        sf = max(0.0,
                            compute_federal_tax(sr+sts, yo, infl, filing) -
                            compute_federal_tax(sts, yo, infl, filing))
                        tot_shx[ti] += sf + sr * state_tx
                        sh_trad[ti] -= sr
                        if sh_trad[ti] < 0: sh_trad[ti] = 0.0
                trad = np.where(rmd_ok & (trad < 0), 0.0, trad)

        all_paths[:, ai] = np.maximum(0.0, eng + ph3_moat)
        yp += 1

    term_vals      = np.maximum(0.0, eng)
    gogo_arr       = tot_gg
    slgo_arr       = tot_sg
    nogo_arr       = tot_ng
    conv_tx_arr    = tot_ctx
    shadow_tx_arr  = tot_shx
    dd_arr         = max_dd
    if use_ph:
        ph_peak_arr   = ph3_peak
        ph_harv_arr   = ph_total_harv
        ph_drawn_arr  = ph_total_drawn
        ph_refill_arr = ph_refills
        ph_funded_arr = ph_funded_age

    # ── Aggregate ─────────────────────────────────────────────────────────────
    ages = list(range(current_age, end_age + 1))
    pcts = [10, 25, 50, 75, 90]
    bands = {"ages": ages}
    for p in pcts:
        bands[f"p{p}"] = [float(np.percentile(all_paths[:, i], p)) for i in range(n_years)]

    # Moat balance bands — only for bridge years where moat is active
    moat_ages = list(range(target_age, min(target_age + bridge_years + 1, end_age + 1)))
    moat_bands = {"ages": moat_ages}
    for p in [10, 50, 90]:
        moat_bands[f"p{p}"] = [
            float(np.percentile(moat_paths[:, a - current_age], p))
            for a in moat_ages
        ]

    milestones = sorted({target_age, min(target_age+5,70), 70, 75, 80, 85, 90, 95} &
                        set(range(current_age, end_age+1)))
    mile_out = []
    for ma in milestones:
        idx = ma - current_age
        mile_out.append({"age": ma,
            "p10": float(np.percentile(all_paths[:,idx], 10)),
            "p50": float(np.percentile(all_paths[:,idx], 50)),
            "p90": float(np.percentile(all_paths[:,idx], 90))})

    ruin_by_age = {}
    for ra in [70, 75, 80, 85, 90]:
        if ra > current_age:
            idx = min(ra - current_age, n_years - 1)
            ruin_by_age[str(ra)] = round(float(np.mean(all_paths[:, idx] <= 0)) * 100, 1)

    sav = shadow_tx_arr - conv_tx_arr

    ss_hist = {}
    for a in range(62, 71):
        ss_hist[str(a)] = int(np.sum(ss_age_arr == a))

    total_spend_arr = gogo_arr + slgo_arr + nogo_arr
    spt = total_spend_arr + term_vals
    med_spt = float(np.median(spt))
    lifetime_spend = {
        "p50_total":  round(float(np.median(total_spend_arr))),
        "p90_total":  round(float(np.percentile(total_spend_arr, 90))),
        "p50_gogo":   round(float(np.median(gogo_arr))),
        "p50_slowgo": round(float(np.median(slgo_arr))),
        "p50_nogo":   round(float(np.median(nogo_arr))),
        "spend_ratio": round(float(np.median(total_spend_arr)) / med_spt * 100, 1) if med_spt > 0 else 0.0,
    }

    ratchet_stats = None
    if use_rat:
        t1 = rat_t1_age_arr[rat_t1_age_arr > 0]
        t2 = rat_t2_age_arr[rat_t2_age_arr > 0]
        t3 = rat_t3_age_arr[rat_t3_age_arr > 0]
        ratchet_stats = {
            "tier1_pct":        round(float(len(t1)/n_trials*100), 1),
            "tier2_pct":        round(float(len(t2)/n_trials*100), 1),
            "tier3_pct":        round(float(len(t3)/n_trials*100), 1),
            "median_tier1_age": round(float(np.median(t1))) if len(t1) else None,
            "median_tier2_age": round(float(np.median(t2))) if len(t2) else None,
            "median_tier3_age": round(float(np.median(t3))) if len(t3) else None,
        }

    prime_harvest_stats = None
    if use_ph:
        funded = ph_funded_arr[ph_funded_arr > 0]
        prime_harvest_stats = {
            "median_peak":       round(float(np.median(ph_peak_arr))),
            "funded_pct":        round(float(len(funded)/n_trials*100), 1),
            "median_funded_age": round(float(np.median(funded))) if len(funded) else None,
            "median_drawn":      round(float(np.median(ph_drawn_arr))),
            "median_refills":    round(float(np.median(ph_refill_arr)), 1),
            "recycled_pct":      round(float(np.mean(ph_refill_arr >= 1)*100), 1),
        }

    ratchet_paths = None
    if use_rat:
        ph3_ages = list(range(target_age, end_age + 1))
        t1_cum, t2_cum, t3_cum = [], [], []
        for a in ph3_ages:
            t1_cum.append(round(float(np.sum((rat_t1_age_arr > 0) & (rat_t1_age_arr <= a)) / n_trials * 100), 1))
            t2_cum.append(round(float(np.sum((rat_t2_age_arr > 0) & (rat_t2_age_arr <= a)) / n_trials * 100), 1))
            t3_cum.append(round(float(np.sum((rat_t3_age_arr > 0) & (rat_t3_age_arr <= a)) / n_trials * 100), 1))
        ratchet_paths = {"ages": ph3_ages, "t1": t1_cum, "t2": t2_cum, "t3": t3_cum}

    spend_scenarios = {
        "labels": ["Go-Go (62\u201375)", "Slow-Go (76\u201385)", "No-Go (86+)"],
        "p10": [round(float(np.percentile(a, 10))) for a in [gogo_arr, slgo_arr, nogo_arr]],
        "p25": [round(float(np.percentile(a, 25))) for a in [gogo_arr, slgo_arr, nogo_arr]],
        "p50": [round(float(np.percentile(a, 50))) for a in [gogo_arr, slgo_arr, nogo_arr]],
        "p75": [round(float(np.percentile(a, 75))) for a in [gogo_arr, slgo_arr, nogo_arr]],
        "p90": [round(float(np.percentile(a, 90))) for a in [gogo_arr, slgo_arr, nogo_arr]],
    }

    return {
        "success_pct":        round(float(np.mean(term_vals > 0)*100), 1),
        "milestones":         mile_out,
        "bands":              bands,
        "moat_bands":         moat_bands,
        "ruin_by_age":        ruin_by_age,
        "ss_histogram":       ss_hist,
        "lifetime_spend":     lifetime_spend,
        "ratchet_stats":      ratchet_stats,
        "ratchet_paths":      ratchet_paths,
        "spend_scenarios":    spend_scenarios,
        "prime_harvest_stats": prime_harvest_stats,
        "stats": {
            "median_arrival":      round(float(np.median(arrival_arr))),
            "median_ss_age":       round(float(np.median(ss_age_arr)), 1),
            "ripcord_rate":        round(float(np.mean(ripcord_arr)*100), 1),
            "moat_breach_rate":    round(float(np.mean(breach_arr)*100), 1),
            "median_terminal":     round(float(np.median(term_vals))),
            "median_gogo_spend":   round(float(np.median(gogo_arr))),
            "median_total_spend":  round(float(np.median(total_spend_arr))),
            "conv_tax_paid":       round(float(np.median(conv_tx_arr))),
            "tax_savings":         round(float(np.median(sav))),
            "median_drawdown":     round(float(np.median(dd_arr)*100), 1),
        },
        "trial_count": n_trials,
        "runtime_ms":  round((time.time()-t0)*1000, 1),
    }

@mcp.tool()
def run_retirement_simulation(
    current_age: int = 45,
    retirement_age: int = 62,
    engine_balance: float = 0,
    sgov_balance: float = 0,
    checking_balance: float = 5000,
    ss_benefit_67: float = 25620,
    annual_contribution: float = 0
) -> str:
    """Runs a Monte Carlo retirement simulation (1,000 trials) using V4-equivalent parameters.
    Returns success rate, wealth percentiles at key ages, and key simulation statistics."""
    params = {
        "current_age": current_age,
        "target_age": retirement_age,
        "start_engine": engine_balance,
        "start_sgov": sgov_balance,
        "start_checking": checking_balance,
        "full_ss": ss_benefit_67,
        "annual_contribution": annual_contribution,
        "trials": 1000,
        "mean_return": 0.09, "volatility": 0.16, "sgov_yield": 0.04,
        "inflation_rate": 0.03, "strict_moat_cost": 96000, "full_moat_cost": 96000,
        "use_mortality_weighting": True,
        "gogo_e": 0.25, "gogo_n": 0.15, "slowgo_e": 0.20, "slowgo_n": 0.10,
        "nogo_e": 0.10, "nogo_n": 0.05, "gk_trigger": 0.20, "gk_cut_rate": 0.50,
        "bear_streak_years": 3, "bear_streak_cut": 0.25,
        "portfolio_cap": 5000000, "cap_inflation": 0.03,
        "cap_gogo": 0.10, "cap_slowgo": 0.05, "cap_nogo": 0.02,
        "tent_skim_rate": 0.50, "dividend_yield": 0.015, "wage_growth": 0.02,
    }
    result = run_monte_carlo(params)
    s = result["stats"]
    lines = [
        f"SUCCESS RATE: {result['success_pct']}%  ({result['trial_count']:,} trials, {result['runtime_ms']}ms)",
        "",
        "WEALTH PERCENTILES:",
    ]
    for m in result["milestones"]:
        lines.append(f"  Age {m['age']:2d}:  P10=${m['p10']:>12,.0f}   P50=${m['p50']:>12,.0f}   P90=${m['p90']:>12,.0f}")
    lines += [
        "",
        "KEY STATISTICS:",
        f"  Arrival wealth (median):  ${s['median_arrival']:>12,.0f}",
        f"  SS claim age (median):    {s['median_ss_age']}",
        f"  Ripcord rate (early SS):  {s['ripcord_rate']}%",
        f"  Moat breach rate:         {s['moat_breach_rate']}%",
        f"  Terminal wealth (median): ${s['median_terminal']:>12,.0f}",
        f"  Go-go spend (median):     ${s['median_gogo_spend']:>12,.0f}",
        f"  Max drawdown (median):    {s['median_drawdown']}%",
    ]
    return "\n".join(lines)

# ── REST API handlers ──────────────────────────────────────────────────────────

async def api_rules(request: Request):
    return JSONResponse({
        "contrib_401k": 24500,
        "catchup_50_plus": 8000,
        "super_catchup_60_63": 11250,
        "ira_roth_limit": 7500,
        "ira_roth_50_plus": 8600,
        "rothification_income_threshold": 145000,
        "year": 2026,
        "return_models": ["normal", "fat_tail", "regime_switch", "garch"],
        "sim_sizes": ["1k", "10k", "100k", "1m"],
    })

async def api_ledger_dashboard(request: Request):
    try:
        data = await asyncio.to_thread(read_dashboard_data)
        return JSONResponse(data)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def api_monte_carlo(request: Request):
    try:
        body = await request.json()
        result = run_monte_carlo(body)
        return JSONResponse(result)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def api_portfolio(request: Request):
    try:
        holdings = await asyncio.to_thread(read_portfolio_data)
        for h in holdings:
            if h['cached_price'] and h['shares']:
                h['cached_value'] = round(h['cached_price'] * h['shares'], 2)
            else:
                h['cached_value'] = None
        return JSONResponse({"holdings": holdings, "as_of": "cached"})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def api_portfolio_refresh(request: Request):
    try:
        body = await request.json()
        api_key = body.get('api_key', '').strip()
        if not api_key:
            return JSONResponse({"error": "API key required"}, status_code=400)

        def _refresh_logic():
            holdings = read_portfolio_data()
            for h in holdings:
                if h.get('is_proxy'):
                    # 401k proxies: use cached Excel value, don't fetch live price
                    h['live_price'] = h['cached_price']
                    h['live_value'] = h.get('cached_value')
                    h['gain_loss'] = None;  h['gain_loss_pct'] = None
                    h['proxy_note'] = 'Voya proxy — cached only'
                    continue
                try:
                    price = fetch_av_price(h['ticker'], api_key, h['is_crypto'])
                    h['live_price'] = price
                    h['live_value'] = round(price * h['shares'], 2) if price else None
                    cost_basis = (h['avg_cost'] or 0) * h['shares']
                    if price and h['avg_cost'] and cost_basis > 0:
                        h['gain_loss']     = round(h['live_value'] - cost_basis, 2)
                        h['gain_loss_pct'] = round(h['gain_loss'] / cost_basis * 100, 2)
                    else:
                        h['gain_loss'] = None;  h['gain_loss_pct'] = None
                except Exception as ex:
                    h['live_price'] = None;  h['error'] = str(ex)
                time.sleep(0.3)

            total_value     = sum(h.get('live_value') or 0 for h in holdings)
            total_cost      = sum((h.get('avg_cost') or 0) * h['shares'] for h in holdings)
            total_gain_loss = total_value - total_cost

            return {
                "holdings": holdings,
                "summary": {
                    "total_value":        round(total_value, 2),
                    "total_cost":         round(total_cost, 2),
                    "total_gain_loss":    round(total_gain_loss, 2),
                    "total_gain_loss_pct": round(total_gain_loss / total_cost * 100, 2) if total_cost else 0,
                },
                "as_of": "live",
            }

        result_data = await asyncio.to_thread(_refresh_logic)
        return JSONResponse(result_data)

    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def api_stock_price(request: Request):
    try:
        body = await request.json()
        ticker = body.get("ticker", "").strip().upper()
        api_key = body.get("api_key", "").strip()
        if not ticker or not api_key:
            return JSONResponse({"error": "Ticker and API key are required."}, status_code=400)
        url = f'https://www.alphavantage.co/query?function=GLOBAL_QUOTE&symbol={ticker}&apikey={api_key}'
        r = requests.get(url, timeout=10)
        data = r.json()
        quote = data.get("Global Quote", {})
        price = quote.get("05. price")
        change_pct = quote.get("10. change percent", "").replace("%", "")
        if not price:
            note = data.get("Note") or data.get("Information")
            return JSONResponse({"error": note or f"No data found for {ticker}."})
        return JSONResponse({
            "ticker": ticker,
            "price": float(price),
            "change_pct": float(change_pct) if change_pct else None
        })
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def api_send_digest(request: Request):
    try:
        body = await request.json()
        host = body.get('smtp_host', 'smtp.gmail.com')
        port = int(body.get('smtp_port', 587))
        user = body.get('smtp_user', '')
        pwd  = body.get('smtp_pass', '')
        to   = body.get('to_email', user)
        if not user or not pwd:
            return JSONResponse({'error': 'smtp_user and smtp_pass required'}, status_code=400)
        data = await asyncio.to_thread(read_dashboard_data)
        m    = data['metrics']
        lnw  = m.get('LIQUID NET WORTH', 0) or 0
        tnw  = m.get('TOTAL NET WORTH', 0) or 0
        pct  = float(m.get('PROGRESS TO FI', 0) or 0) * 100
        fl_levels = data.get('freedom_levels', [])
        fl_rows = ''.join(
            f'<tr><td>{fl["name"]}</td><td align="right">'
            f'{"&#10003;" if fl.get("status")=="Achieved" else fl.get("status","—")}'
            f'</td></tr>'
            for fl in fl_levels if fl.get('name')
        )
        html_body = (
            '<html><body style="font-family:sans-serif;max-width:520px;margin:auto;padding:20px;">'
            '<div style="background:#080d1a;color:#f0f4ff;padding:24px;border-radius:12px;">'
            '<h2 style="color:#ec4899;margin:0 0 16px;">&#x1F4B0; Road To FI Digest</h2>'
            '<table width="100%" cellpadding="8" style="font-size:14px;border-collapse:collapse;">'
            f'<tr><td>Liquid Net Worth</td><td align="right"><b>${lnw:,.0f}</b></td></tr>'
            f'<tr><td>Total Net Worth</td><td align="right"><b>${tnw:,.0f}</b></td></tr>'
            f'<tr><td>Progress to FI</td><td align="right"><b>{pct:.1f}%</b></td></tr>'
            '</table>'
            + (f'<hr style="border-color:#1a2540;margin:12px 0;"><table width="100%" cellpadding="6" style="font-size:13px;">{fl_rows}</table>' if fl_rows else '')
            + '</div>'
            '<p style="color:#666;font-size:11px;text-align:center;margin-top:12px;">'
            'Road To FI &mdash; automated digest</p>'
            '</body></html>'
        )
        msg = MIMEMultipart('alternative')
        msg['Subject'] = body.get('subject', 'Road To FI Weekly Digest')
        msg['From']    = user
        msg['To']      = to
        msg.attach(MIMEText(html_body, 'html'))
        with smtplib.SMTP(host, port) as s:
            s.ehlo()
            s.starttls()
            s.login(user, pwd)
            s.sendmail(user, to, msg.as_string())
        return JSONResponse({'ok': True, 'sent_to': to})
    except Exception as e:
        return JSONResponse({'error': str(e)}, status_code=500)

async def api_optimize_contribution(request: Request):
    try:
        body = await request.json()
        target_pct = float(body.pop('target_success_pct', 95))
        lo, hi = 0.0, 150000.0
        last_result = None
        for _ in range(18):
            mid = (lo + hi) / 2
            body['annual_contribution'] = mid
            last_result = run_monte_carlo(body)
            if last_result['success_pct'] >= target_pct:
                hi = mid
            else:
                lo = mid
        return JSONResponse({
            'optimal_contribution': round(hi),
            'achieved_success_pct': round(last_result['success_pct'], 1) if last_result else 0
        })
    except Exception as e:
        return JSONResponse({'error': str(e)}, status_code=500)

async def api_sensitivity(request: Request):
    try:
        base = await request.json()
        base.setdefault('trials', 500)
        base_pct = run_monte_carlo(base)['success_pct']
        moat_val  = base.get('moat_target', 360000)
        floor_val = base.get('strict_moat_cost', 54292)
        tests = [
            ('mean_return',         0.02,  'Return +2%',      'Return -2%'),
            ('volatility',          0.05,  'Vol -5%',         'Vol +5%'),
            ('inflation_rate',      0.01,  'Inflation -1%',   'Inflation +1%'),
            ('annual_contribution', 10000, 'Contrib +$10k',   'Contrib -$10k'),
            ('moat_target',         moat_val * 0.10, 'Moat +10%', 'Moat -10%'),
        ]
        results = []
        for key, delta, pos_label, neg_label in tests:
            base_val = base.get(key, 0)
            pos_p = {**base, key: base_val + delta}
            neg_p = {**base, key: max(0, base_val - delta)}
            pos_delta = round(run_monte_carlo(pos_p)['success_pct'] - base_pct, 1)
            neg_delta = round(run_monte_carlo(neg_p)['success_pct'] - base_pct, 1)
            results.append({
                'param': key, 'pos_label': pos_label, 'neg_label': neg_label,
                'pos_delta': pos_delta, 'neg_delta': neg_delta
            })
        results.sort(key=lambda x: max(abs(x['pos_delta']), abs(x['neg_delta'])), reverse=True)
        return JSONResponse({'base_pct': base_pct, 'results': results})
    except Exception as e:
        return JSONResponse({'error': str(e)}, status_code=500)

async def api_roadmap(request: Request):
    try:
        data = await asyncio.to_thread(read_roadmap_data)
        return JSONResponse(data)
    except Exception as e:
        return JSONResponse({'error': str(e)}, status_code=500)

async def api_transactions(request: Request):
    try:
        qs = request.query_params
        page   = int(qs.get('page', 1))
        limit  = min(int(qs.get('limit', 50)), 200)
        month  = qs.get('month') or None
        txtype = qs.get('type') or None
        data = await asyncio.to_thread(read_transactions_data, page, limit, month, txtype)
        return JSONResponse(data)
    except Exception as e:
        return JSONResponse({'error': str(e)}, status_code=500)

async def api_forecast(request: Request):
    try:
        data = await asyncio.to_thread(read_forecast_data)
        return JSONResponse(data)
    except Exception as e:
        return JSONResponse({'error': str(e)}, status_code=500)

async def api_tax_loss(request: Request):
    try:
        data = await asyncio.to_thread(read_tax_loss_data)
        return JSONResponse(data)
    except Exception as e:
        return JSONResponse({'error': str(e)}, status_code=500)

async def api_grid_search(request: Request):
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({'error': 'invalid JSON'}, status_code=400)
    base_params = dict(body.get('base_params', {}))
    grid_axes   = body.get('grid_axes', {})
    if not grid_axes:
        return JSONResponse({'error': 'grid_axes required'}, status_code=400)
    axis_names  = list(grid_axes.keys())
    axis_values = [grid_axes[k] for k in axis_names]
    combos      = list(itertools.product(*axis_values))
    MAX_COMBOS  = 300
    if len(combos) > MAX_COMBOS:
        return JSONResponse({'error': f'Too many combinations ({len(combos)}). Max {MAX_COMBOS}.'}, status_code=400)
    params_list = []
    for combo in combos:
        p = {**base_params}
        for name, val in zip(axis_names, combo):
            p[name] = val
        p.setdefault('trials', 500)
        p['trials'] = min(int(p['trials']), 500)
        params_list.append(p)
    t0 = time.time()
    try:
        results_raw = await asyncio.to_thread(_run_grid_sync, params_list)
    except Exception as e:
        return JSONResponse({'error': str(e)}, status_code=500)
    output = []
    for combo, res, p in zip(combos, results_raw, params_list):
        row = {name: round(float(val), 6) for name, val in zip(axis_names, combo)}
        row['success_pct']      = res['success_pct']
        row['median_terminal']  = res['stats']['median_terminal']
        row['ripcord_rate']     = res['stats']['ripcord_rate']
        row['moat_breach_rate'] = res['stats']['moat_breach_rate']
        output.append(row)
    output.sort(key=lambda r: r['success_pct'], reverse=True)
    return JSONResponse({
        'results':     output,
        'axis_names':  axis_names,
        'combo_count': len(combos),
        'runtime_ms':  round((time.time()-t0)*1000, 1),
    })

def _build_context_string(sim_data, dashboard_data):
    lines = []
    if sim_data:
        s = sim_data.get('stats', {})
        lines += [
            f"MONTE CARLO SIMULATION ({sim_data.get('trial_count',0):,} trials):",
            f"  Success rate (alive at 95): {sim_data.get('success_pct')}%",
            f"  Arrival wealth at retirement (median): ${s.get('median_arrival',0):,.0f}",
            f"  SS claim age (median): {s.get('median_ss_age')}",
            f"  Early SS (ripcord) rate: {s.get('ripcord_rate')}%",
            f"  SGOV moat breach rate: {s.get('moat_breach_rate')}%",
            f"  Terminal wealth at 95 (median): ${s.get('median_terminal',0):,.0f}",
            f"  Go-Go discretionary spend (median): ${s.get('median_gogo_spend',0):,.0f}",
            f"  Max drawdown (median): {s.get('median_drawdown')}%",
        ]
        if s.get('conv_tax_paid', 0) > 0:
            lines.append(f"  Roth conversion tax (median): ${s['conv_tax_paid']:,.0f}")
            lines.append(f"  Roth conversion tax savings (median): ${max(0, s.get('tax_savings',0)):,.0f}")
        ruin = sim_data.get('ruin_by_age', {})
        if ruin:
            lines.append("  Ruin probability by age: " + ", ".join(f"age {a}: {v}%" for a, v in ruin.items()))
        ms = sim_data.get('milestones', [])
        if ms:
            lines.append("  Wealth percentiles (P10 / P50 / P90):")
            for m in ms:
                lines.append(f"    Age {m['age']}: ${m['p10']:,.0f} / ${m['p50']:,.0f} / ${m['p90']:,.0f}")
        rs = sim_data.get('ratchet_stats')
        if rs:
            lines.append(f"  Abundance ratchet Tier 1 (150%): {rs.get('tier1_pct')}% of trials, median age {rs.get('median_tier1_age')}")
        ls = sim_data.get('lifetime_spend')
        if ls:
            lines.append(f"  Lifetime spend P50: ${ls.get('p50_total',0):,.0f} (go-go ${ls.get('p50_gogo',0):,.0f} / slow-go ${ls.get('p50_slowgo',0):,.0f} / no-go ${ls.get('p50_nogo',0):,.0f})")
    if dashboard_data:
        nw = dashboard_data.get('net_worth', {})
        sp = dashboard_data.get('spending', {})
        mc = dashboard_data.get('mc_prefill', {})
        if nw:
            lines += [
                "\nCURRENT FINANCIAL SNAPSHOT:",
                f"  Liquid net worth: ${nw.get('liquid_nw',0):,.0f}",
                f"  Total net worth: ${nw.get('total_nw',0):,.0f}",
            ]
        if mc:
            if mc.get('engine_balance'): lines.append(f"  VTI/brokerage balance: ${mc['engine_balance']:,.0f}")
            if mc.get('sgov_balance'):   lines.append(f"  SGOV balance: ${mc['sgov_balance']:,.0f}")
            if mc.get('full_ss_annual'): lines.append(f"  Projected SS benefit: ${mc['full_ss_annual']:,.0f}/yr")
        if sp:
            lines.append(f"  Monthly spending: ${sp.get('monthly_total',0):,.0f}")
            if sp.get('savings_rate'): lines.append(f"  Savings rate: {sp['savings_rate']:.1f}%")
    # Freedom levels
    if dashboard_data and dashboard_data.get('freedom_levels'):
        lines.append("\nFREEDOM LEVELS:")
        for lv in dashboard_data['freedom_levels']:
            prog = lv.get('progress')
            status = lv.get('status') or ('✓ Achieved' if (prog and prog >= 1) else '… Pending')
            goal = f"  (goal: ${lv['goal']:,.0f})" if isinstance(lv.get('goal'), (int, float)) else ''
            lines.append(f"  {lv['name']}{goal}: {status}")

    # Asset allocation
    if dashboard_data and dashboard_data.get('allocation'):
        lines.append("\nASSET ALLOCATION:")
        for k, v in dashboard_data['allocation'].items():
            lines.append(f"  {k}: {v:.1f}%")

    # Spending by category — most recent month
    if dashboard_data and dashboard_data.get('spending') and dashboard_data.get('spending_months'):
        months = dashboard_data['spending_months']
        lines.append(f"\nSPENDING (most recent: {months[0] if months else 'N/A'}):")
        for cat, vals in dashboard_data['spending'].items():
            if vals and isinstance(vals[0], (int, float)):
                lines.append(f"  {cat}: ${vals[0]:,.0f}")

    # Additional balance details
    if dashboard_data:
        mc = dashboard_data.get('mc_prefill', {})
        extras = []
        if mc.get('checking_balance'): extras.append(f"  Checking balance: ${mc['checking_balance']:,.0f}")
        if mc.get('monthly_burn'):     extras.append(f"  Monthly burn: ${mc['monthly_burn']:,.0f}")
        if mc.get('annual_floor_cost'): extras.append(f"  Annual floor cost: ${mc['annual_floor_cost']:,.0f}")
        if extras:
            lines.append("\nADDITIONAL BALANCES:")
            lines.extend(extras)

    return '\n'.join(lines) if lines else 'No financial data available.'

async def api_chat_stream(request: Request):
    import json as _json
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({'error': 'invalid JSON'}, status_code=400)
    message        = str(body.get('message', '')).strip()
    context_type   = body.get('context_type', 'all')
    model          = body.get('model', 'llama3.1:8b')
    history        = body.get('history', [])
    sim_data       = body.get('sim_data') if context_type in ('all', 'simulation') else None
    dashboard_data = None
    if context_type in ('all', 'dashboard'):
        try:
            dashboard_data = await asyncio.to_thread(read_dashboard_data)
        except Exception:
            pass
    ctx = _build_context_string(sim_data, dashboard_data)
    system_prompt = SYSTEM_PROMPT.format(today=time.strftime('%Y-%m-%d'))
    system_prompt = system_prompt + f"\n\nLIVE FINANCIAL DATA:\n{ctx}"
    messages = [{'role': 'system', 'content': system_prompt}]
    for turn in history[-8:]:
        if turn.get('role') in ('user', 'assistant') and turn.get('content'):
            messages.append({'role': turn['role'], 'content': turn['content']})
    messages.append({'role': 'user', 'content': message})

    def generate():
        try:
            resp = requests.post(
                f"{OLLAMA_URL}/api/chat",
                json={'model': model, 'messages': messages, 'stream': True},
                stream=True,
                timeout=120
            )
            for line in resp.iter_lines():
                if line:
                    chunk = _json.loads(line)
                    token = chunk.get('message', {}).get('content', '')
                    if token:
                        yield f"data: {_json.dumps({'token': token})}\n\n"
                    if chunk.get('done'):
                        yield "data: [DONE]\n\n"
                        break
        except requests.exceptions.ConnectionError:
            yield f"data: {_json.dumps({'error': f'Ollama unreachable at {OLLAMA_URL}'})}\n\n"
        except Exception as e:
            yield f"data: {_json.dumps({'error': str(e)})}\n\n"

    from starlette.responses import StreamingResponse
    return StreamingResponse(generate(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


async def api_chat(request: Request):
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({'error': 'invalid JSON'}, status_code=400)
    message        = str(body.get('message', '')).strip()
    context_type   = body.get('context_type', 'all')
    model          = body.get('model', 'llama3.1:8b')
    history        = body.get('history', [])
    sim_data       = body.get('sim_data') if context_type in ('all', 'simulation') else None
    dashboard_data = None
    if context_type in ('all', 'dashboard'):
        try:
            dashboard_data = await asyncio.to_thread(read_dashboard_data)
        except Exception:
            pass
    ctx = _build_context_string(sim_data, dashboard_data)
    system_prompt = SYSTEM_PROMPT.format(today=time.strftime('%Y-%m-%d'))
    system_prompt = system_prompt + f"\n\nLIVE FINANCIAL DATA:\n{ctx}"
    messages = [{'role': 'system', 'content': system_prompt}]
    for turn in history[-8:]:
        if turn.get('role') in ('user', 'assistant') and turn.get('content'):
            messages.append({'role': turn['role'], 'content': turn['content']})
    messages.append({'role': 'user', 'content': message})
    t0 = time.time()
    try:
        resp = await asyncio.to_thread(
            requests.post,
            f"{OLLAMA_URL}/api/chat",
            json={'model': model, 'messages': messages, 'stream': False},
            timeout=120
        )
        data = resp.json()
        if 'error' in data:
            return JSONResponse({'error': f'Ollama error: {data["error"]}'})
        reply = data['message']['content']
    except requests.exceptions.ConnectionError as e:
        return JSONResponse({'error': f'Ollama unreachable at {OLLAMA_URL} — is the Ollama container running?'})
    except Exception as e:
        return JSONResponse({'error': f'Ollama error: {e}'})
    return JSONResponse({'reply': reply, 'model': model, 'elapsed_ms': round((time.time()-t0)*1000)})

async def api_summarize(request: Request):
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({'error': 'invalid JSON'}, status_code=400)
    model        = body.get('model', 'llama3.1:8b')
    summary_type = body.get('summary_type', 'playbook')
    sim_data     = body.get('sim_data')
    dash_data    = body.get('dashboard_data')
    if not dash_data:
        try:
            dash_data = await asyncio.to_thread(read_dashboard_data)
        except Exception:
            pass
    ctx = _build_context_string(sim_data, dash_data)
    today = time.strftime('%Y-%m-%d')
    if summary_type == 'playbook':
        prompt = (
            f"You are a fiduciary retirement planning AI. Today is {today}.\n"
            f"The user plans to retire at 62, claim SS at 67, and uses a SGOV bridge moat strategy.\n\n"
            f"SIMULATION DATA:\n{ctx}\n\n"
            f"Generate a 3-5 bullet Advisor Playbook. Each bullet must:\n"
            f"- Reference a specific dollar amount or percentage from the data\n"
            f"- Identify either a strength, a risk, or a concrete action item\n"
            f"- Draw on historical SWR research or sequence-of-returns literature where relevant\n"
            f"Format: Start each bullet with a symbol (✓ for strength, ⚠ for risk, → for action). "
            f"No headers. Plain text only. Be specific and non-generic."
        )
    else:
        prompt = (
            f"You are a fiduciary retirement planning AI. Today is {today}.\n"
            f"The user plans to retire at 62, claim SS at 67, and uses a SGOV bridge moat strategy.\n\n"
            f"SIMULATION DATA:\n{ctx}\n\n"
            f"Write a 3-paragraph Plan Details narrative:\n"
            f"Paragraph 1: Bridge period analysis (62-67) — SGOV moat adequacy, moat breach risk, "
            f"early SS risk, how the bridge handles adverse returns. Reference specific rates.\n"
            f"Paragraph 2: Post-67 retirement smile trajectory — go-go/slow-go/no-go spending, "
            f"ratchet trigger probability, terminal wealth range across percentiles.\n"
            f"Paragraph 3: Key tail risks — the 10th-percentile scenario, what drives ruin probability, "
            f"and one concrete mitigation strategy based on the data.\n"
            f"Use specific numbers throughout. No generic advice. Plain text, no headers or bullets."
        )
    sys_prompt = SYSTEM_PROMPT.format(today=time.strftime('%Y-%m-%d'))
    messages = [
        {'role': 'system', 'content': sys_prompt},
        {'role': 'user',   'content': prompt},
    ]
    t0 = time.time()
    try:
        resp = await asyncio.to_thread(
            requests.post,
            f"{OLLAMA_URL}/api/chat",
            json={'model': model, 'messages': messages, 'stream': False},
            timeout=120
        )
        data = resp.json()
        if 'error' in data:
            return JSONResponse({'error': f'Ollama error: {data["error"]}'})
        summary = data['message']['content']
    except requests.exceptions.ConnectionError:
        return JSONResponse({'error': f'Ollama unreachable at {OLLAMA_URL} — is the Ollama container running?'})
    except Exception as e:
        return JSONResponse({'error': f'Ollama error: {e}'})
    return JSONResponse({'summary': summary, 'model': model, 'elapsed_ms': round((time.time()-t0)*1000)})

async def api_ss_sensitivity(request: Request):
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON"}, status_code=400)
    base_params = dict(body)
    base_params['trials'] = min(int(base_params.get('trials', 500)), 1000)
    base_terminal = None
    results = []
    for test_age in [62, 64, 67, 69, 70]:
        p = {**base_params, 'ss_age': test_age}
        r = run_monte_carlo(p)
        terminal = r['stats']['median_terminal']
        results.append({
            "ss_age":          test_age,
            "success_pct":     r['success_pct'],
            "median_terminal": terminal,
            "median_ss_age":   r['stats']['median_ss_age'],
        })
        if test_age == 67:
            base_terminal = terminal
    for row in results:
        row['delta_from_67'] = round(row['median_terminal'] - (base_terminal or 0))
    return JSONResponse({"results": results})

# ── Dashboard HTML ─────────────────────────────────────────────────────────────

DASHBOARD_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Road To FI</title>
<meta name="theme-color" content="#080d1a">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="Road To FI">
<link rel="manifest" href="/manifest.json">
<link rel="apple-touch-icon" href="/icon-192.svg">
<script>if('serviceWorker' in navigator) navigator.serviceWorker.register('/sw.js');</script>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;600;700&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.4/dist/chart.umd.min.js" defer></script>
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  :root {
    --bg:          #080d1a;
    --surface:     #0f1829;
    --surface2:    #131e30;
    --surface3:    #1a2540;
    --border:      rgba(255,255,255,0.06);
    --border2:     rgba(255,255,255,0.11);
    --border-hi:   rgba(236,72,153,0.40);
    --text:        #f0f4ff;
    --muted:       #8899b8;
    --muted2:      #526070;
    --green:       #10b981;
    --green-glow:  rgba(16,185,129,0.25);
    --red:         #f43f5e;
    --red-glow:    rgba(244,63,94,0.25);
    --amber:       #f59e0b;
    --pink:        #ec4899;
    --pink-dim:    #9d174d;
    --pink-glow:   rgba(236,72,153,0.30);
    --pink-glow2:  rgba(236,72,153,0.12);
    --pink-glass:  rgba(236,72,153,0.07);
    --gold:        #ec4899;
    --gold-dim:    #9d174d;
    --grad-card:   linear-gradient(135deg,#131e30 0%,#0f1829 100%);
    --grad-pink:   linear-gradient(135deg,#ec4899 0%,#be185d 100%);
    --grad-success:linear-gradient(135deg,#10b981 0%,#059669 100%);
    --grad-amber:  linear-gradient(135deg,#f59e0b 0%,#d97706 100%);
    --grad-teal:   linear-gradient(135deg,#14b8a6 0%,#0d9488 100%);
    --grad-indigo: linear-gradient(135deg,#6366f1 0%,#4338ca 100%);
    --grad-sky:    linear-gradient(135deg,#38bdf8 0%,#0284c7 100%);
    --grad-violet: linear-gradient(135deg,#a78bfa 0%,#7c3aed 100%);
    --font-ui:     'Inter', system-ui, -apple-system, 'Segoe UI', sans-serif;
    --font-mono:   'JetBrains Mono', ui-monospace, 'Cascadia Code', 'Fira Mono', monospace;
    --radius:      12px;
    --radius-sm:   8px;
    --radius-pill: 9999px;
    --bottom-nav-h:64px;
  }
  body { font-family: var(--font-ui); background: var(--bg); color: var(--text); min-height: 100vh; -webkit-font-smoothing: antialiased; -moz-osx-font-smoothing: grayscale; }

  /* ── Header ── */
  header {
    padding: 0 28px; height: 58px;
    display: flex; align-items: center; justify-content: space-between;
    background: var(--surface);
    backdrop-filter: blur(12px); -webkit-backdrop-filter: blur(12px);
    position: sticky; top: 0; z-index: 100;
    border-bottom: 1px solid var(--border2);
  }
  .header-brand { display: flex; align-items: center; gap: 10px; }
  .header-logo {
    width: 30px; height: 30px;
    background: var(--grad-pink);
    border-radius: 8px; display: flex; align-items: center; justify-content: center;
    box-shadow: 0 0 14px var(--pink-glow); flex-shrink: 0;
  }
  header h1 { color: var(--text); font-size: 0.88rem; letter-spacing: 0.02em; font-weight: 700; }
  header h1 span { color: var(--pink); }
  .refresh-ts {
    display: flex; align-items: center; gap: 5px;
    color: var(--muted2); font-size: 0.60rem; font-family: var(--font-mono);
  }
  .refresh-ts::before {
    content: ''; width: 5px; height: 5px; border-radius: 50%;
    background: var(--green); flex-shrink: 0;
    animation: livePulse 2s ease-in-out infinite;
  }
  @keyframes livePulse { 0%,100% { opacity:1; } 50% { opacity:0.2; } }

  /* ── Desktop Tab nav ── */
  .tab-nav {
    display: flex; border-bottom: 1px solid var(--border);
    padding: 0 20px; background: var(--surface);
    position: sticky; top: 58px; z-index: 90; gap: 2px;
  }
  .tab-btn {
    background: none; border: none; border-bottom: 2px solid transparent;
    color: var(--muted2); font-family: var(--font-ui); font-size: 0.72rem;
    font-weight: 600; letter-spacing: 0.06em; text-transform: uppercase;
    padding: 14px 16px 12px; cursor: pointer; transition: color 0.2s;
    margin-bottom: -1px; white-space: nowrap; position: relative; min-height: unset;
  }
  .tab-btn::after {
    content: ''; position: absolute; bottom: -1px; left: 16px; right: 16px;
    height: 2px; background: var(--grad-pink); border-radius: 2px 2px 0 0;
    transform: scaleX(0); transition: transform 0.2s ease;
  }
  .tab-btn:hover { color: var(--text); border-color: transparent; background: none; }
  .tab-btn.active { color: var(--pink); border-bottom-color: transparent; }
  .tab-btn.active::after { transform: scaleX(1); }
  @media (max-width: 768px) { .tab-nav { display: none; } }

  /* ── Mobile bottom nav ── */
  .bottom-nav { display: none; }
  @media (max-width: 768px) {
    .bottom-nav {
      display: flex; position: fixed; bottom: 0; left: 0; right: 0;
      height: var(--bottom-nav-h);
      background: rgba(8,13,26,0.95);
      backdrop-filter: blur(20px); -webkit-backdrop-filter: blur(20px);
      border-top: 1px solid var(--border2); z-index: 200;
      padding: 0 4px; padding-bottom: env(safe-area-inset-bottom,0px);
      overflow-x: auto; overflow-y: hidden;
      scrollbar-width: none; -ms-overflow-style: none;
    }
    .bottom-nav::-webkit-scrollbar { display: none; }
    .bn-item {
      flex: 0 0 64px; display: flex; flex-direction: column; align-items: center; justify-content: center; gap: 3px;
      background: none; border: none; color: var(--muted2);
      font-family: var(--font-ui); font-size: 0.55rem; font-weight: 600;
      letter-spacing: 0.04em; text-transform: uppercase; cursor: pointer;
      padding: 6px 2px; min-height: 44px; transition: color 0.15s;
      -webkit-tap-highlight-color: transparent; border-radius: 0; box-shadow: none;
    }
    .bn-item:hover { color: var(--muted); border-color: transparent; background: none; transform: none; }
    .bn-item svg { transition: transform 0.15s, filter 0.15s; }
    .bn-item.active { color: var(--pink); }
    .bn-item.active svg { filter: drop-shadow(0 0 5px rgba(236,72,153,0.7)); transform: translateY(-1px); }
    .bn-item:active { opacity: 0.7; }
  }

  /* ── Tab content ── */
  .tab-pane { display: none; }
  .tab-pane.active { display: block; animation: tabFadeIn 0.22s ease forwards; }
  @keyframes tabFadeIn { from { opacity:0; transform:translateY(6px); } to { opacity:1; transform:translateY(0); } }
  main { max-width: 980px; margin: 0 auto; padding: 28px 24px; display: flex; flex-direction: column; gap: 20px; }

  /* ── Glass panels ── */
  .panel, .glass {
    background: linear-gradient(135deg, var(--surface2) 0%, var(--surface) 100%);
    border: 1px solid var(--border); border-radius: var(--radius); padding: 22px 26px;
    position: relative; overflow: hidden;
  }
  .panel::before, .glass::before {
    content: ''; position: absolute; top: 0; left: 0; right: 0; height: 1px;
    background: linear-gradient(90deg, transparent 0%, var(--border2) 30%, var(--border2) 70%, transparent 100%);
    pointer-events: none;
  }
  .panel::after, .glass::after {
    content: ''; position: absolute; top: -60px; left: -60px;
    width: 200px; height: 200px;
    background: var(--panel-glow, transparent);
    border-radius: 50%; pointer-events: none; opacity: 0.6;
  }
  .panel h2, .glass h2 {
    font-family: var(--font-ui); font-size: 0.68rem; font-weight: 700;
    letter-spacing: 0.1em; text-transform: uppercase; color: var(--muted);
    margin-bottom: 18px; position: relative; z-index: 1;
  }

  /* ── Accent system (ambient glow replaces border-left) ── */
  .accent-pink    { --panel-glow: radial-gradient(circle,rgba(236,72,153,0.10) 0%,transparent 70%); }
  .accent-emerald { --panel-glow: radial-gradient(circle,rgba(16,185,129,0.10) 0%,transparent 70%); }
  .accent-indigo  { --panel-glow: radial-gradient(circle,rgba(99,102,241,0.10) 0%,transparent 70%); }
  .accent-amber   { --panel-glow: radial-gradient(circle,rgba(245,158,11,0.10) 0%,transparent 70%); }
  .accent-sky     { --panel-glow: radial-gradient(circle,rgba(56,189,248,0.10) 0%,transparent 70%); }
  .accent-violet  { --panel-glow: radial-gradient(circle,rgba(167,139,250,0.10) 0%,transparent 70%); }
  .accent-teal    { --panel-glow: radial-gradient(circle,rgba(20,184,166,0.10) 0%,transparent 70%); }
  .accent-rose    { --panel-glow: radial-gradient(circle,rgba(244,63,94,0.10) 0%,transparent 70%); }

  /* ── Stat cards ── */
  .stat-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(155px, 1fr)); gap: 12px; margin-bottom: 20px; }
  .stat-card {
    background: var(--grad-card); border: 1px solid var(--border); border-radius: var(--radius);
    padding: 16px 18px; position: relative; overflow: hidden;
    transition: transform 0.2s ease, box-shadow 0.2s ease, border-color 0.2s ease;
  }
  .stat-card::before {
    content: ''; position: absolute; top: 0; left: 0; right: 0; height: 2px;
    background: var(--card-accent, var(--grad-pink));
  }
  .stat-card::after {
    content: ''; position: absolute; top: -24px; left: -24px;
    width: 80px; height: 80px; background: var(--card-glow, var(--pink-glass)); border-radius: 50%; pointer-events: none;
  }
  .stat-card:hover { transform: translateY(-2px); box-shadow: 0 8px 32px rgba(0,0,0,0.45); border-color: var(--border2); }
  .stat-card.accent-pink    { --card-accent: var(--grad-pink);    --card-glow: var(--pink-glass); }
  .stat-card.accent-emerald { --card-accent: var(--grad-success);  --card-glow: rgba(16,185,129,0.07); }
  .stat-card.accent-indigo  { --card-accent: var(--grad-indigo);   --card-glow: rgba(99,102,241,0.07); }
  .stat-card.accent-amber   { --card-accent: var(--grad-amber);    --card-glow: rgba(245,158,11,0.07); }
  .stat-card.accent-sky     { --card-accent: var(--grad-sky);      --card-glow: rgba(56,189,248,0.07); }
  .stat-card.accent-violet  { --card-accent: var(--grad-violet);   --card-glow: rgba(167,139,250,0.07); }
  .stat-card.accent-teal    { --card-accent: var(--grad-teal);     --card-glow: rgba(20,184,166,0.07); }
  .stat-card.accent-rose    { --card-accent: linear-gradient(135deg,#f43f5e,#be123c); --card-glow: rgba(244,63,94,0.07); }
  .stat-card .sc-label { font-family: var(--font-ui); font-size: 0.65rem; font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.08em; color: var(--muted); margin-bottom: 8px; position: relative; z-index: 1; }
  .stat-card .sc-value { font-family: var(--font-mono); font-size: 1.45rem; font-weight: 700; line-height: 1;
    color: var(--text); letter-spacing: -0.02em; position: relative; z-index: 1; }
  .stat-card .sc-sub   { font-family: var(--font-ui); font-size: 0.62rem; color: var(--muted2); margin-top: 5px; position: relative; z-index: 1; }
  .stat-card.gl-pos .sc-value { color: var(--green); }
  .stat-card.gl-neg .sc-value { color: var(--red); }

  /* ── Progress bar ── */
  .fi-progress-wrap { margin-bottom: 20px; }
  .fi-label { display: flex; justify-content: space-between; font-size: 0.68rem; color: var(--muted); margin-bottom: 6px; }
  .fi-label .pct { color: var(--pink); font-weight: 700; font-family: var(--font-mono); }
  .progress-bar { background: rgba(255,255,255,0.06); border-radius: 6px; height: 8px; overflow: hidden; }
  .progress-fill { background: var(--grad-pink); height: 100%; border-radius: 6px;
    transition: width 0.8s cubic-bezier(0.4,0,0.2,1); box-shadow: 0 0 8px var(--pink-glow); }

  /* ── Freedom levels ── */
  .levels-grid { display: flex; flex-direction: column; gap: 6px; }
  .level-row { display: grid; grid-template-columns: 1fr auto auto; align-items: center; gap: 12px;
    background: var(--grad-card); border: 1px solid var(--border); border-radius: var(--radius-sm); padding: 9px 14px;
    transition: border-color 0.2s, transform 0.2s; }
  .level-row:hover { border-color: var(--border2); transform: translateX(2px); }
  .level-row .lname  { font-size: 0.75rem; font-weight: 500; }
  .level-row .lgoal  { font-size: 0.72rem; color: var(--muted2); text-align: right; font-family: var(--font-mono); }
  .level-row .lbadge { font-size: 0.60rem; font-weight: 700; letter-spacing: 0.06em;
    padding: 3px 9px; border-radius: var(--radius-pill); white-space: nowrap; }
  .badge-achieved { background: rgba(16,185,129,0.10); color: var(--green); border: 1px solid rgba(16,185,129,0.25); }
  .badge-progress { background: var(--pink-glass); color: var(--pink); border: 1px solid rgba(236,72,153,0.25); }

  /* ── Two-column ── */
  .two-col { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; margin-top: 14px; }
  @media (max-width: 600px) { .two-col { grid-template-columns: 1fr; } }
  .sub-panel { background: var(--grad-card); border: 1px solid var(--border); border-radius: var(--radius-sm); padding: 14px 16px; position: relative; }
  .sub-panel h3 { color: var(--muted2); font-size: 0.62rem; text-transform: uppercase; letter-spacing: 0.09em; margin-bottom: 12px; font-weight: 700; }
  .alloc-row { display: flex; justify-content: space-between; font-size: 0.75rem; padding: 5px 0; border-bottom: 1px solid var(--border); }
  .alloc-row:last-child { border-bottom: none; }
  .alloc-row .aname { color: var(--muted); }
  .alloc-row span:last-child { font-family: var(--font-mono); }
  .cashflow-row { display: flex; justify-content: space-between; font-size: 0.75rem; padding: 5px 0; }
  .cashflow-row .cname { color: var(--muted); }
  .cval.pos { color: var(--green); font-family: var(--font-mono); }
  .cval.neg { color: var(--red); font-family: var(--font-mono); }
  .cval.neu { color: var(--pink); font-family: var(--font-mono); }

  /* ── Spending table ── */
  .spending-table { width: 100%; border-collapse: collapse; font-size: 0.73rem; margin-top: 18px; overflow-x: auto; display: block; }
  .spending-table th { color: var(--muted2); font-weight: 600; text-align: right; padding: 6px 8px;
    font-size: 0.63rem; text-transform: uppercase; letter-spacing: 0.06em; border-bottom: 1px solid var(--border2); white-space: nowrap; }
  .spending-table th:first-child { text-align: left; }
  .spending-table td { padding: 6px 8px; text-align: right; border-bottom: 1px solid var(--border); font-family: var(--font-mono); }
  .spending-table td:first-child { text-align: left; color: var(--muted); font-family: var(--font-ui); white-space: nowrap; }
  .spending-table tr:last-child td { border-bottom: none; }
  .spending-table tr:hover td { background: rgba(236,72,153,0.04); }
  .val-pos { color: var(--green); } .val-neg { color: var(--red); } .val-neu { color: var(--pink); }

  /* ── Inputs & buttons ── */
  input[type="number"], input[type="text"], input[type="password"], select {
    background: rgba(255,255,255,0.03); border: 1px solid var(--border2); color: var(--text);
    font-family: var(--font-mono); font-size: 0.88rem; padding: 9px 12px;
    border-radius: var(--radius-sm); outline: none; width: 100%;
    transition: border-color 0.15s, box-shadow 0.15s; min-height: 44px;
  }
  input:focus, select:focus { border-color: var(--pink); box-shadow: 0 0 0 3px var(--pink-glow2); }
  label { font-family: var(--font-ui); font-size: 0.68rem; font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.06em; color: var(--muted); display: block; margin-bottom: 5px; }
  button {
    background: var(--surface3); color: var(--muted); border: 1px solid var(--border2);
    border-radius: var(--radius-sm); font-family: var(--font-ui); font-size: 0.78rem; font-weight: 600;
    letter-spacing: 0.04em; text-transform: uppercase; padding: 9px 18px;
    cursor: pointer; transition: all 0.2s ease; white-space: nowrap; min-height: 44px;
  }
  button:hover { background: var(--surface2); border-color: var(--border-hi); color: var(--text); }
  button:disabled { opacity: 0.35; cursor: not-allowed; }
  button.btn-primary {
    background: var(--grad-pink); color: #fff; border: none; font-weight: 700; letter-spacing: 0.06em;
    box-shadow: 0 0 20px var(--pink-glow), 0 4px 12px rgba(0,0,0,0.4), inset 0 1px 0 rgba(255,255,255,0.15);
    position: relative; overflow: hidden;
  }
  button.btn-primary::after { content: ''; position: absolute; inset: 0;
    background: linear-gradient(180deg,rgba(255,255,255,0.10) 0%,transparent 60%); pointer-events: none; }
  button.btn-primary:hover {
    box-shadow: 0 0 32px rgba(236,72,153,0.50), 0 6px 20px rgba(0,0,0,0.4), inset 0 1px 0 rgba(255,255,255,0.2);
    transform: translateY(-1px); border: none; color: #fff; background: var(--grad-pink);
  }
  button.btn-primary:active { transform: translateY(0); box-shadow: 0 0 16px var(--pink-glow), 0 2px 8px rgba(0,0,0,0.4); }
  .btn { background: var(--surface3); color: var(--muted); border: 1px solid var(--border2);
    border-radius: var(--radius-sm); font-family: var(--font-ui); font-size: 0.78rem; font-weight: 600;
    letter-spacing: 0.04em; text-transform: uppercase; padding: 9px 18px;
    cursor: pointer; transition: all 0.2s ease; white-space: nowrap; min-height: 44px; }
  .btn:hover { background: var(--surface2); border-color: var(--border-hi); color: var(--text); }
  .btn-sm { font-size: 0.60rem; padding: 4px 10px; border-radius: 6px; min-height: 28px; }
  .price { color: var(--pink); font-size: 1.5rem; font-weight: bold; font-family: var(--font-mono); }
  .change.up { color: var(--green); } .change.down { color: var(--red); }
  .err { color: var(--red); font-size: 0.8rem; }
  .key-status { font-size: 0.65rem; color: var(--muted2); margin-top: 5px; }
  .key-status .saved { color: var(--green); }

  /* ── Rules ── */
  .rules-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(175px, 1fr)); gap: 12px; }
  .note { color: var(--muted); font-size: 0.72rem; margin-top: 16px; border-top: 1px solid var(--border2); padding-top: 12px; }
  .note span { color: #f59e0b; }

  /* ── Forms ── */
  .form-row { display: flex; gap: 10px; flex-wrap: wrap; align-items: flex-end; margin-bottom: 14px; }
  .field { display: flex; flex-direction: column; gap: 4px; }
  .gain { color: var(--green); font-family: var(--font-mono); }
  .loss { color: var(--red); font-family: var(--font-mono); }

  /* ── Portfolio ── */
  .port-summary { display: grid; grid-template-columns: repeat(4,1fr); gap: 10px; margin-bottom: 22px; }
  @media (max-width: 650px) { .port-summary { grid-template-columns: repeat(2,1fr); } }
  .port-accounts { display: flex; flex-direction: column; gap: 12px; }
  .port-account { border: 1px solid var(--border); border-radius: var(--radius); overflow: hidden; transition: border-color 0.2s; }
  .port-account:hover { border-color: var(--border2); }
  .port-acct-hdr {
    display: flex; justify-content: space-between; align-items: center;
    padding: 11px 16px; background: var(--grad-card); border-bottom: 1px solid var(--border);
  }
  .port-acct-left { display: flex; align-items: center; gap: 9px; }
  .port-acct-tag {
    font-size: 0.58rem; font-weight: 700; letter-spacing: 0.08em; padding: 3px 8px;
    border-radius: var(--radius-pill); text-transform: uppercase; white-space: nowrap; min-height: unset;
  }
  .tag-roth    { background: rgba(236,72,153,0.12);  color: #ec4899; border: 1px solid rgba(236,72,153,0.3); }
  .tag-taxable { background: rgba(56,189,248,0.08);  color: #38bdf8; border: 1px solid rgba(56,189,248,0.2); }
  .tag-401k    { background: rgba(245,158,11,0.08);  color: #f59e0b; border: 1px solid rgba(245,158,11,0.2); }
  .tag-crypto  { background: rgba(167,139,250,0.08); color: #a78bfa; border: 1px solid rgba(167,139,250,0.2); }
  .port-acct-name  { color: var(--text); font-size: 0.74rem; font-weight: 500; }
  .port-acct-total { color: var(--pink); font-size: 0.92rem; font-weight: 700; font-family: var(--font-mono); }
  .port-htable { width: 100%; border-collapse: collapse; background: rgba(8,13,26,0.5); }
  .port-htable td { padding: 9px 16px; font-size: 0.73rem; border-bottom: 1px solid var(--border); vertical-align: middle; }
  .port-htable tr:last-child td { border-bottom: none; }
  .port-htable tr:hover td { background: rgba(236,72,153,0.04); }
  .port-htable .h-ticker { width: 1%; white-space: nowrap; }
  .port-htable .h-name   { color: var(--muted); font-size: 0.71rem; }
  .port-htable .h-r      { text-align: right; white-space: nowrap; }
  .port-htable .h-muted  { color: var(--muted2); font-size: 0.68rem; font-family: var(--font-mono); }
  .port-htable .h-val    { color: var(--text); font-weight: 700; text-align: right; font-family: var(--font-mono); }
  .port-htable .h-gl     { text-align: right; }
  .ticker-badge {
    display: inline-block; font-family: var(--font-mono); font-size: 0.7rem; font-weight: 700;
    color: var(--pink); background: var(--pink-glass); border: 1px solid rgba(236,72,153,0.25);
    padding: 2px 8px; border-radius: 5px; letter-spacing: 0.04em;
  }
  .gl-pill {
    display: inline-block; font-size: 0.63rem; font-weight: 700; padding: 2px 9px;
    border-radius: var(--radius-pill); white-space: nowrap; letter-spacing: 0.03em; font-family: var(--font-mono);
  }
  .gl-pill.gain  { background: rgba(16,185,129,0.10); color: var(--green); border: 1px solid rgba(16,185,129,0.20); }
  .gl-pill.loss  { background: rgba(244,63,94,0.10);  color: var(--red);   border: 1px solid rgba(244,63,94,0.20); }
  .gl-pill.proxy { background: transparent; color: var(--muted2); border: 1px solid var(--border2); font-style: italic; font-weight: 400; }
  .port-refresh-bar { display: flex; gap: 10px; align-items: flex-end; flex-wrap: wrap; margin-top: 18px;
    padding-top: 16px; border-top: 1px solid var(--border2); }

  /* ── Monte Carlo ── */
  .mc-layout { display: grid; grid-template-columns: 330px 1fr; gap: 20px; align-items: start; }
  @media (max-width: 768px) { .mc-layout { grid-template-columns: 1fr; } }
  .mc-inputs { display: flex; flex-direction: column; gap: 10px; }
  .mc-section-title {
    color: var(--muted2); font-size: 0.62rem; text-transform: uppercase; letter-spacing: 0.1em;
    border-bottom: 1px solid var(--border2); padding-bottom: 5px; margin-top: 8px; font-weight: 600;
  }
  .mc-row-3 { display: grid; grid-template-columns: repeat(3,1fr); gap: 8px; }
  .mc-row-2 { display: grid; grid-template-columns: repeat(2,1fr); gap: 8px; }
  .mc-field { display: flex; flex-direction: column; gap: 4px; }
  .mc-select {
    background: rgba(255,255,255,0.03); border: 1px solid var(--border2); color: var(--text);
    font-family: var(--font-mono); font-size: 0.88rem; padding: 9px 10px; border-radius: var(--radius-sm); outline: none; width: 100%; min-height: 44px;
  }
  .mc-select:focus { border-color: var(--pink); }
  .mc-toggle-row { display: flex; flex-wrap: wrap; gap: 12px; }
  .mc-toggle { display: flex; align-items: center; gap: 5px; cursor: pointer; font-size: 0.72rem; color: var(--muted); }
  .mc-toggle input { accent-color: var(--pink); }
  .mc-radio-row { display: flex; gap: 18px; }
  .mc-radio-row label { display: flex; align-items: center; gap: 5px; cursor: pointer; font-size: 0.72rem; color: var(--muted); }
  .mc-radio-row input { accent-color: var(--pink); }

  /* Presets */
  .mc-presets { display: flex; gap: 7px; flex-wrap: wrap; margin-bottom: 2px; }
  .preset-pill {
    font-family: var(--font-ui); font-size: 0.65rem; font-weight: 700; text-transform: uppercase;
    letter-spacing: 0.06em; padding: 6px 14px; border-radius: var(--radius-pill); cursor: pointer;
    transition: all 0.2s; border: 1px solid var(--border2); background: var(--surface3); color: var(--muted2); min-height: unset;
  }
  .preset-pill.active, .preset-pill.baseline:hover, .preset-pill.baseline.active
    { background: var(--grad-pink); border-color: transparent; color: #fff; box-shadow: 0 0 14px var(--pink-glow); }
  .preset-pill.cautious:hover, .preset-pill.cautious.active
    { background: var(--grad-amber); border-color: transparent; color: #000; }
  .preset-pill.stress:hover, .preset-pill.stress.active
    { background: linear-gradient(135deg,#f43f5e,#be123c); border-color: transparent; color: #fff; }
  .preset-pill.optimist:hover, .preset-pill.optimist.active
    { background: var(--grad-success); border-color: transparent; color: #fff; }

  /* Advanced collapsible */
  details.mc-advanced { border: 1px solid var(--border); border-radius: var(--radius-sm); overflow: hidden; margin-top: 4px; }
  details.mc-advanced summary {
    padding: 10px 12px; cursor: pointer; font-size: 0.72rem; font-weight: 600;
    color: var(--muted); letter-spacing: 0.04em; list-style: none;
    background: var(--surface2); user-select: none;
  }
  details.mc-advanced summary::-webkit-details-marker { display: none; }
  details.mc-advanced[open] summary { border-bottom: 1px solid var(--border2); color: var(--text); }
  details.mc-advanced .adv-body { padding: 12px; display: flex; flex-direction: column; gap: 10px; }
  .mc-defaults-note {
    font-size: 0.65rem; color: var(--muted2); padding: 6px 10px; border-radius: var(--radius-sm);
    background: rgba(16,185,129,0.05); border: 1px solid rgba(16,185,129,0.15); margin-bottom: 2px;
  }

  /* Success Badge with pulsing rings */
  .success-badge {
    text-align: center; padding: 30px 22px 24px; border-radius: var(--radius);
    background: linear-gradient(135deg, var(--surface2) 0%, var(--surface) 100%);
    border: 1px solid var(--border); margin-bottom: 16px; position: relative; overflow: hidden;
  }
  .success-badge::before {
    content: ''; position: absolute; inset: 0;
    background: radial-gradient(circle at 50% 35%, var(--badge-glow, rgba(16,185,129,0.10)) 0%, transparent 65%);
    pointer-events: none;
  }
  .sb-ring-wrap { position: relative; display: inline-block; margin-bottom: 12px; }
  .sb-ring {
    position: absolute; inset: -12px; border-radius: 50%;
    border: 1.5px solid var(--badge-ring-color, var(--green)); opacity: 0;
    animation: ringPulse 2.4s ease-out infinite;
  }
  .sb-ring:nth-child(2) { animation-delay: 0.8s; }
  .sb-ring:nth-child(3) { animation-delay: 1.6s; }
  @keyframes ringPulse { 0% { transform:scale(0.85); opacity:0.6; } 100% { transform:scale(1.7); opacity:0; } }
  .success-badge.badge-great { --badge-glow:rgba(16,185,129,0.12); --badge-ring-color:var(--green); }
  .success-badge.badge-ok    { --badge-glow:rgba(245,158,11,0.12);  --badge-ring-color:var(--amber); }
  .success-badge.badge-warn  { --badge-glow:rgba(244,63,94,0.12);   --badge-ring-color:var(--red); }
  .success-badge .sbval { font-family:var(--font-mono); font-size:3.4rem; font-weight:700; line-height:1; letter-spacing:-0.03em; position:relative; z-index:1; }
  .success-badge .sblabel { font-size:0.62rem; text-transform:uppercase; letter-spacing:0.14em; color:var(--muted2); margin-top:10px; position:relative; z-index:1; }
  .clr-green { color: var(--green); } .clr-amber { color: #f59e0b; } .clr-red { color: var(--red); }
  .mc-table { width: 100%; border-collapse: collapse; font-size: 0.74rem; }
  .mc-table th { color: var(--muted2); font-weight: 600; text-align: right; padding: 7px 8px;
    font-size: 0.62rem; text-transform: uppercase; letter-spacing: 0.07em; border-bottom: 1px solid var(--border2); }
  .mc-table th:first-child { text-align: left; }
  .mc-table td { padding: 7px 8px; text-align: right; border-bottom: 1px solid var(--border); font-family: var(--font-mono); }
  .mc-table td:first-child { text-align: left; color: var(--muted); font-family: var(--font-ui); }
  .mc-table tr:last-child td { border-bottom: none; }
  .mc-table tr:hover td { background: rgba(255,255,255,0.03); }
  .mc-stats-strip { display: grid; grid-template-columns: repeat(auto-fit, minmax(125px, 1fr)); gap: 8px; }
  .mc-runtime { color: var(--muted2); font-size: 0.62rem; text-align: right; margin-top: 10px; font-family: var(--font-mono); }
  .mc-result-hint { color: var(--muted2); font-size: 0.78rem; text-align: center; padding: 60px 0; }

  /* Playbook */
  .playbook-panel {
    border: 1px solid var(--border); border-radius: var(--radius); padding: 20px 22px;
    margin-top: 14px; background: linear-gradient(135deg, var(--surface2) 0%, var(--surface) 100%);
    position: relative; overflow: hidden;
  }
  .playbook-panel::after { content: ''; position: absolute; top:-40px; left:-40px;
    width:150px; height:150px; background:radial-gradient(circle,rgba(99,102,241,0.06) 0%,transparent 70%);
    border-radius:50%; pointer-events:none; }
  .playbook-panel h3 { font-size: 0.65rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.1em; color: var(--muted2); margin-bottom: 14px; position: relative; z-index: 1; }
  .playbook-line { display: flex; align-items: flex-start; gap: 10px; padding: 9px 0; border-bottom: 1px solid var(--border); font-size: 0.76rem; line-height: 1.55; position: relative; z-index: 1; }
  .playbook-line:last-child { border-bottom: none; }
  .playbook-dot { width: 8px; height: 8px; border-radius: 50%; flex-shrink: 0; margin-top: 6px; }
  .pb-emerald .playbook-dot { background: #10b981; } .pb-emerald { color: var(--text); }
  .pb-amber   .playbook-dot { background: #f59e0b; } .pb-amber   { color: #fcd34d; }
  .pb-rose    .playbook-dot { background: #f43f5e; } .pb-rose    { color: #fca5a5; }
  .pb-teal    .playbook-dot { background: #14b8a6; } .pb-teal    { color: #5eead4; }

  /* Collapsible ledger sections */
  .section-toggle { width: 100%; background: none; border: none; border-top: 1px solid var(--border);
    display: flex; justify-content: space-between; align-items: center; padding: 11px 0 7px;
    cursor: pointer; color: var(--muted); font-family: var(--font-ui);
    font-size: 0.66rem; font-weight: 700; text-transform: uppercase;
    letter-spacing: 0.09em; margin-top: 14px; transition: color 0.15s; min-height: unset; box-shadow: none; }
  .section-toggle:hover { color: var(--text); border-color: transparent; background: none; }
  .section-toggle .st-arrow { font-size: 0.7rem; transition: transform 0.2s; }
  .section-toggle.open .st-arrow { transform: rotate(90deg); }
  .section-body { overflow: hidden; transition: max-height 0.3s ease; }

  /* Allocation donut legend */
  .donut-wrap { display: flex; align-items: center; gap: 24px; margin-top: 14px; flex-wrap: wrap; }
  .donut-wrap canvas { flex-shrink: 0; }
  .donut-legend { display: flex; flex-direction: column; gap: 6px; font-size: 0.72rem; }
  .dl-row { display: flex; align-items: center; gap: 7px; }
  .dl-dot { width: 9px; height: 9px; border-radius: 50%; flex-shrink: 0; }
  .dl-name { color: var(--muted); flex: 1; }
  .dl-val  { font-family: var(--font-mono); color: var(--text); }

  /* Tax summary */
  .tax-summary { margin-top: 18px; padding: 16px 18px; background: var(--grad-card); border: 1px solid var(--border); border-radius: var(--radius); }
  .tax-header { font-size: 0.65rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.09em; color: var(--muted2); margin-bottom: 12px; }
  .tax-summary h3 { font-size: 0.65rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.09em; color: var(--muted2); margin-bottom: 12px; }
  .tax-row { display: flex; justify-content: space-between; font-size: 0.74rem; padding: 6px 0; border-bottom: 1px solid var(--border); }
  .tax-row:last-child { border-bottom: none; }
  .tax-row.tax-net { font-weight: 700; }
  .tax-note { font-size: 0.63rem; color: var(--muted2); margin-top: 10px; line-height: 1.5; }

  /* Scenario compare strip */
  .compare-strip { margin-top: 14px; border: 1px solid var(--border); border-radius: var(--radius); overflow: hidden; background: var(--grad-card); }
  .compare-strip h3 { font-size: 0.65rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.09em; color: var(--muted2); padding: 12px 16px; border-bottom: 1px solid var(--border2); }
  .compare-table { width: 100%; border-collapse: collapse; font-size: 0.74rem; }
  .compare-table th { color: var(--muted2); font-weight: 600; text-align: right; padding: 7px 12px; font-size: 0.63rem; text-transform: uppercase; border-bottom: 1px solid var(--border2); }
  .compare-table th:first-child { text-align: left; }
  .compare-table td { padding: 7px 12px; text-align: right; border-bottom: 1px solid var(--border); font-family: var(--font-mono); }
  .compare-table td:first-child { text-align: left; color: var(--muted); font-family: var(--font-ui); }
  .compare-table tr:last-child td { border-bottom: none; }

  /* ── MC Inner Tabs (Results / Plan Details) ── */
  .mc-inner-tabs { display: flex; gap: 6px; margin-bottom: 12px; border-bottom: 1px solid var(--border); padding-bottom: 8px; }
  .mc-itab { background: none; border: none; color: var(--muted2); font-size: 0.72rem; font-weight: 600; padding: 5px 12px; border-radius: 6px; cursor: pointer; letter-spacing: 0.04em; transition: color 0.15s, background 0.15s; }
  .mc-itab:hover { color: var(--text); background: var(--surface2); }
  .mc-itab.active { color: var(--pink); background: rgba(236,72,153,0.12); }

  /* ── Plan Status Hero Card ── */
  .plan-hero { border-radius: var(--radius); padding: 14px 16px; margin-bottom: 12px; display: flex; flex-direction: column; gap: 7px; }
  .hero-secure { background: rgba(16,185,129,0.10); border: 1px solid rgba(16,185,129,0.30); }
  .hero-ok     { background: rgba(245,158,11,0.10); border: 1px solid rgba(245,158,11,0.30); }
  .hero-caution{ background: rgba(249,115,22,0.10); border: 1px solid rgba(249,115,22,0.30); }
  .hero-risk   { background: rgba(244,63,94,0.10);  border: 1px solid rgba(244,63,94,0.30); }
  .ph-verdict  { font-size: 0.82rem; font-weight: 700; letter-spacing: 0.06em; text-transform: uppercase; }
  .hero-secure .ph-verdict { color: #10b981; }
  .hero-ok     .ph-verdict { color: #f59e0b; }
  .hero-caution .ph-verdict{ color: #f97316; }
  .hero-risk   .ph-verdict { color: #f43f5e; }
  .ph-conf-row { display: flex; align-items: center; gap: 10px; }
  .ph-bar { flex: 1; height: 6px; background: var(--border); border-radius: 3px; overflow: hidden; }
  .hero-secure .ph-fill { background: #10b981; }
  .hero-ok     .ph-fill { background: #f59e0b; }
  .hero-caution .ph-fill{ background: #f97316; }
  .hero-risk   .ph-fill { background: #f43f5e; }
  .ph-fill { height: 100%; border-radius: 3px; transition: width 0.8s cubic-bezier(0.4,0,0.2,1); }
  .ph-pct  { font-size: 0.78rem; font-weight: 700; font-family: var(--font-mono); color: var(--text); min-width: 38px; text-align: right; }
  .ph-text { font-size: 0.73rem; color: var(--muted2); line-height: 1.5; }

  /* ── Plan Details Tab 2 ── */
  .pd-section { border-radius: var(--radius-sm); padding: 12px 14px; margin-bottom: 10px; background: var(--grad-card); border: 1px solid var(--border); }
  .pd-section h3 { font-size: 0.65rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.09em; color: var(--muted2); margin-bottom: 10px; }
  .pd-row  { display: flex; justify-content: space-between; font-size: 0.74rem; padding: 5px 0; border-bottom: 1px solid var(--border); }
  .pd-row:last-child { border-bottom: none; }
  .pd-key  { color: var(--muted2); }
  .pd-val  { font-family: var(--font-mono); color: var(--text); font-weight: 600; }
  .pd-warn { background: rgba(244,63,94,0.08); border: 1px solid rgba(244,63,94,0.25); border-radius: 6px; padding: 8px 12px; font-size: 0.72rem; color: #fca5a5; margin-bottom: 7px; }
  .pd-note { background: rgba(245,158,11,0.07); border: 1px solid rgba(245,158,11,0.22); border-radius: 6px; padding: 8px 12px; font-size: 0.72rem; color: #fcd34d; margin-bottom: 7px; }
  .pd-good { background: rgba(16,185,129,0.07); border: 1px solid rgba(16,185,129,0.22); border-radius: 6px; padding: 8px 12px; font-size: 0.72rem; color: #6ee7b7; margin-bottom: 7px; }
  .pd-prose{ font-size: 0.73rem; color: var(--muted2); line-height: 1.65; }

  /* ── Tax Loss Tab ── */
  .tl-summary { display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin-bottom: 14px; }
  .tl-table { width: 100%; border-collapse: collapse; font-size: 0.73rem; }
  .tl-table th { color: var(--muted2); font-weight: 600; padding: 7px 10px; text-align: left; font-size: 0.62rem; text-transform: uppercase; letter-spacing: 0.06em; border-bottom: 1px solid var(--border2); }
  .tl-table td { padding: 8px 10px; border-bottom: 1px solid var(--border); vertical-align: middle; }
  .tl-table tr:last-child td { border-bottom: none; }
  .tl-empty { text-align: center; padding: 32px 16px; color: var(--muted2); font-size: 0.75rem; }

  /* ── Grid Search Panel ── */
  .gs-panel { margin-top: 12px; padding: 14px 16px; background: var(--grad-card); border: 1px solid var(--border); border-radius: var(--radius); }
  .gs-panel h3 { font-size: 0.65rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.09em; color: var(--muted2); margin-bottom: 12px; }
  .gs-axis-row { display: grid; grid-template-columns: 1fr 80px 80px 70px auto; gap: 6px; align-items: center; margin-bottom: 8px; }
  .gs-axis-row select, .gs-axis-row input { background: var(--surface3); border: 1px solid var(--border2); border-radius: var(--radius-sm); color: var(--text); padding: 5px 8px; font-size: 0.72rem; }
  .gs-axis-label { font-size: 0.62rem; color: var(--muted2); text-align: center; }
  .gs-combo-preview { font-size: 0.68rem; color: var(--amber); margin: 8px 0; font-family: var(--font-mono); }
  .gs-result-table { width: 100%; border-collapse: collapse; font-size: 0.72rem; margin-top: 12px; }
  .gs-result-table th { color: var(--muted2); font-weight: 600; padding: 6px 8px; text-align: right; font-size: 0.60rem; text-transform: uppercase; border-bottom: 1px solid var(--border2); }
  .gs-result-table th:first-child { text-align: left; }
  .gs-result-table td { padding: 6px 8px; text-align: right; border-bottom: 1px solid var(--border); font-family: var(--font-mono); }
  .gs-result-table td:first-child { text-align: left; font-family: var(--font-ui); }
  .gs-result-table tr:last-child td { border-bottom: none; }
  .gs-top { background: rgba(16,185,129,0.07); }
  .gs-heatmap-wrap { margin-top: 14px; }

  /* ── AI Playbook (Ollama-generated) ── */
  .ai-playbook { margin-top: 14px; padding: 14px 16px; background: var(--grad-card); border: 1px solid var(--border); border-radius: var(--radius); }
  .ai-playbook-hdr { display: flex; align-items: center; justify-content: space-between; margin-bottom: 12px; }
  .ai-playbook-hdr h3 { font-size: 0.65rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.09em; color: var(--muted2); }
  .ai-playbook-hdr .ai-badge { font-size: 0.58rem; background: rgba(236,72,153,0.15); color: var(--pink); border: 1px solid rgba(236,72,153,0.25); border-radius: 4px; padding: 2px 6px; font-weight: 600; }
  .ai-playbook-body { font-size: 0.74rem; color: var(--text); line-height: 1.65; white-space: pre-wrap; }
  .ai-thinking { display: flex; align-items: center; gap: 8px; color: var(--muted2); font-size: 0.72rem; padding: 8px 0; }
  .ai-thinking-dots { display: inline-flex; gap: 3px; }
  .ai-thinking-dots span { width: 5px; height: 5px; border-radius: 50%; background: var(--pink); animation: aiDot 1.2s infinite; }
  .ai-thinking-dots span:nth-child(2) { animation-delay: 0.2s; }
  .ai-thinking-dots span:nth-child(3) { animation-delay: 0.4s; }
  @keyframes aiDot { 0%,80%,100%{opacity:0.2;transform:scale(0.8)} 40%{opacity:1;transform:scale(1)} }

  /* ── Chat FAB + Overlay ── */
  .chat-fab { position: fixed; bottom: calc(var(--bottom-nav-h) + 16px); right: 16px; width: 48px; height: 48px; border-radius: 50%; background: var(--pink); z-index: 200; font-size: 1.3rem; border: none; cursor: pointer; box-shadow: 0 4px 16px rgba(236,72,153,0.45); transition: transform 0.15s; }
  .chat-fab:hover { transform: scale(1.08); }
  .chat-overlay { position: fixed; bottom: calc(var(--bottom-nav-h) + 76px); right: 16px; width: 320px; max-height: 460px; background: var(--surface2); border: 1px solid var(--border); border-radius: var(--radius); z-index: 199; display: flex; flex-direction: column; box-shadow: 0 8px 32px rgba(0,0,0,0.55); }
  @media (max-width: 430px) { .chat-overlay { width: calc(100vw - 32px); right: 16px; } }
  .chat-hdr { display: flex; align-items: center; gap: 6px; padding: 8px 12px; border-bottom: 1px solid var(--border2); font-size: 0.72rem; font-weight: 700; color: var(--text); }
  .chat-hdr span { flex: 1; }
  .chat-hdr select { background: var(--surface3); border: 1px solid var(--border2); border-radius: 4px; color: var(--text); font-size: 0.65rem; padding: 3px 5px; }
  .chat-hdr button { background: none; border: none; color: var(--muted2); cursor: pointer; font-size: 0.9rem; padding: 2px 4px; }
  .chat-hist { flex: 1; overflow-y: auto; padding: 10px; display: flex; flex-direction: column; gap: 8px; }
  .chat-msg-user { align-self: flex-end; background: var(--pink); color: #fff; border-radius: 12px 12px 2px 12px; padding: 8px 12px; font-size: 0.74rem; max-width: 86%; line-height: 1.5; }
  .chat-msg-ai { align-self: flex-start; background: var(--surface3); border: 1px solid var(--border); border-radius: 2px 12px 12px 12px; padding: 8px 12px; font-size: 0.74rem; max-width: 92%; line-height: 1.6; color: var(--text); }
  .chat-msg-ai strong { color: var(--pink); }
  .chat-suggested { display: flex; flex-direction: column; gap: 6px; padding: 8px 0; }
  .chat-sug-btn { background: var(--surface3); border: 1px solid var(--border2); border-radius: 8px; color: var(--muted2); font-size: 0.68rem; padding: 7px 10px; text-align: left; cursor: pointer; transition: border-color 0.15s, color 0.15s; }
  .chat-sug-btn:hover { border-color: var(--pink); color: var(--text); }
  .chat-input-row { display: flex; gap: 6px; padding: 8px 10px; border-top: 1px solid var(--border2); align-items: flex-end; }
  .chat-input-row textarea { flex: 1; background: var(--surface3); border: 1px solid var(--border2); border-radius: 8px; color: var(--text); font-size: 0.73rem; padding: 7px 10px; resize: none; line-height: 1.4; font-family: var(--font-ui); }
  .chat-send { background: var(--pink); border: none; border-radius: 8px; color: #fff; width: 36px; height: 36px; font-size: 1rem; cursor: pointer; flex-shrink: 0; transition: opacity 0.15s; }
  .chat-send:disabled { opacity: 0.4; cursor: not-allowed; }
  .chat-cursor { display: inline-block; animation: blink 0.7s step-end infinite; }
  @keyframes blink { 0%,100%{opacity:1} 50%{opacity:0} }
  .compare-better { color: var(--green); font-weight: 700; } .compare-worse { color: var(--red); }
  .pin-btn { float: right; font-size: 0.60rem; padding: 4px 10px; margin-top: -2px; border-radius: 6px; min-height: 28px; }

  /* Roth Ladder */
  .rl-layout { display: grid; grid-template-columns: 280px 1fr; gap: 20px; align-items: start; }
  @media (max-width: 768px) { .rl-layout { grid-template-columns: 1fr; } }
  .rl-inputs { display: flex; flex-direction: column; gap: 10px; }
  .rl-result-hint { color: var(--muted2); font-size: 0.78rem; text-align: center; padding: 60px 0; }

  /* ── Home / Landing ── */
  .home-hero {
    display: flex; flex-direction: column; align-items: center; justify-content: center;
    padding: 28px 0 20px; gap: 6px; position: relative;
  }
  .fi-ring-wrap { position: relative; width: 164px; height: 164px; flex-shrink: 0; }
  .fi-ring-wrap svg { transform: rotate(-90deg); }
  .fi-ring-track { fill: none; stroke: rgba(255,255,255,0.07); }
  .fi-ring-fill  { fill: none; stroke: url(#ringGrad); stroke-linecap: round; transition: stroke-dashoffset 1.2s cubic-bezier(.4,0,.2,1); }
  .fi-ring-label {
    position: absolute; inset: 0; display: flex; flex-direction: column;
    align-items: center; justify-content: center; gap: 1px;
  }
  .fi-ring-pct   { font-size: 2rem; font-weight: 800; color: var(--text); font-family: var(--font-mono); line-height: 1; }
  .fi-ring-sub   { font-size: 0.58rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.12em; color: var(--muted2); }
  .home-lnw      { font-size: 2.2rem; font-weight: 800; color: var(--text); letter-spacing: -0.02em; }
  .home-lnw-sub  { font-size: 0.72rem; color: var(--muted); display: flex; gap: 10px; align-items: center; }
  .home-lnw-sub .sep { color: var(--border2); }

  .home-month-bars { display: flex; gap: 10px; margin-top: 4px; }
  .hmb-item { flex: 1; display: flex; flex-direction: column; gap: 4px; }
  .hmb-label { font-size: 0.58rem; text-transform: uppercase; letter-spacing: 0.08em; font-weight: 700; }
  .hmb-label.inc { color: var(--green); } .hmb-label.exp { color: var(--red); } .hmb-label.inv { color: var(--amber); }
  .hmb-bar-track { height: 4px; background: rgba(255,255,255,0.07); border-radius: 2px; overflow: hidden; }
  .hmb-bar-fill  { height: 4px; border-radius: 2px; transition: width 0.8s ease; }
  .hmb-fill-inc  { background: var(--green); } .hmb-fill-exp { background: var(--red); } .hmb-fill-inv { background: var(--amber); }
  .hmb-val { font-size: 0.78rem; font-family: var(--font-mono); font-weight: 600; }

  .home-nav-grid { display: grid; grid-template-columns: repeat(3,1fr); gap: 8px; }
  .home-nav-tile {
    background: var(--surface2); border: 1px solid var(--border); border-radius: var(--radius);
    padding: 14px 10px; display: flex; flex-direction: column; align-items: center; gap: 6px;
    cursor: pointer; transition: background 0.15s, border-color 0.15s;
    font-size: 0.65rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.06em;
    color: var(--muted); -webkit-tap-highlight-color: transparent; user-select: none;
  }
  .home-nav-tile:hover  { background: var(--surface3); border-color: var(--border2); color: var(--text); }
  .home-nav-tile:active { opacity: 0.7; }
  .home-nav-tile svg    { color: var(--pink); }

  .home-tx-row { display: flex; align-items: center; gap: 10px; padding: 9px 0; border-bottom: 1px solid var(--border); }
  .home-tx-row:last-child { border-bottom: none; }
  .home-tx-dot  { width: 8px; height: 8px; border-radius: 50%; flex-shrink: 0; }
  .home-tx-info { flex: 1; min-width: 0; }
  .home-tx-cat  { font-size: 0.70rem; font-weight: 600; color: var(--text); white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
  .home-tx-memo { font-size: 0.62rem; color: var(--muted2); white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
  .home-tx-amt  { font-family: var(--font-mono); font-size: 0.78rem; font-weight: 700; flex-shrink: 0; }
  .home-tx-date { font-size: 0.60rem; color: var(--muted2); flex-shrink: 0; }

  /* ── Loading skeletons ── */
  .skeleton { background: linear-gradient(90deg, var(--surface2) 25%, rgba(255,255,255,0.04) 50%, var(--surface2) 75%);
    background-size: 200% 100%; border-radius: var(--radius-sm); animation: shimmer 1.8s ease-in-out infinite; }
  @keyframes shimmer { 0% { background-position:200% 0; } 100% { background-position:-200% 0; } }
  .skeleton-wrap { display: flex; flex-direction: column; gap: 12px; padding: 4px 0; }
  .sk-line { height: 14px; }
  .sk-line.wide { width: 100%; } .sk-line.med { width: 65%; } .sk-line.short { width: 40%; }
  .sk-card { height: 88px; border-radius: var(--radius); }
  .sk-chart { height: 200px; border-radius: var(--radius); }

  /* ── Mobile overrides ── */
  @media (max-width: 768px) {
    main { max-width: 100%; padding: 16px 14px; padding-bottom: calc(var(--bottom-nav-h) + 16px + env(safe-area-inset-bottom,0px)); gap: 14px; }
    .stat-grid { grid-template-columns: repeat(2, 1fr); }
    .panel, .glass { padding: 16px; }
    .success-badge .sbval { font-size: 2.6rem; }
    .mc-stats-strip { grid-template-columns: repeat(2, 1fr); }
    header { padding: 0 14px; }
    .spending-table { font-size: 0.66rem; }
    .spending-table td, .spending-table th { padding: 5px 4px; }
  }

  /* ── Budget vs Actual ── */
  .budget-table { width:100%; border-collapse:collapse; font-size:0.76rem; margin-top:8px; }
  .budget-table th { text-align:left; color:var(--muted); font-weight:500; font-size:0.68rem; text-transform:uppercase; letter-spacing:.04em; padding:6px 8px; border-bottom:1px solid var(--border); }
  .budget-table td { padding:6px 8px; border-bottom:1px solid var(--border); }
  .delta-over  { color:var(--red); font-weight:600; }
  .delta-under { color:var(--green); font-weight:600; }
  .budget-edit-grid { display:grid; grid-template-columns:repeat(auto-fill,minmax(160px,1fr)); gap:8px; margin-bottom:10px; }

  /* ── Scenario save/load ── */
  .scenario-list { display:flex; flex-direction:column; gap:6px; }
  .scenario-item { display:flex; align-items:center; gap:8px; padding:8px 10px; background:var(--surface3); border-radius:var(--radius-sm); font-size:0.74rem; }
  .scenario-item .sn-name { flex:1; font-weight:600; }
  .scenario-item .sn-date { color:var(--muted2); font-size:0.65rem; }
  .scenario-item .sn-pct  { font-family:var(--font-mono); color:var(--green); font-weight:700; margin-right:4px; }

  /* ── Contribution optimizer result ── */
  .contrib-result { background:var(--surface2); border:1px solid var(--border2); border-radius:var(--radius-sm); padding:12px 14px; margin-top:10px; font-size:0.78rem; }
  .contrib-result .cr-val { font-family:var(--font-mono); font-size:1.2rem; font-weight:700; color:var(--pink); display:block; margin:4px 0; }

  /* ── Sensitivity tornado ── */
  .tornado-wrap { margin-top:14px; padding:16px 18px; }
  .tornado-wrap canvas { max-height:260px; }

  /* ── Ruin table ── */
  .ruin-table { width:100%; border-collapse:collapse; font-size:0.76rem; }
  .ruin-table th { text-align:left; color:var(--muted); font-weight:500; font-size:0.68rem; text-transform:uppercase; letter-spacing:.04em; padding:6px 8px; border-bottom:1px solid var(--border); }
  .ruin-table td { padding:7px 8px; border-bottom:1px solid var(--border); }
  .ruin-table td:last-child { font-family:var(--font-mono); font-weight:700; text-align:right; }

  /* ── Email digest panel ── */
  .digest-grid { display:grid; grid-template-columns:repeat(auto-fill,minmax(200px,1fr)); gap:8px; margin-bottom:10px; }
  .digest-note { font-size:0.65rem; color:var(--muted2); margin-top:6px; }

  /* ── Print ── */
  @media print {
    .tab-nav, .bottom-nav, .mc-inputs, .port-refresh-bar, #budgetEditForm,
    #savedScenariosList, .pin-btn, .mc-presets, .mc-advanced,
    #mc_run_btn, #mc_spinner, .contrib-opt-btn, #sensitivityBtn,
    .digest-panel { display:none !important; }
    .tab-pane { display:none !important; }
    .tab-pane.active { display:block !important; }
    body { background:#fff; color:#111; }
    .panel, .glass { background:#fff !important; border:1px solid #ccc !important;
      box-shadow:none !important; color:#111 !important; }
    .sc-value, .sbval, .clr-green, .clr-amber, .clr-red { color:#111 !important; }
    header { background:#fff; border-bottom:2px solid #333; position:relative; }
    .header-logo, .bottom-nav { display:none !important; }
    main { padding:12px 0; }
  }
</style>
</head>
<body>
<header>
  <div class="header-brand">
    <div class="header-logo">
      <svg width="16" height="16" fill="none" viewBox="0 0 24 24" stroke="white" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
        <path d="M12 2L2 7l10 5 10-5-10-5zM2 17l10 5 10-5M2 12l10 5 10-5"/>
      </svg>
    </div>
    <h1>Road To <span>FI</span></h1>
  </div>
  <div style="display:flex;align-items:center;gap:10px;">
    <button class="btn-sm" onclick="window.print()" style="font-size:0.62rem;padding:4px 10px;">Export PDF</button>
    <span class="refresh-ts" id="refreshTs"></span>
  </div>
</header>

<!-- Desktop Tab navigation -->
<nav class="tab-nav">
  <button class="tab-btn active" data-tab="home" onclick="switchTab('home',this)">Home</button>
  <button class="tab-btn" data-tab="ledger" onclick="switchTab('ledger',this)">Ledger</button>
  <button class="tab-btn" data-tab="portfolio" onclick="switchTab('portfolio',this)">Portfolio</button>
  <button class="tab-btn" data-tab="montecarlo" onclick="switchTab('montecarlo',this)">Monte Carlo</button>
  <button class="tab-btn" data-tab="rothlad" onclick="switchTab('rothlad',this)">Roth Ladder</button>
  <button class="tab-btn" data-tab="rules" onclick="switchTab('rules',this)">Rules</button>
  <button class="tab-btn" data-tab="roadmap" onclick="switchTab('roadmap',this)">Roadmap</button>
  <button class="tab-btn" data-tab="transactions" onclick="switchTab('transactions',this)">Transactions</button>
  <button class="tab-btn" data-tab="forecast" onclick="switchTab('forecast',this)">Forecast</button>
  <button class="tab-btn" data-tab="chat" onclick="switchTab('chat',this)">&#129302; AI Advisor</button>
</nav>

<!-- Mobile bottom navigation -->
<nav class="bottom-nav">
  <button class="bn-item active" data-tab="home" onclick="switchTab('home',this)">
    <svg width="20" height="20" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M3 12l2-2m0 0l7-7 7 7M5 10v10a1 1 0 001 1h3m10-11l2 2m-2-2v10a1 1 0 01-1 1h-3m-6 0a1 1 0 001-1v-4a1 1 0 011-1h2a1 1 0 011 1v4a1 1 0 001 1m-6 0h6"/></svg>
    Home
  </button>
  <button class="bn-item" data-tab="ledger" onclick="switchTab('ledger',this)">
    <svg width="20" height="20" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M9 17v-2m3 2v-4m3 4v-6m2 10H7a2 2 0 01-2-2V5a2 2 0 012-2h5.586a1 1 0 01.707.293l5.414 5.414a1 1 0 01.293.707V19a2 2 0 01-2 2z"/></svg>
    Ledger
  </button>
  <button class="bn-item" data-tab="portfolio" onclick="switchTab('portfolio',this)">
    <svg width="20" height="20" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M9 19v-6a2 2 0 00-2-2H5a2 2 0 00-2 2v6a2 2 0 002 2h2a2 2 0 002-2zm0 0V9a2 2 0 012-2h2a2 2 0 012 2v10m-6 0a2 2 0 002 2h2a2 2 0 002-2m0 0V5a2 2 0 012-2h2a2 2 0 012 2v14a2 2 0 01-2 2h-2a2 2 0 01-2-2z"/></svg>
    Portfolio
  </button>
  <button class="bn-item" data-tab="montecarlo" onclick="switchTab('montecarlo',this)">
    <svg width="20" height="20" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M13 7h8m0 0v8m0-8l-8 8-4-4-6 6"/></svg>
    Sim
  </button>
  <button class="bn-item" data-tab="rothlad" onclick="switchTab('rothlad',this)">
    <svg width="20" height="20" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M4 4v5h.582m15.356 2A8.001 8.001 0 004.582 9m0 0H9m11 11v-5h-.581m0 0a8.003 8.003 0 01-15.357-2m15.357 2H15"/></svg>
    Roth
  </button>
  <button class="bn-item" data-tab="rules" onclick="switchTab('rules',this)">
    <svg width="20" height="20" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M9 5H7a2 2 0 00-2 2v12a2 2 0 002 2h10a2 2 0 002-2V7a2 2 0 00-2-2h-2M9 5a2 2 0 002 2h2a2 2 0 002-2M9 5a2 2 0 012-2h2a2 2 0 012 2m-3 7h3m-3 4h3m-6-4h.01M9 16h.01"/></svg>
    Rules
  </button>
  <button class="bn-item" data-tab="roadmap" onclick="switchTab('roadmap',this)">
    <svg width="20" height="20" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M9 20l-5.447-2.724A1 1 0 013 16.382V5.618a1 1 0 011.447-.894L9 7m0 13l6-3m-6 3V7m6 10l4.553 2.276A1 1 0 0021 18.382V7.618a1 1 0 00-.553-.894L15 4m0 13V4m0 0L9 7"/></svg>
    Roadmap
  </button>
  <button class="bn-item" data-tab="transactions" onclick="switchTab('transactions',this)">
    <svg width="20" height="20" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M3 10h18M7 15h1m4 0h1m-7 4h12a3 3 0 003-3V8a3 3 0 00-3-3H6a3 3 0 00-3 3v8a3 3 0 003 3z"/></svg>
    Ledger
  </button>
  <button class="bn-item" data-tab="forecast" onclick="switchTab('forecast',this)">
    <svg width="20" height="20" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M8 7V3m8 4V3m-9 8h10M5 21h14a2 2 0 002-2V7a2 2 0 00-2-2H5a2 2 0 00-2 2v12a2 2 0 002 2z"/></svg>
    Forecast
  </button>
  <button class="bn-item" data-tab="taxloss" onclick="switchTab('taxloss',this)">
    <svg width="20" height="20" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M9 14l6-6m-5.5.5h.01m4.99 5h.01M19 21V5a2 2 0 00-2-2H7a2 2 0 00-2 2v16l3.5-2 3.5 2 3.5-2 3.5 2z"/></svg>
    Tax Loss
  </button>
  <button class="bn-item" data-tab="chat" onclick="switchTab('chat',this)">
    <svg width="20" height="20" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M8 10h.01M12 10h.01M16 10h.01M9 16H5a2 2 0 01-2-2V6a2 2 0 012-2h14a2 2 0 012 2v8a2 2 0 01-2 2h-5l-5 5v-5z"/></svg>
    AI Advisor
  </button>
</nav>

<!-- ══ HOME TAB ════════════════════════════════════════════════════════════ -->
<div id="tab-home" class="tab-pane active">
<main>
  <div id="homeContent">
    <div class="skeleton-wrap">
      <div class="skeleton sk-chart" style="height:180px;border-radius:var(--radius);"></div>
      <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;">
        <div class="skeleton sk-card"></div><div class="skeleton sk-card"></div><div class="skeleton sk-card"></div>
      </div>
    </div>
  </div>
</main>
</div>

<!-- ══ LEDGER TAB ══════════════════════════════════════════════════════════ -->
<div id="tab-ledger" class="tab-pane">
<main>
  <div class="panel accent-pink">
    <h2>FI Dashboard Snapshot <button class="btn-sm" onclick="loadLedger()">Refresh</button></h2>
    <div id="ledgerContent"><div style="color:var(--muted2);font-size:0.8rem;">Loading ledger&hellip;</div></div>
    <div id="allocChartWrap" style="display:none;">
      <div class="donut-wrap">
        <canvas id="allocChart" width="180" height="180" style="max-width:180px;"></canvas>
        <div class="donut-legend" id="allocLegend"></div>
      </div>
    </div>
  </div>

  <!-- Net Worth History -->
  <div id="nwhPanel" style="display:none;margin-top:14px;">
    <div class="panel accent-sky">
      <h2>Net Worth History <button class="btn-sm" onclick="clearNWHistory()" style="font-size:0.6rem;">Clear</button></h2>
      <canvas id="nwhChart" style="max-height:200px;"></canvas>
    </div>
  </div>

  <!-- Budget vs Actual -->
  <div class="panel accent-amber" style="margin-top:14px;" id="budgetPanel">
    <h2>Budget vs Actual <button class="btn-sm" onclick="toggleBudgetEdit()">Edit Budgets</button></h2>
    <div id="budgetEditForm" style="display:none;margin-bottom:12px;">
      <div class="budget-edit-grid" id="budgetInputs"></div>
      <button class="btn-sm btn-primary" onclick="saveBudget()">Save Budget</button>
    </div>
    <div id="budgetTable"><div style="color:var(--muted2);font-size:0.75rem;">Budgets auto-populate once Ledger loads.</div></div>
  </div>

  <!-- Email Digest Config -->
  <div class="panel digest-panel" style="margin-top:14px;">
    <h2>Weekly Email Digest</h2>
    <div class="digest-grid" id="digestForm">
      <div class="field"><label>Gmail / SMTP User</label><input id="dig_user" type="email" placeholder="you@gmail.com"></div>
      <div class="field"><label>App Password</label><input id="dig_pass" type="password" placeholder="Google App Password"></div>
      <div class="field"><label>Send To</label><input id="dig_to" type="email" placeholder="same as above"></div>
    </div>
    <div style="display:flex;align-items:center;gap:10px;margin-top:8px;">
      <button class="btn-sm btn-primary" onclick="sendDigest()">Send Now</button>
      <span id="digestStatus" style="font-size:0.72rem;color:var(--muted2);"></span>
    </div>
    <div class="digest-note">For Gmail: enable 2FA then create an App Password at myaccount.google.com &rarr; Security. Credentials saved locally in your browser.</div>
  </div>
</main>
</div>

<!-- ══ PORTFOLIO TAB ═══════════════════════════════════════════════════════ -->
<div id="tab-portfolio" class="tab-pane">
<main>
  <div class="panel accent-teal">
    <h2>Holdings <button class="btn-sm" onclick="loadPortfolio()">Reload from Ledger</button></h2>

    <!-- Summary stat-cards -->
    <div class="port-summary">
      <div class="stat-card"><div class="sc-label">Total Value</div><div class="sc-value" id="ps_value">—</div><div class="sc-sub" id="ps_asof">Cached from ledger</div></div>
      <div class="stat-card"><div class="sc-label">Cost Basis</div><div class="sc-value" id="ps_cost">—</div><div class="sc-sub">Avg cost × shares</div></div>
      <div class="stat-card" id="ps_gl_card"><div class="sc-label">Total Gain / Loss</div><div class="sc-value" id="ps_gl">—</div><div class="sc-sub" id="ps_gl_pct">—</div></div>
      <div class="stat-card"><div class="sc-label">Holdings</div><div class="sc-value" id="ps_count">—</div><div class="sc-sub">across accounts</div></div>
    </div>

    <!-- Holdings table -->
    <div id="portTable"><div style="color:var(--muted2);font-size:0.8rem;">Loading holdings&hellip;</div></div>

    <!-- Refresh bar -->
    <div class="port-refresh-bar">
      <div class="field" style="min-width:220px;">
        <label>Alpha Vantage API Key</label>
        <input id="portApiKey" type="password" placeholder="your key">
      </div>
      <button onclick="refreshPortfolio()">Refresh Live Prices</button>
      <div class="key-status" id="portKeyStatus"></div>
    </div>
    <div id="portRefreshStatus" style="font-size:0.72rem;color:var(--muted2);margin-top:6px;"></div>
  </div>
</main>
</div>

<!-- ══ MONTE CARLO TAB ════════════════════════════════════════════════════ -->
<div id="tab-montecarlo" class="tab-pane">
<main>
  <div class="panel accent-indigo">
    <h2>Monte Carlo Retirement Simulation</h2>
    <div class="mc-layout">

      <!-- LEFT: Inputs -->
      <div class="mc-inputs">

        <!-- Scenario presets -->
        <div class="mc-presets">
          <button class="preset-pill baseline" onclick="applyPreset('baseline')">Baseline</button>
          <button class="preset-pill cautious" onclick="applyPreset('cautious')">Cautious</button>
          <button class="preset-pill stress"   onclick="applyPreset('stress')">Stress Test</button>
          <button class="preset-pill optimist" onclick="applyPreset('optimist')">Optimist</button>
        </div>

        <div class="mc-section-title">Core Parameters</div>
        <div class="mc-row-3">
          <div class="field"><label>Current Age</label><input type="number" id="mc_age" value="45" min="20" max="70"></div>
          <div class="field"><label>Retire Age</label><input type="number" id="mc_ret_age" value="62" min="50" max="75"></div>
          <div class="field"><label>Filing Status</label>
            <select id="mc_filing" class="mc-select">
              <option value="single">Single</option>
              <option value="mfj">Married (MFJ)</option>
            </select>
          </div>
        </div>

        <div class="mc-section-title">Portfolio Balances</div>
        <div class="mc-row-3">
          <div class="field"><label>Engine (Market)</label><input type="number" id="mc_engine" value="0" min="0"></div>
          <div class="field"><label>SGOV Moat ($)</label><input type="number" id="mc_sgov" value="0" min="0"></div>
          <div class="field"><label>Checking ($)</label><input type="number" id="mc_checking" value="5000" min="0"></div>
        </div>
        <div class="mc-row-2">
          <div class="field"><label>Annual Contribution ($)</label><input type="number" id="mc_contrib" value="0" min="0"></div>
          <div class="field"><label>SS Benefit @67 ($/yr)</label><input type="number" id="mc_full_ss" value="25620" min="0"></div>
        </div>

        <p class="mc-defaults-note">&#10003; Smart defaults active &mdash; expand Advanced to customize</p>

        <details class="mc-advanced">
          <summary>&#9881; Advanced Settings</summary>
          <div class="adv-body">

            <div class="mc-section-title">Return Assumptions</div>
            <div class="mc-row-3">
              <div class="field"><label>Mean Return (%)</label><input type="number" id="mc_mu" value="9" step="0.1" min="1" max="20"></div>
              <div class="field"><label>Volatility (%)</label><input type="number" id="mc_sigma" value="16" step="0.1" min="1" max="40"></div>
              <div class="field"><label>SGOV Yield (%)</label><input type="number" id="mc_syld" value="4" step="0.1" min="0" max="15"></div>
            </div>
            <div class="mc-row-3">
              <div class="field"><label>Inflation (%)</label><input type="number" id="mc_infl" value="3" step="0.1" min="0" max="15"></div>
              <div class="field"><label>Annual Floor Draw ($)</label><input type="number" id="mc_floor" value="54292" min="0"></div>
              <div class="field"><label>Moat Target ($)</label><input type="number" id="mc_moat_target" value="360000" min="0"></div>
            </div>

            <div class="mc-section-title">Roth Conversion</div>
            <div class="mc-toggle-row">
              <label class="mc-toggle"><input type="checkbox" id="mc_use_conv" onchange="document.getElementById('mc_conv_opts').style.display=this.checked?'grid':'none'"> Enable Roth Conversion</label>
            </div>
            <div id="mc_conv_opts" class="mc-row-3" style="display:none">
              <div class="field"><label>Trad Balance ($)</label><input type="number" id="mc_trad" value="0" min="0"></div>
              <div class="field"><label>Target Bracket</label>
                <select id="mc_tgt_bkt" class="mc-select">
                  <option value="0.12">12%</option>
                  <option value="0.22">22%</option>
                </select>
              </div>
              <div class="field"><label>State Tax (%)</label><input type="number" id="mc_state_tx" value="3.99" step="0.01" min="0" max="15"></div>
            </div>

            <div class="mc-section-title">Options</div>
            <div class="mc-toggle-row">
              <label class="mc-toggle"><input type="checkbox" id="mc_haircut"> SS Haircut (21%)</label>
              <label class="mc-toggle"><input type="checkbox" id="mc_tail"> Tail Shock</label>
              <label class="mc-toggle"><input type="checkbox" id="mc_mort" checked> Mortality Weight</label>
            </div>

            <div class="mc-section-title">Stochastic Inflation</div>
            <div class="mc-toggle-row">
              <label class="mc-toggle"><input type="checkbox" id="mc_use_si" onchange="document.getElementById('mc_si_opts').style.display=this.checked?'grid':'none'"> Variable Inflation</label>
            </div>
            <div id="mc_si_opts" class="mc-row-3" style="display:none">
              <div class="field"><label>Infl. Vol (%)</label><input type="number" id="mc_infl_vol" value="1" step="0.1" min="0" max="5"></div>
              <div class="field"><label>Min (%)</label><input type="number" id="mc_infl_min" value="1" step="0.1" min="0" max="5"></div>
              <div class="field"><label>Max (%)</label><input type="number" id="mc_infl_max" value="8" step="0.1" min="2" max="20"></div>
              <div class="field"><label>Stagflation Corr (%)</label><input type="number" id="mc_stag_corr" value="30" step="5" min="0" max="100"></div>
            </div>

            <div class="mc-section-title">Macro Shocks</div>
            <div class="mc-toggle-row">
              <label class="mc-toggle"><input type="checkbox" id="mc_use_aca" onchange="document.getElementById('mc_aca_opts').style.display=this.checked?'grid':'none'"> ACA Subsidy Shock</label>
              <label class="mc-toggle"><input type="checkbox" id="mc_use_taxrev" onchange="document.getElementById('mc_taxrev_opts').style.display=this.checked?'grid':'none'"> Tax Legislative Risk</label>
            </div>
            <div id="mc_aca_opts" class="mc-row-3" style="display:none">
              <div class="field"><label>Annual Prob (%)</label><input type="number" id="mc_aca_prob" value="30" step="5" min="0" max="100"></div>
              <div class="field"><label>Shock Amount ($)</label><input type="number" id="mc_aca_mag" value="15000" step="1000" min="0"></div>
            </div>
            <div id="mc_taxrev_opts" class="mc-row-3" style="display:none">
              <div class="field"><label>Near-term Prob (%)</label><input type="number" id="mc_tax_near" value="20" step="5" min="0" max="100"></div>
              <div class="field"><label>Mid-term Prob (%)</label><input type="number" id="mc_tax_mid" value="40" step="5" min="0" max="100"></div>
              <div class="field"><label>Long-term Prob (%)</label><input type="number" id="mc_tax_late" value="60" step="5" min="0" max="100"></div>
            </div>
            <div style="font-size:0.62rem;color:var(--muted2);margin-top:3px;">ACA: random premium spike ages 62&ndash;64. Tax Risk: TCJA expiry raises conversion taxes ~13%.</div>

            <div class="mc-section-title">AI Model</div>
            <div class="field" style="margin-bottom:4px;">
              <label>Ollama Model</label>
              <input type="text" id="aiModel" value="llama3.1:8b" placeholder="llama3.1:8b" style="font-family:var(--font-mono);font-size:0.72rem;">
            </div>
            <div style="font-size:0.61rem;color:var(--muted2);margin-bottom:4px;">Used for AI Playbook and chat. Set OLLAMA_URL env var if not on host network.</div>

            <div class="mc-section-title">Sim Size</div>
            <div class="mc-radio-row" id="simSizeBtns" style="gap:4px;">
              <label><input type="radio" name="mc_sim_size" value="1k" checked> 1K</label>
              <label><input type="radio" name="mc_sim_size" value="10k"> 10K</label>
              <label><input type="radio" name="mc_sim_size" value="100k"> 100K</label>
              <label><input type="radio" name="mc_sim_size" value="1m"> 1M</label>
            </div>
            <div style="font-size:0.61rem;color:var(--muted2);margin-bottom:4px;">Higher = more accurate, slower. 1M uses full server CPU.</div>

            <div class="mc-section-title">Market Model</div>
            <div class="mc-radio-row">
              <label><input type="radio" name="mc_return_model" value="normal" checked> Normal</label>
              <label><input type="radio" name="mc_return_model" value="fat_tail"> Fat Tail</label>
              <label><input type="radio" name="mc_return_model" value="regime_switch"> Regime</label>
              <label><input type="radio" name="mc_return_model" value="garch"> GARCH</label>
            </div>
            <div style="font-size:0.61rem;color:var(--muted2);margin-bottom:4px;">Fat Tail: heavier crash risk &bull; Regime: bull/bear cycles &bull; GARCH: volatility clustering</div>

            <div class="mc-section-title">Sequence-of-Returns Stress</div>
            <div class="mc-radio-row">
              <label><input type="radio" name="mc_seq" value="0" checked> None</label>
              <label><input type="radio" name="mc_seq" value="1"> Year 1</label>
              <label><input type="radio" name="mc_seq" value="3"> Year 3</label>
              <label><input type="radio" name="mc_seq" value="5"> Year 5</label>
            </div>
            <div style="font-size:0.62rem;color:var(--muted2);margin-top:3px;">Forces a &minus;25% return in that retirement year across all trials</div>

          </div>
        </details>

        <div id="mc_prefill_note" style="display:none;font-size:0.62rem;color:var(--green);margin-top:8px;padding:6px 10px;border:1px solid rgba(16,185,129,0.25);border-radius:6px;">
          &#10003; Balances &amp; SS benefit auto-filled from Road To FI. Verify before running.
        </div>
        <button class="btn-primary" id="mc_run_btn" onclick="runMC()" style="width:100%;margin-top:12px;">Run Simulation</button>
        <button class="btn-sm contrib-opt-btn" onclick="optimizeContribution()" style="width:100%;margin-top:6px;">&#128269; Find Optimal Contribution (95%)</button>
        <div id="mc_spinner" style="display:none;text-align:center;padding-top:8px;color:var(--muted2);font-size:0.72rem;">Simulating&hellip;</div>
        <div id="contribResult"></div>

        <!-- Saved Scenarios -->
        <div class="panel" style="margin-top:14px;padding:14px 16px;">
          <h2 style="font-size:0.82rem;margin-bottom:10px;">Saved Scenarios</h2>
          <div style="display:flex;gap:8px;margin-bottom:10px;align-items:center;">
            <input id="scenNameInput" type="text" placeholder="Scenario name&hellip;" style="flex:1;padding:6px 10px;background:var(--surface3);border:1px solid var(--border2);border-radius:var(--radius-sm);color:var(--text);font-size:0.76rem;">
            <button class="btn-sm" onclick="saveScenario()">Save</button>
          </div>
          <div id="savedScenariosList" class="scenario-list"></div>
        </div>
      </div>

      <!-- RIGHT: Results -->
      <div id="mc_results_panel">
        <div class="mc-result-hint">Enter parameters and click Run Simulation</div>
      </div>

    </div><!-- .mc-layout -->
  </div>
</main>
</div>

<!-- ══ RULES TAB ══════════════════════════════════════════════════════════ -->
<div id="tab-rules" class="tab-pane">
<main>
  <div class="panel accent-amber">
    <h2>2026 Contribution Limits</h2>
    <div class="rules-grid" id="rulesGrid">
      <div class="stat-card"><div class="sc-label">Loading&hellip;</div></div>
    </div>
    <div class="note" id="rothNote"></div>
  </div>
</main>
</div>

<!-- ══ ROTH LADDER TAB ══════════════════════════════════════════════════════ -->
<div id="tab-rothlad" class="tab-pane">
<main>
  <div class="panel accent-emerald" style="margin-bottom:16px;">
    <h2>Roth Conversion Ladder</h2>
    <div class="rl-layout">
      <div class="rl-inputs">
        <div class="mc-field"><label>Traditional IRA Balance</label><input type="number" id="rl_trad" value="0" min="0" step="1000"></div>
        <div class="mc-field"><label>Current Age</label><input type="number" id="rl_age" value="45" min="20" max="80"></div>
        <div class="mc-field"><label>Retirement Age</label><input type="number" id="rl_ret" value="62" min="40" max="75"></div>
        <div class="mc-field"><label>Filing Status</label>
          <select id="rl_filing">
            <option value="single">Single</option>
            <option value="mfj" selected>Married Filing Jointly</option>
          </select>
        </div>
        <div class="mc-field"><label>SS Monthly Benefit @67</label><input type="number" id="rl_ss" value="1800" min="0"></div>
        <div class="mc-field"><label>State Tax Rate %</label><input type="number" id="rl_state" value="0" min="0" max="15" step="0.1"></div>
        <div class="mc-field"><label>Target Bracket</label>
          <select id="rl_bracket">
            <option value="0.12">12%</option>
            <option value="0.22" selected>22%</option>
          </select>
        </div>
        <div class="mc-field"><label>Annual Return on Roth %</label><input type="number" id="rl_ret_rate" value="8" min="0" max="20" step="0.5"></div>
        <button class="btn" onclick="calcRothLadder()" style="width:100%;margin-top:8px;">Calculate</button>
      </div>
      <div id="rl_results">
        <div class="rl-result-hint">Enter your Traditional IRA details and click Calculate.</div>
      </div>
    </div>
  </div>
</main>
</div>

<!-- ══ ROADMAP TAB ══════════════════════════════════════════════════════════ -->
<div id="tab-roadmap" class="tab-pane">
<main>
  <div class="panel accent-teal">
    <h2>FI Roadmap — GPS to Age 62</h2>
    <div id="roadmapContent"><div style="color:var(--muted2);font-size:0.8rem;">Loading roadmap&hellip;</div></div>
  </div>
</main>
</div>

<!-- ══ TRANSACTIONS TAB ════════════════════════════════════════════════════ -->
<div id="tab-transactions" class="tab-pane">
<main>
  <div class="panel accent-indigo">
    <h2>Transaction Ledger</h2>
    <div id="txFilters" style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;">
      <select id="txMonthFilter" onchange="loadTransactions(1)" style="flex:1;min-width:120px;background:var(--surface3);border:1px solid var(--border2);color:var(--text);border-radius:6px;padding:6px 8px;font-size:0.72rem;">
        <option value="">All Months</option>
      </select>
      <select id="txTypeFilter" onchange="loadTransactions(1)" style="flex:1;min-width:120px;background:var(--surface3);border:1px solid var(--border2);color:var(--text);border-radius:6px;padding:6px 8px;font-size:0.72rem;">
        <option value="">All Types</option>
      </select>
    </div>
    <div id="txContent"><div style="color:var(--muted2);font-size:0.8rem;">Loading transactions&hellip;</div></div>
    <div id="txPager" style="display:flex;align-items:center;justify-content:space-between;margin-top:12px;font-size:0.72rem;color:var(--muted2);"></div>
  </div>
</main>
</div>

<!-- ══ FORECAST TAB ════════════════════════════════════════════════════════ -->
<div id="tab-forecast" class="tab-pane">
<main>
  <div class="panel accent-amber">
    <h2>30-Day Cash Flow Forecast</h2>
    <div id="forecastContent"><div style="color:var(--muted2);font-size:0.8rem;">Loading forecast&hellip;</div></div>
  </div>
</main>
</div>

<!-- ══ TAX LOSS TAB ═════════════════════════════════════════════════════════ -->
<div id="tab-taxloss" class="tab-pane">
<main>
  <div class="panel accent-rose" style="margin-bottom:14px;">
    <h2>Tax-Loss Harvesting Log</h2>
    <div class="tl-summary" id="tlSummary">
      <div class="stat-card"><div class="sc-label">Loading&hellip;</div></div>
    </div>
    <div id="tlContent"><div style="color:var(--muted2);font-size:0.8rem;">Loading&hellip;</div></div>
  </div>
</main>
</div>

<script>
const fm = n => '$' + Number(n).toLocaleString('en-US', {minimumFractionDigits:0, maximumFractionDigits:0});
const fp = n => (n * 100).toFixed(1) + '%';
const mdLite = t => t.replace(/\*\*(.+?)\*\*/g,'<strong>$1</strong>').replace(/\*(.+?)\*/g,'<em>$1</em>').replace(/\\n/g,'<br>').replace(/^[\\u2713\\u26a0\\u2192\\u2022]\\s/gm, m => `<span style="color:var(--pink)">${m}</span>`);

// ── Tab switching ──────────────────────────────────────────────────────────
let mcPrefilled = false;
let portfolioLoaded = false;
let rulesLoaded = false;
let roadmapLoaded = false;
let transactionsLoaded = false;
let forecastLoaded = false;
let taxlossLoaded = false;
let homeLoaded = false;
let _txState = { page: 1, pages: 1 };
let scenarioA = null;
function switchTab(name, btn) {
  document.querySelectorAll('.tab-pane').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('[data-tab]').forEach(b => b.classList.remove('active'));
  document.getElementById('tab-' + name).classList.add('active');
  document.querySelectorAll('[data-tab="' + name + '"]').forEach(b => b.classList.add('active'));
  if (name === 'home'         && !homeLoaded)         loadHome();
  if (name === 'rules'        && !rulesLoaded)        loadRules();
  if (name === 'montecarlo'   && !mcPrefilled)        prefillMC();
  if (name === 'portfolio'    && !portfolioLoaded)    loadPortfolio();
  if (name === 'roadmap'      && !roadmapLoaded)      loadRoadmap();
  if (name === 'transactions' && !transactionsLoaded) loadTransactions(1);
  if (name === 'forecast'     && !forecastLoaded)     loadForecast();
  if (name === 'taxloss'      && !taxlossLoaded)      loadTaxLoss();
  if (name === 'chat') {
    if (chatMessages.length === 0) showChatSuggestions();
    setTimeout(() => document.getElementById('chatInput')?.focus(), 80);
  }
  if (name === 'rothlad') {
    // Pre-fill from ledger data if available
    const rl = document.getElementById('rl_trad');
    if (rl && !rl._prefilled) {
      rl._prefilled = true;
      fetch('/api/ledger/dashboard').then(r=>r.json()).then(d=>{
        if (d.mc_prefill) {
          const p = d.mc_prefill;
          if (p.trad_balance > 0) rl.value = p.trad_balance;
          if (p.full_ss_annual > 0) document.getElementById('rl_ss').value = Math.round(p.full_ss_annual/12);
        }
      }).catch(()=>{});
    }
  }
}

// ── Home / Landing ─────────────────────────────────────────────────────────
async function loadHome() {
  homeLoaded = true;
  const el = document.getElementById('homeContent');
  if (!el) return;
  try {
    const [dashRes, txRes] = await Promise.all([
      fetch('/api/ledger/dashboard'),
      fetch('/api/transactions?limit=5')
    ]);
    const d  = await dashRes.json();
    const tx = await txRes.json();
    if (d.error) { el.innerHTML = `<div style="color:#f87171;padding:16px;font-size:0.9rem;word-break:break-all"><b>API ERROR:</b><br>${d.error}</div>`; return; }
    if (d.error) { el.innerHTML = `<span class="err">${d.error}</span>`; return; }
    window._lastDashboard = d;  // make available for AI chat context

    const m       = d.metrics || {};
    const lnw     = m['LIQUID NET WORTH']   || 0;
    const tnw     = m['TOTAL NET WORTH']    || 0;
    const target  = m['FI TARGET (Age 62)'] || 0;
    const progress= m['PROGRESS TO FI']     || 0;
    const runway  = m['SURVIVAL RUNWAY']    || '—';
    const pct     = Math.min(progress * 100, 100);

    // ── FI ring (SVG arc) ──
    const R = 68; const C = 2 * Math.PI * R;
    const fill = C - (pct / 100) * C;

    // ── Monthly cash flow from latest spending month ──
    const spend  = d.spending || {};
    const months = d.spending_months || [];
    const latestIdx = months.length - 1;
    const income  = latestIdx >= 0 ? (spend['Income']   || [])[latestIdx] || 0 : 0;
    const expense = latestIdx >= 0 ? (spend['Expense']  || [])[latestIdx] || 0 : 0;
    const invest  = latestIdx >= 0 ? (spend['Investment']|| [])[latestIdx]|| 0 : 0;
    const srVals  = spend['SAVINGS RATE'] || [];
    const sr      = srVals.length ? srVals[srVals.length-1] : null;
    const maxFlow = Math.max(Math.abs(income), Math.abs(expense), Math.abs(invest), 1);
    const barW    = v => Math.round(Math.abs(v) / maxFlow * 100);

    // ── Recent transactions ──
    const txTypeColor = t => ({Income:'#10b981',Investment:'#38bdf8',Expense:'#f43f5e',Fun:'#f59e0b'}[t]||'#8899b8');
    const txRows = (tx.rows||[]).slice(0,5).map(r => `
      <div class="home-tx-row">
        <div class="home-tx-dot" style="background:${txTypeColor(r.type)}"></div>
        <div class="home-tx-info">
          <div class="home-tx-cat">${r.category}</div>
          <div class="home-tx-memo">${r.memo}</div>
        </div>
        <div class="home-tx-amt" style="color:${r.signed>=0?'var(--green)':'var(--red)'}">${r.signed>=0?'+':''}${fm(r.signed)}</div>
        <div class="home-tx-date">${r.date.slice(5)}</div>
      </div>`).join('');

    // ── Gap to target ──
    const gap      = Math.max(0, target - lnw);
    const monthStr = months[latestIdx] || '';

    el.innerHTML = `
      <div class="panel" style="padding:20px 18px 16px;border-color:rgba(236,72,153,0.25);">

        <!-- FI Ring + LNW hero -->
        <div class="home-hero">
          <div class="fi-ring-wrap">
            <svg width="164" height="164" viewBox="0 0 164 164">
              <defs>
                <linearGradient id="ringGrad" x1="0%" y1="0%" x2="100%" y2="0%">
                  <stop offset="0%" stop-color="#ec4899"/>
                  <stop offset="100%" stop-color="#6366f1"/>
                </linearGradient>
              </defs>
              <circle class="fi-ring-track" cx="82" cy="82" r="${R}" stroke-width="10"/>
              <circle class="fi-ring-fill" cx="82" cy="82" r="${R}" stroke-width="10"
                stroke-dasharray="${C.toFixed(1)}"
                stroke-dashoffset="${fill.toFixed(1)}"
                id="fiRingFill"/>
            </svg>
            <div class="fi-ring-label">
              <span class="fi-ring-pct" id="fiRingPct">0%</span>
              <span class="fi-ring-sub">to FI</span>
            </div>
          </div>
          <div class="home-lnw">${fm(lnw)}</div>
          <div class="home-lnw-sub">
            <span>Liquid NW</span>
            <span class="sep">|</span>
            <span style="color:var(--pink)">Target ${fm(target)}</span>
            <span class="sep">|</span>
            <span style="color:var(--amber)">Gap ${fm(gap)}</span>
          </div>
        </div>

        <!-- Key stats row -->
        <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(100px,1fr));gap:8px;margin-bottom:16px;">
          <div class="stat-card accent-indigo"><div class="sc-label">Total NW</div><div class="sc-value" style="font-size:0.95rem;">${fm(tnw)}</div></div>
          <div class="stat-card accent-sky"><div class="sc-label">Runway</div><div class="sc-value" style="font-size:0.95rem;">${runway}</div></div>
          ${sr !== null ? `<div class="stat-card accent-emerald"><div class="sc-label">Savings Rate</div><div class="sc-value" style="font-size:0.95rem;">${(sr*100).toFixed(1)}%</div></div>` : ''}
        </div>

        <!-- Monthly cash flow -->
        ${monthStr ? `
        <div style="margin-bottom:16px;">
          <div style="font-size:0.62rem;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:var(--muted2);margin-bottom:10px;">${monthStr} Cash Flow</div>
          <div class="home-month-bars">
            <div class="hmb-item">
              <span class="hmb-label inc">Income</span>
              <div class="hmb-bar-track"><div class="hmb-bar-fill hmb-fill-inc" style="width:${barW(income)}%"></div></div>
              <span class="hmb-val" style="color:var(--green)">${fm(income)}</span>
            </div>
            <div class="hmb-item">
              <span class="hmb-label exp">Expenses</span>
              <div class="hmb-bar-track"><div class="hmb-bar-fill hmb-fill-exp" style="width:${barW(expense)}%"></div></div>
              <span class="hmb-val" style="color:var(--red)">${fm(expense)}</span>
            </div>
            <div class="hmb-item">
              <span class="hmb-label inv">Invested</span>
              <div class="hmb-bar-track"><div class="hmb-bar-fill hmb-fill-inv" style="width:${barW(invest)}%"></div></div>
              <span class="hmb-val" style="color:var(--amber)">${fm(invest)}</span>
            </div>
          </div>
        </div>` : ''}

        <!-- Quick nav grid -->
        <div style="font-size:0.62rem;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:var(--muted2);margin-bottom:8px;">Quick Access</div>
        <div class="home-nav-grid" style="margin-bottom:16px;">
          <button class="home-nav-tile" onclick="switchTab('ledger')">
            <svg width="22" height="22" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8"><path d="M9 17v-2m3 2v-4m3 4v-6m2 10H7a2 2 0 01-2-2V5a2 2 0 012-2h5.586a1 1 0 01.707.293l5.414 5.414a1 1 0 01.293.707V19a2 2 0 01-2 2z"/></svg>
            Ledger
          </button>
          <button class="home-nav-tile" onclick="switchTab('portfolio')">
            <svg width="22" height="22" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8"><path d="M9 19v-6a2 2 0 00-2-2H5a2 2 0 00-2 2v6a2 2 0 002 2h2a2 2 0 002-2zm0 0V9a2 2 0 012-2h2a2 2 0 012 2v10m-6 0a2 2 0 002 2h2a2 2 0 002-2m0 0V5a2 2 0 012-2h2a2 2 0 012 2v14a2 2 0 01-2 2h-2a2 2 0 01-2-2z"/></svg>
            Portfolio
          </button>
          <button class="home-nav-tile" onclick="switchTab('montecarlo')">
            <svg width="22" height="22" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8"><path d="M13 7h8m0 0v8m0-8l-8 8-4-4-6 6"/></svg>
            Simulate
          </button>
          <button class="home-nav-tile" onclick="switchTab('roadmap')">
            <svg width="22" height="22" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8"><path d="M9 20l-5.447-2.724A1 1 0 013 16.382V5.618a1 1 0 011.447-.894L9 7m0 13l6-3m-6 3V7m6 10l4.553 2.276A1 1 0 0021 18.382V7.618a1 1 0 00-.553-.894L15 4m0 13V4m0 0L9 7"/></svg>
            Roadmap
          </button>
          <button class="home-nav-tile" onclick="switchTab('transactions')">
            <svg width="22" height="22" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8"><path d="M3 10h18M7 15h1m4 0h1m-7 4h12a3 3 0 003-3V8a3 3 0 00-3-3H6a3 3 0 00-3 3v8a3 3 0 003 3z"/></svg>
            Transactions
          </button>
          <button class="home-nav-tile" onclick="switchTab('forecast')">
            <svg width="22" height="22" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.8"><path d="M8 7V3m8 4V3m-9 8h10M5 21h14a2 2 0 002-2V7a2 2 0 00-2-2H5a2 2 0 00-2 2v12a2 2 0 002 2z"/></svg>
            Forecast
          </button>
        </div>

        <!-- Recent transactions -->
        ${txRows ? `
        <div style="font-size:0.62rem;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:var(--muted2);margin-bottom:4px;">
          Recent Transactions
          <button class="btn-sm" onclick="switchTab('transactions')" style="float:right;font-size:0.58rem;padding:2px 8px;">View All</button>
        </div>
        <div>${txRows}</div>` : ''}
      </div>
    `;

    // Animate ring counter
    const pctEl = document.getElementById('fiRingPct');
    if (pctEl) {
      const dur = 1200, t0 = performance.now();
      const tick = now => {
        const prog = Math.min((now-t0)/dur, 1);
        const eased = 1 - Math.pow(1-prog, 3);
        pctEl.textContent = (eased * pct).toFixed(1) + '%';
        if (prog < 1) requestAnimationFrame(tick);
      };
      requestAnimationFrame(tick);
    }
  } catch(e) {
    el.innerHTML = `<div style="color:#f87171;padding:16px;font-size:0.9rem;word-break:break-all"><b>JS ERROR:</b><br>${e.message}<br><pre style="font-size:0.75rem;margin-top:8px">${e.stack||''}</pre></div>`;
  }
}

// ── Ledger Dashboard ───────────────────────────────────────────────────────
async function loadLedger() {
  const el = document.getElementById('ledgerContent');
  el.innerHTML = `<div class="skeleton-wrap">
    <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(155px,1fr));gap:12px;">
      <div class="skeleton sk-card"></div><div class="skeleton sk-card"></div>
      <div class="skeleton sk-card"></div><div class="skeleton sk-card"></div>
      <div class="skeleton sk-card"></div>
    </div>
    <div class="skeleton sk-line wide" style="margin-top:8px;height:20px;border-radius:6px;"></div>
    <div class="skeleton sk-line med" style="height:14px;"></div>
    <div class="skeleton sk-chart" style="margin-top:4px;"></div>
  </div>`;
  try {
    const r = await fetch('/api/ledger/dashboard');
    const d = await r.json();
    if (d.error) { el.innerHTML = `<span class="err">${d.error}</span>`; return; }

    const m        = d.metrics;
    const lnw      = m['LIQUID NET WORTH']   || 0;
    const tnw      = m['TOTAL NET WORTH']    || 0;
    const cash     = m['LIQUID CASH']        || 0;
    const runway   = m['SURVIVAL RUNWAY']    || '—';
    const target   = m['FI TARGET (Age 62)'] || 0;
    const progress = m['PROGRESS TO FI']     || 0;
    const pct  = (progress * 100).toFixed(2);
    const barW = Math.min(progress * 100, 100).toFixed(1);

    const levels = d.freedom_levels || [];
    const levelRows = levels.map(lv => {
      let badge, goalStr;
      if (lv.status === 'ACHIEVED') {
        badge   = '<span class="lbadge badge-achieved">ACHIEVED</span>';
        goalStr = lv.goal_text || '—';
      } else if (lv.progress !== null && lv.progress !== undefined) {
        badge   = `<span class="lbadge badge-progress">${fp(lv.progress)}</span>`;
        goalStr = lv.goal ? fm(lv.goal) : '—';
      } else {
        badge   = `<span class="lbadge badge-progress">${lv.goal ? fm(lv.goal) : '—'}</span>`;
        goalStr = '';
      }
      return `<div class="level-row"><span class="lname">${lv.name}</span><span class="lgoal">${goalStr}</span>${badge}</div>`;
    }).join('');

    const alloc = d.allocation || {};
    const allocRows = Object.entries(alloc).map(([k,v]) =>
      `<div class="alloc-row"><span class="aname">${k}</span><span>${fm(v)}</span></div>`
    ).join('');

    const cf = d.cashflow || {};
    const cfRows = Object.entries(cf).map(([k,v]) => {
      const cls = k.includes('Surplus') ? 'neu' : v >= 0 ? 'pos' : 'neg';
      return `<div class="cashflow-row"><span class="cname">${k}</span><span class="cval ${cls}">${k.includes('Rate') ? fp(v) : fm(v)}</span></div>`;
    }).join('');

    const months    = d.spending_months || [];
    const spend     = d.spending || {};
    const srVals    = spend['SAVINGS RATE'] || [];
    const latestSR  = srVals.length ? srVals[srVals.length - 1] : null;
    const spendRows = Object.entries(spend).filter(([k]) => k !== 'SAVINGS RATE').map(([k, vals]) => {
      const cls = k === 'Income' ? 'val-pos' : (k === 'TRUE SAVING RATE') ? 'val-neu' : 'val-neg';
      const cells = vals.map(v => {
        if (v === null || v === undefined) return '<td>—</td>';
        if (typeof v === 'number' && Math.abs(v) < 2) return `<td class="val-neu">${(v*100).toFixed(1)}%</td>`;
        return `<td class="${cls}">${fm(v)}</td>`;
      }).join('');
      return `<tr><td>${k}</td>${cells}</tr>`;
    }).join('');
    const monthHeaders = months.map(m => `<th>${m}</th>`).join('');

    document.getElementById('refreshTs').textContent = 'Last read: ' + new Date().toLocaleTimeString();
    el.innerHTML = `
      <div class="stat-grid">
        <div class="stat-card accent-pink"><div class="sc-label">Liquid Net Worth</div><div class="sc-value">${fm(lnw)}</div><div class="sc-sub">Freedom Number</div></div>
        <div class="stat-card accent-indigo"><div class="sc-label">Total Net Worth</div><div class="sc-value">${fm(tnw)}</div><div class="sc-sub">Estate Number</div></div>
        <div class="stat-card accent-sky"><div class="sc-label">Liquid Cash</div><div class="sc-value">${fm(cash)}</div><div class="sc-sub">Safety Net</div></div>
        <div class="stat-card accent-amber"><div class="sc-label">Survival Runway</div><div class="sc-value" style="font-size:1.05rem;">${runway}</div><div class="sc-sub">at current burn</div></div>
        <div class="stat-card accent-teal"><div class="sc-label">FI Target (Age 62)</div><div class="sc-value">${fm(target)}</div><div class="sc-sub">The Goal</div></div>
        ${latestSR !== null ? `<div class="stat-card accent-emerald"><div class="sc-label">Savings Rate</div><div class="sc-value">${(latestSR*100).toFixed(1)}%</div><div class="sc-sub">latest month</div></div>` : ''}
      </div>
      <div class="fi-progress-wrap">
        <div class="fi-label"><span>Road to FI</span><span class="pct">${pct}%</span></div>
        <div class="progress-bar"><div class="progress-fill" style="width:${barW}%"></div></div>
      </div>
      ${sectionToggle('Freedom Levels', 'sec-levels', `<div class="levels-grid">${levelRows}</div>`, true)}
      ${sectionToggle('Allocation &amp; Cash Flow', 'sec-alloc', `
        <div class="two-col">
          <div class="sub-panel"><h3>Asset Allocation</h3>${allocRows}</div>
          <div class="sub-panel"><h3>Monthly Cash Flow</h3>${cfRows}</div>
        </div>
        <div id="allocChartWrap2" style="margin-top:14px;">
          <div class="donut-wrap">
            <canvas id="allocChart2" width="180" height="180" style="max-width:180px;"></canvas>
            <div class="donut-legend" id="allocLegend2"></div>
          </div>
        </div>
      `, false)}
      ${spendRows ? sectionToggle('Spending Table', 'sec-spend', `<table class="spending-table"><thead><tr><th>Category</th>${monthHeaders}</tr></thead><tbody>${spendRows}</tbody></table>`, false) : ''}
    `;
    // Render allocation donut after DOM settles
    setTimeout(() => renderAllocChart(alloc, 'allocChart2', 'allocLegend2'), 80);

    // ── Net Worth History ──
    const lnwNum = parseFloat(lnw) || 0;
    const tnwNum = parseFloat(tnw) || 0;
    if (lnwNum || tnwNum) {
      let hist = JSON.parse(localStorage.getItem('retAdv_nwHistory') || '[]');
      const today = new Date().toLocaleDateString('en-US', {month:'short', day:'numeric'});
      if (!hist.length || hist[hist.length-1].label !== today) {
        hist.push({ label: today, lnw: lnwNum, tnw: tnwNum });
        if (hist.length > 90) hist = hist.slice(-90);
        localStorage.setItem('retAdv_nwHistory', JSON.stringify(hist));
      }
      renderNWHistoryChart(hist);
    }

    // ── Contribution Inputs + Time-to-FI ──
    const targetNum = parseFloat(target) || 0;
    window._tfiCtx = { lnwNum, targetNum };
    const rulesData = window._rules || await fetch('/api/rules').then(r=>r.json()).catch(()=>({}));
    window._rules = rulesData;
    const sv = JSON.parse(localStorage.getItem('retAdv_contrib') || '{}');
    el.insertAdjacentHTML('beforeend', `
      <div style="margin-top:14px;padding:12px;background:rgba(255,255,255,0.025);border:1px solid var(--border2);border-radius:var(--radius);">
        <div style="font-size:0.68rem;font-weight:700;color:var(--gold);letter-spacing:0.07em;text-transform:uppercase;margin-bottom:10px;">Contribution Inputs</div>
        <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:8px;">
          <label>Weekly Taxable ($)<input id="ci_taxable" type="number" step="0.01" min="0" value="${sv.weeklyTaxable||''}"></label>
          <label>Weekly IRA ($)<input id="ci_ira" type="number" step="0.01" min="0" value="${sv.weeklyIRA||''}"></label>
          <label>401k Employee (%)<input id="ci_401k_pct" type="number" step="0.1" min="0" max="100" value="${sv.emp401kPct||''}"></label>
          <label>Employer Match (%)<input id="ci_match_pct" type="number" step="0.1" min="0" max="100" value="${sv.matchPct||''}"></label>
          <label>Annual Gross ($)<input id="ci_gross" type="number" step="1" min="0" value="${sv.annualGross||''}"></label>
        </div>
      </div>
      <div id="tfiCardWrap"></div>
    `);
    ['ci_taxable','ci_ira','ci_401k_pct','ci_match_pct','ci_gross'].forEach(id => {
      document.getElementById(id)?.addEventListener('input', () =>
        calcContribTFI(window._tfiCtx.lnwNum, window._tfiCtx.targetNum, window._rules || {})
      );
    });
    calcContribTFI(lnwNum, targetNum, rulesData);

    // ── Budget vs Actual ──
    renderBudgetTable(d.spending, d.spending_months);
    loadDigestCreds();

  } catch(e) {
    el.innerHTML = `<span class="err">Error: ${e.message}</span>`;
  }
}

// ── Rules ──────────────────────────────────────────────────────────────────
async function loadRules() {
  try {
    const r = await fetch('/api/rules');
    const d = await r.json();
    document.getElementById('rulesGrid').innerHTML = `
      <div class="stat-card accent-amber"><div class="sc-label">401k / 403b</div><div class="sc-value" style="font-size:1.2rem;">$${d.contrib_401k.toLocaleString()}</div><div class="sc-sub">base limit</div></div>
      <div class="stat-card accent-sky"><div class="sc-label">Catch-up (50+)</div><div class="sc-value" style="font-size:1.2rem;">$${d.catchup_50_plus.toLocaleString()}</div><div class="sc-sub">additional</div></div>
      <div class="stat-card accent-violet"><div class="sc-label">Super Catch-up (60–63)</div><div class="sc-value" style="font-size:1.2rem;">$${d.super_catchup_60_63.toLocaleString()}</div><div class="sc-sub">replaces standard</div></div>
      <div class="stat-card accent-emerald"><div class="sc-label">IRA / Roth IRA</div><div class="sc-value" style="font-size:1.2rem;">$${d.ira_roth_limit.toLocaleString()}</div><div class="sc-sub">$${d.ira_roth_50_plus.toLocaleString()} if 50+</div></div>
    `;
    document.getElementById('rothNote').innerHTML =
      `<span>&#9888; Roth-ification rule:</span> If ${d.year-1} income &gt; $${d.rothification_income_threshold.toLocaleString()}, catch-up contributions <strong>must</strong> go Roth.`;
    rulesLoaded = true;
  } catch(e) {
    document.getElementById('rulesGrid').innerHTML = '<div class="stat-card"><div class="sc-label" style="color:var(--red)">Failed to load</div></div>';
  }
}

// ── Portfolio ──────────────────────────────────────────────────────────────
const PORT_KEY_STORE = 'retAdv_avKey';

function acctType(section) {
  const s = section.toLowerCase();
  if (s.includes('roth') || s.includes('converted')) return 'roth';
  if (s.includes('401') || s.includes('proxy') || s.includes('voya')) return '401k';
  if (s.includes('crypto') || s.includes('bitcoin')) return 'crypto';
  return 'taxable';
}
function acctTypeLabel(section) {
  const t = acctType(section);
  return {roth:'ROTH', '401k':'401K', crypto:'CRYPTO', taxable:'TAXABLE'}[t];
}

function buildPortTable(holdings, priceField, valueField) {
  if (!holdings.length) return '<div style="color:var(--muted2);font-size:0.8rem;">No holdings found in ledger.</div>';

  const sections = {};
  for (const h of holdings) {
    if (!sections[h.section]) sections[h.section] = [];
    sections[h.section].push(h);
  }

  let grandValue = 0, grandCost = 0, grandCount = 0;
  let html = '<div class="port-accounts">';

  for (const [section, rows] of Object.entries(sections)) {
    let secValue = 0;
    for (const h of rows) {
      const v = h[valueField] || 0;
      secValue   += v;
      grandValue += v;
      grandCost  += (h.avg_cost || 0) * h.shares;
      grandCount++;
    }

    const type    = acctType(section);
    const tag     = acctTypeLabel(section);
    const secName = section.replace(/\s*-\s*Money Machine/i,'').replace(/\s*-\s*Bridge Asset [A-Z]/i,'').replace(/\s*-\s*Moat Fund/i,' \u2014 Moat').trim();

    html += `
    <div class="port-account">
      <div class="port-acct-hdr">
        <div class="port-acct-left">
          <span class="port-acct-tag tag-${type}">${tag}</span>
          <span class="port-acct-name">${secName}</span>
        </div>
        <span class="port-acct-total">${secValue ? fm(secValue) : '—'}</span>
      </div>
      <table class="port-htable"><tbody>`;

    for (const h of rows) {
      const price  = h[priceField];
      const value  = h[valueField];
      const shares = h.shares < 0.01 ? h.shares.toFixed(6) : h.shares < 1 ? h.shares.toFixed(4) : h.shares.toFixed(3);
      const priceStr = price      ? '$' + Number(price).toFixed(2) : '—';
      const valueStr = value      ? fm(value)                      : '—';
      const costStr  = h.avg_cost ? '$' + h.avg_cost.toFixed(2) + ' \u2192' : '';

      let glCell;
      if (h.proxy_note) {
        glCell = `<span class="gl-pill proxy">Voya proxy</span>`;
      } else if (h.gain_loss != null) {
        const sign = h.gain_loss >= 0 ? '+' : '';
        const cls  = h.gain_loss >= 0 ? 'gain' : 'loss';
        const pct  = h.gain_loss_pct != null ? `${sign}${h.gain_loss_pct}%` : '';
        glCell = `<span class="gl-pill ${cls}">${pct || (sign + fm(h.gain_loss))}</span>`;
      } else if (h.error) {
        glCell = `<span style="color:var(--muted2);font-size:0.62rem;">${h.error}</span>`;
      } else {
        glCell = '<span style="color:var(--muted2);">\u2014</span>';
      }

      html += `
        <tr>
          <td class="h-ticker"><span class="ticker-badge">${h.ticker}</span></td>
          <td class="h-name">${h.name.length > 35 ? h.name.slice(0,35)+'\u2026' : h.name}</td>
          <td class="h-r h-muted">${shares} sh</td>
          <td class="h-r h-muted">${costStr}</td>
          <td class="h-r" style="font-family:var(--font-mono);">${priceStr}</td>
          <td class="h-val">${valueStr}</td>
          <td class="h-gl">${glCell}</td>
        </tr>`;
    }
    html += `</tbody></table></div>`;
  }
  html += '</div>';

  // Update summary stat-cards
  document.getElementById('ps_value').textContent = grandValue ? fm(grandValue) : '—';
  document.getElementById('ps_cost').textContent  = grandCost  ? fm(grandCost)  : '—';
  document.getElementById('ps_count').textContent = grandCount;
  const totalGl = grandValue - grandCost;
  const glCard  = document.getElementById('ps_gl_card');
  if (grandValue && grandCost) {
    const sign  = totalGl >= 0 ? '+' : '';
    const glPct = (totalGl / grandCost * 100).toFixed(1);
    document.getElementById('ps_gl').textContent     = sign + fm(totalGl);
    document.getElementById('ps_gl_pct').textContent = sign + glPct + '%';
    glCard.className = 'stat-card ' + (totalGl >= 0 ? 'gl-pos' : 'gl-neg');
  }

  // Tax summary panel
  if (grandValue && grandCost) {
    let totalGain = 0, totalLoss = 0;
    // Re-iterate holdings for per-position G/L (available via closure over holdings param)
    // holdings is the first param; we need to walk sections again
    // Instead compute from grandValue/grandCost which we have
    // For per-row breakdown we need to recompute from sections built above
    // Walk sections object (it's in scope)
    for (const rows of Object.values(sections)) {
      for (const h of rows) {
        if (h.gain_loss != null) {
          if (h.gain_loss >= 0) totalGain += h.gain_loss;
          else totalLoss += h.gain_loss;
        }
      }
    }
    const netGl = totalGain + totalLoss;
    const estLtcg = netGl > 0 ? netGl * 0.15 : 0;
    html += `
    <div class="tax-summary">
      <div class="tax-header">Unrealized G/L Tax Summary</div>
      <div class="tax-row"><span>Total Unrealized Gains</span><span class="clr-green">+${fm(totalGain)}</span></div>
      <div class="tax-row"><span>Total Unrealized Losses</span><span class="clr-red">${fm(totalLoss)}</span></div>
      <div class="tax-row tax-net"><span>Net Unrealized G/L</span><span class="${netGl >= 0 ? 'clr-green' : 'clr-red'}">${netGl >= 0 ? '+' : ''}${fm(netGl)}</span></div>
      <div class="tax-row"><span>Est. LTCG Tax @15%</span><span style="color:var(--muted);">${estLtcg > 0 ? fm(estLtcg) : '—'}</span></div>
      <div class="tax-note">Assumes long-term treatment. Verify hold periods. Does not include state tax.</div>
    </div>`;
  }

  return html;
}

async function loadPortfolio() {
  portfolioLoaded = true;
  const savedKey = localStorage.getItem(PORT_KEY_STORE);
  const el = document.getElementById('portTable');

  if (savedKey) {
    document.getElementById('portApiKey').value = savedKey;
    document.getElementById('portKeyStatus').innerHTML =
      '<span class="saved">\u25b6 Auto-refreshing live prices\u2026</span>';
    await refreshPortfolio();
  } else {
    el.innerHTML = `<div class="skeleton-wrap">
      <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(155px,1fr));gap:12px;">
        <div class="skeleton sk-card"></div><div class="skeleton sk-card"></div>
        <div class="skeleton sk-card"></div><div class="skeleton sk-card"></div>
      </div>
      <div class="skeleton sk-chart" style="margin-top:8px;"></div>
    </div>`;
    try {
      const r = await fetch('/api/portfolio');
      const d = await r.json();
      if (d.error) { el.innerHTML = `<span class="err">${d.error}</span>`; return; }
      el.innerHTML = buildPortTable(d.holdings, 'cached_price', 'cached_value');
      document.getElementById('ps_asof').textContent = 'Cached \u2014 add API key to enable live prices';
    } catch(e) {
      el.innerHTML = `<span class="err">Error: ${e.message}</span>`;
    }
  }
}

async function refreshPortfolio() {
  const key    = document.getElementById('portApiKey').value.trim();
  const status = document.getElementById('portRefreshStatus');
  const el     = document.getElementById('portTable');
  if (!key) { status.innerHTML = '<span class="err">Enter your Alpha Vantage API key.</span>'; return; }
  localStorage.setItem(PORT_KEY_STORE, key);
  document.getElementById('portKeyStatus').innerHTML = '<span class="saved">\u2713 API key saved</span>';
  el.innerHTML = '<div style="color:var(--muted2);font-size:0.8rem;">Fetching live prices\u2026</div>';
  status.textContent = 'Fetching live prices (may take ~5s for all holdings)\u2026';
  try {
    const r = await fetch('/api/portfolio/refresh', {
      method: 'POST', headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({api_key: key})
    });
    const d = await r.json();
    if (d.error) { status.innerHTML = `<span class="err">${d.error}</span>`; return; }
    el.innerHTML = buildPortTable(d.holdings, 'live_price', 'live_value');
    const now = new Date().toLocaleTimeString();
    document.getElementById('ps_asof').textContent = 'Live \u2014 ' + now;
    status.innerHTML = `<span class="saved">\u2713 Prices updated at ${now}</span>`;
  } catch(e) {
    status.innerHTML = `<span class="err">Error: ${e.message}</span>`;
  }
}

// ── Monte Carlo ────────────────────────────────────────────────────────────
const PRESETS = {
  baseline: { mc_mu: 9,  mc_sigma: 16, mc_infl: 3.0, mc_haircut: false, mc_tail: false },
  cautious: { mc_mu: 7,  mc_sigma: 16, mc_infl: 3.5, mc_haircut: true,  mc_tail: false },
  stress:   { mc_mu: 6,  mc_sigma: 22, mc_infl: 4.0, mc_haircut: true,  mc_tail: true  },
  optimist: { mc_mu: 11, mc_sigma: 14, mc_infl: 2.5, mc_haircut: false, mc_tail: false },
};

function applyPreset(name) {
  document.querySelectorAll('.preset-pill').forEach(p => p.classList.remove('active'));
  document.querySelector('.preset-pill.' + name).classList.add('active');
  const p = PRESETS[name];
  document.getElementById('mc_mu').value    = p.mc_mu;
  document.getElementById('mc_sigma').value = p.mc_sigma;
  document.getElementById('mc_infl').value  = p.mc_infl;
  document.getElementById('mc_haircut').checked = p.mc_haircut;
  document.getElementById('mc_tail').checked    = p.mc_tail;
  runMC();
}

async function prefillMC() {
  mcPrefilled = true;
  try {
    const r = await fetch('/api/ledger/dashboard');
    const d = await r.json();
    if (d.error) return;
    const p = d.mc_prefill;
    if (!p) return;
    const set = (id, val) => { if (val != null && val > 0) document.getElementById(id).value = val; };
    set('mc_engine',   p.engine_balance);
    set('mc_sgov',     p.sgov_balance);
    set('mc_checking', p.checking_balance);
    set('mc_full_ss',  p.full_ss_annual);
    if (p.annual_floor_cost > 0) set('mc_floor', p.annual_floor_cost);
    const filled = ['mc_engine','mc_sgov','mc_checking','mc_full_ss']
      .filter(id => parseFloat(document.getElementById(id).value) > 0);
    if (filled.length > 0) {
      const note = document.getElementById('mc_prefill_note');
      if (note) note.style.display = 'block';
    }
  } catch(e) {}
}

async function runMC() {
  const btn     = document.getElementById('mc_run_btn');
  const spinner = document.getElementById('mc_spinner');
  const panel   = document.getElementById('mc_results_panel');
  const simSizeEl = document.querySelector('input[name="mc_sim_size"]:checked');
  const simSize   = simSizeEl ? simSizeEl.value : '1k';
  const returnModelEl = document.querySelector('input[name="mc_return_model"]:checked');
  const returnModel   = returnModelEl ? returnModelEl.value : 'normal';
  const pct     = id => parseFloat(document.getElementById(id).value) / 100;

  const params = {
    current_age:             parseInt(document.getElementById('mc_age').value),
    target_age:              parseInt(document.getElementById('mc_ret_age').value),
    filing_status:           document.getElementById('mc_filing').value,
    start_engine:            parseFloat(document.getElementById('mc_engine').value)   || 0,
    start_sgov:              parseFloat(document.getElementById('mc_sgov').value)     || 0,
    start_checking:          parseFloat(document.getElementById('mc_checking').value) || 0,
    annual_contribution:     parseFloat(document.getElementById('mc_contrib').value)  || 0,
    full_ss:                 parseFloat(document.getElementById('mc_full_ss').value)  || 25620,
    mean_return:             pct('mc_mu'),
    volatility:              pct('mc_sigma'),
    sgov_yield:              pct('mc_syld'),
    inflation_rate:          pct('mc_infl'),
    strict_moat_cost:        parseFloat(document.getElementById('mc_floor').value)        || 54292,
    moat_target:             parseFloat(document.getElementById('mc_moat_target').value) || 360000,
    use_ss_haircut:          document.getElementById('mc_haircut').checked,
    ss_haircut_pct:          0.21,
    use_tail_shock:          document.getElementById('mc_tail').checked,
    tail_shock_return:       -0.25,
    tail_shock_count:        1,
    use_mortality_weighting: document.getElementById('mc_mort').checked,
    use_conversion:          document.getElementById('mc_use_conv').checked,
    trad_balance:            parseFloat(document.getElementById('mc_trad').value)     || 0,
    target_bracket:          parseFloat(document.getElementById('mc_tgt_bkt').value)  || 0.12,
    state_tax_rate:          pct('mc_state_tx'),
    seq_shock_year:          parseInt(document.querySelector('input[name="mc_seq"]:checked')?.value || '0'),
    sim_size: simSize,
    return_model: returnModel,
    wage_growth: 0.02, dividend_yield: 0.015, ss_age: 67,
    gogo_e: 0.25, gogo_n: 0.15, slowgo_e: 0.20, slowgo_n: 0.10, nogo_e: 0.10, nogo_n: 0.05,
    euphoric_offset: 0.03,
    gk_trigger: 0.20, gk_cut_rate: 0.50, bear_streak_years: 3, bear_streak_cut: 0.25,
    portfolio_cap: 5000000, cap_inflation: 0.03, cap_gogo: 0.10, cap_slowgo: 0.05, cap_nogo: 0.02,
    tent_skim_rate: 0.50, phase3_moat_years: 2,
  };

  btn.disabled = true; btn.textContent = 'Running\u2026';
  spinner.style.display = 'block';
  panel.innerHTML = `<div class="mc-result-hint">Running ${simSize.toUpperCase()} trials\u2026</div>`;

  try {
    const r = await fetch('/api/monte-carlo', {
      method: 'POST', headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(params)
    });
    const d = await r.json();
    if (d.error) { panel.innerHTML = `<span class="err">${d.error}</span>`; return; }
    renderMCResults(d);
  } catch(e) {
    panel.innerHTML = `<span class="err">Network error: ${e.message}</span>`;
  } finally {
    btn.disabled = false; btn.textContent = 'Run Simulation';
    spinner.style.display = 'none';
  }
}

function renderMCChart(d) {
  if (!d.bands || !d.bands.ages) return;
  const ctx = document.getElementById('mcChart');
  if (!ctx) return;
  if (window._mcChart) window._mcChart.destroy();
  window._mcChart = new Chart(ctx, {
    type: 'line',
    data: {
      labels: d.bands.ages,
      datasets: [
        { label: 'P90', data: d.bands.p90, fill: '+1', borderColor: 'rgba(205,133,63,0.55)', backgroundColor: 'rgba(205,133,63,0.07)', tension: 0.4, pointRadius: 0, borderWidth: 1.5 },
        { label: 'P75', data: d.bands.p75, fill: '+1', borderColor: 'rgba(205,133,63,0.25)', backgroundColor: 'rgba(205,133,63,0.13)', tension: 0.4, pointRadius: 0, borderWidth: 0 },
        { label: 'P50', data: d.bands.p50, fill: false, borderColor: '#cd853f', borderWidth: 2.5, backgroundColor: 'transparent', tension: 0.4, pointRadius: 0 },
        { label: 'P25', data: d.bands.p25, fill: '-1',  borderColor: 'rgba(205,133,63,0.25)', backgroundColor: 'rgba(205,133,63,0.13)', tension: 0.4, pointRadius: 0, borderWidth: 0 },
        { label: 'P10', data: d.bands.p10, fill: '-1',  borderColor: 'rgba(205,133,63,0.55)', backgroundColor: 'rgba(205,133,63,0.07)', tension: 0.4, pointRadius: 0, borderWidth: 1.5 },
      ]
    },
    options: {
      responsive: true,
      animation: { duration: 900, easing: 'easeOutQuart' },
      interaction: { mode: 'index', intersect: false },
      plugins: {
        legend: { display: false },
        tooltip: {
          backgroundColor: 'rgba(5,5,5,0.95)', borderColor: 'rgba(205,133,63,0.30)', borderWidth: 1,
          titleColor: '#cd853f', bodyColor: '#c8b89a', padding: 12, cornerRadius: 6,
          callbacks: { label: ctx => `  ${ctx.dataset.label}: $${(ctx.raw/1000).toFixed(0)}k` }
        }
      },
      scales: {
        x: { ticks: { color: '#5a5040', font: { family: 'JetBrains Mono, ui-monospace', size: 10 }, maxTicksLimit: 8 }, grid: { color: 'rgba(205,133,63,0.06)' }, border: { color: 'rgba(205,133,63,0.12)' } },
        y: { ticks: { color: '#5a5040', callback: v => v >= 1000000 ? '$' + (v/1000000).toFixed(1) + 'M' : '$' + (v/1000).toFixed(0) + 'k', font: { family: 'JetBrains Mono, ui-monospace', size: 10 } }, grid: { color: 'rgba(205,133,63,0.06)' }, border: { color: 'rgba(205,133,63,0.12)' } }
      }
    }
  });
}

function renderPlaybook(d) {
  const s = d.stats || {};
  const lines = [];
  if (d.success_pct >= 90)
    lines.push({ cls: 'emerald', text: `Strong plan \u2014 ${d.success_pct}% of simulated futures leave you solvent at 95.` });
  else if (d.success_pct >= 75)
    lines.push({ cls: 'amber',   text: `Viable plan \u2014 ${d.success_pct}% success. Consider building moat or delaying SS.` });
  else
    lines.push({ cls: 'rose',    text: `At risk \u2014 only ${d.success_pct}% of futures survive to 95. Increase contributions or reduce floor.` });

  if (s.ripcord_rate > 30)
    lines.push({ cls: 'rose',  text: `${s.ripcord_rate}% of futures trigger early SS (before 67) \u2014 moat is undersized for your target.` });
  if (s.moat_breach_rate > 20)
    lines.push({ cls: 'amber', text: `Moat breaches in ${s.moat_breach_rate}% of futures \u2014 bridge reserve runs short before SS kicks in.` });
  if (s.median_drawdown > 40)
    lines.push({ cls: 'amber', text: `Median max drawdown ${s.median_drawdown}% \u2014 expect significant paper losses in bad scenarios.` });
  if (s.median_ss_age <= 63)
    lines.push({ cls: 'amber', text: `Median SS claim age ${s.median_ss_age} \u2014 most futures claim early, reducing lifetime benefit.` });
  if (s.tax_savings > 0)
    lines.push({ cls: 'teal',  text: `Roth conversions save ~${fm(s.tax_savings)} in median lifetime taxes vs. no-conversion.` });

  if (!lines.length) return '';
  const lineHTML = lines.map(l =>
    `<div class="playbook-line pb-${l.cls}"><div class="playbook-dot"></div><div>${l.text}</div></div>`
  ).join('');
  return `<div class="playbook-panel"><h3>&#128214; Advisor Playbook</h3>${lineHTML}</div>`;
}

function renderMCResults(d) {
  const pct = d.success_pct;
  const cls = pct >= 90 ? 'clr-green' : pct >= 75 ? 'clr-amber' : 'clr-red';
  const s   = d.stats || {};

  const milestoneRows = (d.milestones || []).map(m =>
    `<tr><td>Age ${m.age}</td><td>${fm(m.p10)}</td><td>${fm(m.p50)}</td><td>${fm(m.p90)}</td></tr>`
  ).join('');

  const convStats = (s.conv_tax_paid > 0)
    ? `<div class="stat-card"><div class="sc-label">Conv. Tax Paid</div><div class="sc-value" style="font-size:1rem;">${fm(s.conv_tax_paid)}</div></div>
       <div class="stat-card accent-emerald"><div class="sc-label">Tax Savings</div><div class="sc-value" style="font-size:1rem;color:var(--green)">${fm(Math.max(0,s.tax_savings))}</div></div>`
    : '';

  // Initialize tab structure
  document.getElementById('mc_results_panel').innerHTML = `
    <div class="mc-inner-tabs">
      <button class="mc-itab active" id="mc_itab_0" onclick="switchMCTab(0)">&#128202; Results</button>
      <button class="mc-itab" id="mc_itab_1" onclick="switchMCTab(1)">&#128214; Plan Details</button>
    </div>
    <div id="mc_t1"></div>
    <div id="mc_t2" style="display:none"></div>
  `;

  // Plan Status Hero Card
  const verdictCls   = pct >= 90 ? 'hero-secure' : pct >= 75 ? 'hero-ok' : pct >= 60 ? 'hero-caution' : 'hero-risk';
  const verdictLabel = pct >= 90 ? 'Plan Secure' : pct >= 75 ? 'On Track' : pct >= 60 ? 'Caution' : 'At Risk';
  const verdictText  = pct >= 90
    ? `Your strategy survives ${pct}% of simulated lifetimes \u2014 well above the 90% confidence target.`
    : pct >= 75
    ? `Solid foundation with tail risk. ${pct}% of simulated paths stay solvent through age 95.`
    : pct >= 60
    ? `Plan survives ${pct}% of paths. Consider boosting contributions or increasing your SGOV moat.`
    : `Only ${pct}% of paths avoid ruin. Immediate action recommended before retirement.`;
  document.getElementById('mc_t1').innerHTML = `
    <div class="plan-hero ${verdictCls}">
      <div class="ph-verdict">${verdictLabel}</div>
      <div class="ph-conf-row"><div class="ph-bar"><div class="ph-fill" style="width:0%;" id="phFill"></div></div><span class="ph-pct">${pct}%</span></div>
      <div class="ph-text">${verdictText}</div>
    </div>
  `;
  setTimeout(() => { const f = document.getElementById('phFill'); if(f) f.style.width = pct + '%'; }, 50);

  document.getElementById('mc_t1').insertAdjacentHTML('beforeend', `
    <div class="success-badge badge-${pct >= 90 ? 'great' : pct >= 75 ? 'ok' : 'warn'}">
      <div class="sb-ring-wrap">
        <div class="sb-ring"></div>
        <div class="sb-ring"></div>
        <div class="sb-ring"></div>
        <div class="sbval ${cls}" id="sbAnimVal">0%</div>
      </div>
      <div class="sblabel">Success Rate &mdash; ${d.trial_count.toLocaleString()} trials</div>
      <button class="pin-btn btn-sm" id="pinScenBtn" onclick="pinScenario()" style="margin-top:10px;">${scenarioA ? 'Scenario A pinned &#10003;' : 'Pin as Scenario A'}</button>
    </div>

    <div class="panel accent-indigo" style="margin-bottom:14px;padding:16px 18px;">
      <h2>Wealth Trajectory</h2>
      <canvas id="mcChart" style="max-height:240px;"></canvas>
    </div>

    <div class="panel accent-sky" style="margin-bottom:14px;padding:16px 18px;">
      <h2>Wealth Percentiles by Age</h2>
      <table class="mc-table">
        <thead><tr><th>Age</th><th>P10 (Bear)</th><th>P50 (Median)</th><th>P90 (Bull)</th></tr></thead>
        <tbody>${milestoneRows}</tbody>
      </table>
    </div>

    <div class="panel accent-violet" style="padding:16px 18px;">
      <h2>Simulation Statistics</h2>
      <div class="mc-stats-strip">
        <div class="stat-card"><div class="sc-label">Arrival Wealth</div><div class="sc-value" style="font-size:1rem;">${fm(s.median_arrival)}</div></div>
        <div class="stat-card"><div class="sc-label">SS Claim Age</div><div class="sc-value" style="font-size:1rem;">${s.median_ss_age}</div></div>
        <div class="stat-card"><div class="sc-label">Ripcord Rate</div><div class="sc-value" style="font-size:1rem;">${s.ripcord_rate}%</div></div>
        <div class="stat-card"><div class="sc-label">Moat Breach</div><div class="sc-value" style="font-size:1rem;">${s.moat_breach_rate}%</div></div>
        <div class="stat-card"><div class="sc-label">Terminal Wealth</div><div class="sc-value" style="font-size:1rem;">${fm(s.median_terminal)}</div></div>
        <div class="stat-card"><div class="sc-label">Go-Go Spend</div><div class="sc-value" style="font-size:1rem;">${fm(s.median_gogo_spend)}</div></div>
        <div class="stat-card"><div class="sc-label">Max Drawdown</div><div class="sc-value" style="font-size:1rem;">${s.median_drawdown}%</div></div>
        ${convStats}
      </div>
      <div class="mc-runtime">Computed in ${d.runtime_ms}ms &mdash; ${d.trial_count.toLocaleString()} trials</div>
    </div>
  `);

  setTimeout(() => renderMCChart(d), 50);

  // Animate success % counter
  const sbEl = document.getElementById('sbAnimVal');
  if (sbEl) {
    const target = pct, duration = 1200, startTime = performance.now();
    function tick(now) {
      const elapsed = now - startTime;
      const progress = Math.min(elapsed / duration, 1);
      const eased = 1 - Math.pow(1 - progress, 3);
      sbEl.textContent = Math.round(eased * target) + '%';
      if (progress < 1) requestAnimationFrame(tick);
    }
    requestAnimationFrame(tick);
  }

  // AI Playbook — replaces canned renderPlaybook with Ollama-generated narrative
  document.getElementById('mc_t1').insertAdjacentHTML('beforeend', `
    <div class="ai-playbook" id="aiPlaybookPanel">
      <div class="ai-playbook-hdr">
        <h3>&#128214; Advisor Playbook</h3>
        <span class="ai-badge">AI</span>
      </div>
      <div class="ai-playbook-body" id="aiPlaybookBody">
        <div class="ai-thinking">
          <div class="ai-thinking-dots"><span></span><span></span><span></span></div>
          <span>Generating AI analysis\u2026</span>
        </div>
      </div>
    </div>`);
  // Fire summarize async — non-blocking
  (async () => {
    const model = document.getElementById('aiModel')?.value || 'llama3.1:8b';
    try {
      const r = await fetch('/api/summarize', {
        method: 'POST',
        headers: {'Content-Type':'application/json'},
        body: JSON.stringify({sim_data: d, summary_type: 'playbook', model})
      });
      const sd = await r.json();
      const bodyEl = document.getElementById('aiPlaybookBody');
      if (bodyEl) bodyEl.innerHTML = sd.error ? mdLite(renderPlaybook(d) || sd.error) : mdLite(sd.summary);
    } catch(_) {
      const bodyEl = document.getElementById('aiPlaybookBody');
      if (bodyEl) { const fb = renderPlaybook(d); bodyEl.innerHTML = fb || '\u26a0 AI unavailable.'; }
    }
  })();

  // Ruin by age table
  const ruinEntries = Object.entries(d.ruin_by_age || {});
  if (ruinEntries.length) {
    const ruinRows = ruinEntries.map(([age, pct2]) => {
      const rc = pct2 < 5 ? 'clr-green' : pct2 < 15 ? 'clr-amber' : 'clr-red';
      return `<tr><td>By Age ${age}</td><td class="${rc}">${pct2}%</td></tr>`;
    }).join('');
    document.getElementById('mc_t1').insertAdjacentHTML('beforeend',
      `<div class="panel accent-red" style="margin-top:14px;padding:16px 18px;">
        <h2>Probability of Ruin</h2>
        <table class="ruin-table">
          <thead><tr><th>Milestone</th><th style="text-align:right;">P(Wealth &le; 0)</th></tr></thead>
          <tbody>${ruinRows}</tbody>
        </table>
      </div>`
    );
  }

  // Moat balance trajectory table
  const mb = d.moat_bands;
  if (mb && mb.ages && mb.ages.length > 1) {
    const mbRows = mb.ages.map((age, i) => {
      const p10 = mb.p10[i], p50 = mb.p50[i], p90 = mb.p90[i];
      const cls = p10 <= 0 ? 'clr-red' : p10 < 50000 ? 'clr-amber' : 'clr-green';
      return `<tr>
        <td>Age ${age}</td>
        <td class="${p10<=0?'clr-red':p10<50000?'clr-amber':''}">${fm(p10)}</td>
        <td>${fm(p50)}</td>
        <td class="clr-green">${fm(p90)}</td>
      </tr>`;
    }).join('');
    document.getElementById('mc_t1').insertAdjacentHTML('beforeend',
      `<div class="panel accent-amber" style="margin-top:14px;padding:16px 18px;">
        <h2>SGOV Moat Balance — Bridge Period</h2>
        <table class="mc-table">
          <thead><tr><th>Age</th><th>P10 (Stress)</th><th>P50 (Median)</th><th>P90 (Bull)</th></tr></thead>
          <tbody>${mbRows}</tbody>
        </table>
        <div style="font-size:0.68rem;color:var(--muted2);margin-top:8px;">Red = moat exhausted before SS. Moat remainder folds into engine at SS activation age.</div>
      </div>`
    );
  }

  // SS timing histogram
  const ssHist = d.ss_histogram || {};
  const ssHistEntries = Object.entries(ssHist).filter(([,v]) => v > 0);
  if (ssHistEntries.length) {
    document.getElementById('mc_t1').insertAdjacentHTML('beforeend',
      `<div class="panel accent-sky" style="margin-top:14px;padding:16px 18px;">
         <h2>Social Security Claim Age Distribution</h2>
         <canvas id="ssHistChart" style="max-height:180px;"></canvas>
       </div>`
    );
    setTimeout(() => {
      const ctx = document.getElementById('ssHistChart');
      if (!ctx) return;
      const ages = Object.keys(ssHist).map(Number);
      const counts = ages.map(a => ssHist[String(a)] || 0);
      const total = counts.reduce((a,b)=>a+b,0);
      const pcts = counts.map(c => total > 0 ? (c/total*100).toFixed(1) : 0);
      const colors = ages.map(a => a < 67 ? '#f43f5e' : a === 67 ? '#8899b8' : '#10b981');
      new Chart(ctx, {
        type: 'bar',
        data: {
          labels: ages.map(a => 'Age ' + a),
          datasets: [{ data: pcts, backgroundColor: colors, borderRadius: 4 }]
        },
        options: {
          plugins: { legend: { display: false }, tooltip: { callbacks: { label: c => c.raw + '% of trials' } } },
          scales: {
            x: { grid: { color: 'rgba(255,255,255,0.05)' }, ticks: { color: '#8899b8', font: { size: 10 } } },
            y: { grid: { color: 'rgba(255,255,255,0.05)' }, ticks: { color: '#8899b8', font: { size: 10 }, callback: v => v + '%' }, beginAtZero: true }
          }
        }
      });
    }, 80);
  }

  // Lifetime spending panel
  const ls = d.lifetime_spend;
  if (ls) {
    document.getElementById('mc_t1').insertAdjacentHTML('beforeend',
      `<div class="panel accent-teal" style="margin-top:14px;padding:16px 18px;">
         <h2>Lifetime Discretionary Spending</h2>
         <div class="mc-stats-strip">
           <div class="stat-card"><div class="sc-label">Total (Median)</div><div class="sc-value" style="font-size:1rem;">${fm(ls.p50_total)}</div></div>
           <div class="stat-card"><div class="sc-label">Total (P90 Bull)</div><div class="sc-value" style="font-size:1rem;">${fm(ls.p90_total)}</div></div>
           <div class="stat-card accent-emerald"><div class="sc-label">Go-Go (62–75)</div><div class="sc-value" style="font-size:1rem;">${fm(ls.p50_gogo)}</div></div>
           <div class="stat-card accent-sky"><div class="sc-label">Slow-Go (76–85)</div><div class="sc-value" style="font-size:1rem;">${fm(ls.p50_slowgo)}</div></div>
           <div class="stat-card accent-violet"><div class="sc-label">No-Go (86–95)</div><div class="sc-value" style="font-size:1rem;">${fm(ls.p50_nogo)}</div></div>
           <div class="stat-card"><div class="sc-label">Spend Ratio</div><div class="sc-value" style="font-size:1rem;">${ls.spend_ratio}%</div><div class="sc-sub">of spend+terminal</div></div>
         </div>
       </div>`
    );
  }

  // Spending by phase stacked bar
  const ss2 = d.spend_scenarios;
  if (ss2) {
    document.getElementById('mc_t1').insertAdjacentHTML('beforeend',
      `<div class="panel accent-teal" style="margin-top:14px;padding:16px 18px;">
         <h2>Spending Distribution by Phase</h2>
         <canvas id="spendPhaseChart" style="max-height:220px;"></canvas>
         <div style="font-size:0.62rem;color:var(--muted2);margin-top:6px;">Bars show P10/P25/P50/P75/P90 discretionary spend per retirement phase.</div>
       </div>`
    );
    setTimeout(() => {
      const ctx = document.getElementById('spendPhaseChart');
      if (!ctx) return;
      if (window._spendPhaseChart) window._spendPhaseChart.destroy();
      const barColor = (pct, alpha) => {
        const colors = ['rgba(99,102,241,', 'rgba(14,165,233,', 'rgba(16,185,129,'];
        return pct < 3 ? colors[0]+alpha+')' : pct < 3 ? colors[1]+alpha+')' : colors[Math.min(pct,2)%3]+alpha+')';
      };
      window._spendPhaseChart = new Chart(ctx, {
        type: 'bar',
        data: {
          labels: ss2.labels,
          datasets: [
            { label: 'P90 Bull',  data: ss2.p90, backgroundColor: 'rgba(16,185,129,0.80)', borderRadius: 3 },
            { label: 'P75',       data: ss2.p75, backgroundColor: 'rgba(16,185,129,0.55)', borderRadius: 3 },
            { label: 'P50 Median',data: ss2.p50, backgroundColor: 'rgba(14,165,233,0.80)', borderRadius: 3 },
            { label: 'P25',       data: ss2.p25, backgroundColor: 'rgba(99,102,241,0.55)', borderRadius: 3 },
            { label: 'P10 Bear',  data: ss2.p10, backgroundColor: 'rgba(244,63,94,0.70)',  borderRadius: 3 },
          ]
        },
        options: {
          responsive: true,
          plugins: {
            legend: { labels: { color: '#8899b8', font: { size: 10 } } },
            tooltip: { callbacks: { label: c => '  ' + c.dataset.label + ': ' + fm(c.raw) } }
          },
          scales: {
            x: { ticks: { color: '#8899b8', font: { size: 10 } }, grid: { color: 'rgba(255,255,255,0.04)' } },
            y: { ticks: { color: '#526070', font: { size: 10 }, callback: v => v >= 1e6 ? '$'+(v/1e6).toFixed(1)+'M' : '$'+(v/1000).toFixed(0)+'k' }, grid: { color: 'rgba(255,255,255,0.04)' } }
          }
        }
      });
    }, 110);
  }

  // Ratchet stats panel + cumulative line chart
  const rs = d.ratchet_stats;
  if (rs) {
    const tierRow = (tier, pct, age) =>
      `<div class="stat-card"><div class="sc-label">Tier ${tier} (${tier === 1 ? '150' : tier === 2 ? '200' : '250'}%)</div>` +
      `<div class="sc-value" style="font-size:1rem;">${pct}%</div>` +
      `<div class="sc-sub">${age ? 'median age ' + age : 'never reached'}</div></div>`;
    document.getElementById('mc_t1').insertAdjacentHTML('beforeend',
      `<div class="panel accent-pink" style="margin-top:14px;padding:16px 18px;">
         <h2>Abundance Ratchet</h2>
         <div class="mc-stats-strip">
           ${tierRow(1, rs.tier1_pct, rs.median_tier1_age)}
           ${tierRow(2, rs.tier2_pct, rs.median_tier2_age)}
           ${tierRow(3, rs.tier3_pct, rs.median_tier3_age)}
         </div>
         <canvas id="ratchetPathChart" style="max-height:180px;margin-top:14px;"></canvas>
       </div>`
    );
    if (d.ratchet_paths) {
      setTimeout(() => {
        const ctx = document.getElementById('ratchetPathChart');
        if (!ctx) return;
        const rp = d.ratchet_paths;
        if (window._ratchetChart) window._ratchetChart.destroy();
        window._ratchetChart = new Chart(ctx, {
          type: 'line',
          data: {
            labels: rp.ages,
            datasets: [
              { label: 'Tier 1 (150%)', data: rp.t1, borderColor: '#ec4899', backgroundColor: 'rgba(236,72,153,0.08)', fill: true, tension: 0.3, pointRadius: 0 },
              { label: 'Tier 2 (200%)', data: rp.t2, borderColor: '#8b5cf6', backgroundColor: 'rgba(139,92,246,0.08)', fill: true, tension: 0.3, pointRadius: 0 },
              { label: 'Tier 3 (250%)', data: rp.t3, borderColor: '#6366f1', backgroundColor: 'rgba(99,102,241,0.08)', fill: true, tension: 0.3, pointRadius: 0 },
            ]
          },
          options: {
            responsive: true,
            plugins: {
              legend: { labels: { color: '#8899b8', font: { size: 10 } } },
              tooltip: { callbacks: { label: c => c.dataset.label + ': ' + c.raw + '% of trials' } }
            },
            scales: {
              x: { ticks: { color: '#526070', font: { size: 10 }, maxTicksLimit: 8 }, grid: { color: 'rgba(255,255,255,0.04)' } },
              y: { min: 0, max: 100, ticks: { color: '#526070', font: { size: 10 }, callback: v => v + '%' }, grid: { color: 'rgba(255,255,255,0.04)' } }
            }
          }
        });
      }, 100);
    }
  }

  // Prime Harvest stats panel
  const ph = d.prime_harvest_stats;
  if (ph) {
    document.getElementById('mc_t1').insertAdjacentHTML('beforeend',
      `<div class="panel accent-amber" style="margin-top:14px;padding:16px 18px;">
         <h2>Prime Harvest Buffer</h2>
         <div class="mc-stats-strip">
           <div class="stat-card"><div class="sc-label">Median Peak</div><div class="sc-value" style="font-size:1rem;">${fm(ph.median_peak)}</div></div>
           <div class="stat-card"><div class="sc-label">Fully Funded</div><div class="sc-value" style="font-size:1rem;">${ph.funded_pct}%</div><div class="sc-sub">${ph.median_funded_age ? 'by age ' + ph.median_funded_age : 'of trials'}</div></div>
           <div class="stat-card"><div class="sc-label">Median Drawn</div><div class="sc-value" style="font-size:1rem;">${fm(ph.median_drawn)}</div></div>
           <div class="stat-card"><div class="sc-label">Refill Cycles</div><div class="sc-value" style="font-size:1rem;">${ph.median_refills}</div><div class="sc-sub">${ph.recycled_pct}% had ≥1</div></div>
         </div>
       </div>`
    );
  }

  // SS timing sensitivity button
  document.getElementById('mc_t1').insertAdjacentHTML('beforeend',
    `<div style="margin-top:14px;"><button class="btn-sm" onclick="runSSSensitivity()" style="width:100%;">&#9200; SS Timing Sensitivity</button></div>
     <div id="ssSensResult"></div>`
  );

  // Sensitivity + Grid Search buttons
  document.getElementById('mc_t1').insertAdjacentHTML('beforeend',
    `<div style="margin-top:8px;display:flex;gap:8px;flex-wrap:wrap;">
       <button class="btn-sm" id="sensitivityBtn" onclick="runSensitivity()" style="flex:1;">&#127755; Sensitivity Analysis</button>
       <button class="btn-sm" id="gsOpenBtn" onclick="toggleGSPanel()" style="flex:1;">&#9783; Grid Search</button>
     </div>
     <div id="tornadoResult"></div>
     <div id="gsPanel" class="gs-panel" style="display:none;">
       <h3>&#9783; Deep-Scale Grid Search</h3>
       <div id="gsAxes"></div>
       <div style="margin-top:6px;">
         <button class="btn-sm" onclick="addGridAxis()" style="font-size:0.65rem;padding:5px 10px;">+ Add Axis</button>
       </div>
       <div id="gsComboPreview" class="gs-combo-preview"></div>
       <div style="margin:10px 0 6px;font-size:0.65rem;color:var(--muted2);">Trials per combination</div>
       <div style="display:flex;gap:12px;font-size:0.72rem;">
         <label><input type="radio" name="gsTrials" value="200" checked> 200</label>
         <label><input type="radio" name="gsTrials" value="500"> 500</label>
       </div>
       <button class="btn-sm" onclick="runGridSearch()" style="width:100%;margin-top:12px;background:var(--pink);color:#fff;">Run Grid Search</button>
       <div id="gsResults"></div>
     </div>`
  );

  // Compare strip
  if (scenarioA && scenarioA !== d) {
    const sa = scenarioA.stats || {}, sb = d.stats || {};
    const better = (a, b, higherBetter=true) => {
      if (a == null || b == null) return ['',''];
      return higherBetter
        ? [a > b ? 'compare-better' : (a < b ? 'compare-worse' : ''), b > a ? 'compare-better' : (b < a ? 'compare-worse' : '')]
        : [a < b ? 'compare-better' : (a > b ? 'compare-worse' : ''), b < a ? 'compare-better' : (b > a ? 'compare-worse' : '')];
    };
    const rows = [
      { label: 'Success Rate', a: scenarioA.success_pct + '%', b: d.success_pct + '%', cls: better(scenarioA.success_pct, d.success_pct) },
      { label: 'SS Claim Age', a: sa.median_ss_age, b: sb.median_ss_age, cls: better(sa.median_ss_age, sb.median_ss_age, false) },
      { label: 'Ripcord Rate', a: (sa.ripcord_rate||0) + '%', b: (sb.ripcord_rate||0) + '%', cls: better(sa.ripcord_rate, sb.ripcord_rate, false) },
      { label: 'Moat Breach', a: (sa.moat_breach_rate||0) + '%', b: (sb.moat_breach_rate||0) + '%', cls: better(sa.moat_breach_rate, sb.moat_breach_rate, false) },
      { label: 'Terminal Wealth', a: fm(sa.median_terminal||0), b: fm(sb.median_terminal||0), cls: better(sa.median_terminal, sb.median_terminal) },
      { label: 'Max Drawdown', a: (sa.median_drawdown||0) + '%', b: (sb.median_drawdown||0) + '%', cls: better(sa.median_drawdown, sb.median_drawdown, false) },
    ];
    const rowsHTML = rows.map(r =>
      `<tr><td>${r.label}</td><td class="${r.cls[0]}">${r.a}</td><td class="${r.cls[1]}">${r.b}</td></tr>`
    ).join('');
    const strip = `<div class="compare-strip">
      <h3>Scenario Compare</h3>
      <table class="compare-table">
        <thead><tr><th>Metric</th><th>Scenario A</th><th>Current</th></tr></thead>
        <tbody>${rowsHTML}</tbody>
      </table>
    </div>`;
    document.getElementById('mc_t1').insertAdjacentHTML('beforeend', strip);
  }

  window._lastMCResult = d;

  // ── Tab 2: Plan Details ───────────────────────────────────────────────────
  const t2 = document.getElementById('mc_t2');
  if (!t2) return;

  const warns = [];
  const notes = [];
  const goods = [];
  if (pct < 75)               warns.push(`Success rate ${pct}% is below the 75% threshold — plan requires action before retirement.`);
  if (s.ripcord_rate > 30)    warns.push(`${s.ripcord_rate}% of simulated paths trigger early SS — SGOV moat is likely undersized.`);
  if (s.moat_breach_rate > 20) warns.push(`Moat runs dry before SS activates in ${s.moat_breach_rate}% of trials — consider a larger bridge reserve.`);
  if (s.ripcord_rate > 10 && s.ripcord_rate <= 30) notes.push(`Early SS triggered in ${s.ripcord_rate}% of futures. A modest moat increase would reduce this.`);
  if (s.moat_breach_rate > 5 && s.moat_breach_rate <= 20) notes.push(`Moat breaches in ${s.moat_breach_rate}% of trials — within acceptable range but worth monitoring.`);
  if (s.median_drawdown > 40)  notes.push(`Median max drawdown ${s.median_drawdown}% — expect significant paper losses in adverse scenarios.`);
  if (s.tax_savings > 0)       goods.push(`Roth conversions save ~${fm(s.tax_savings)} median lifetime taxes vs. no-conversion baseline.`);
  if (pct >= 90)               goods.push(`At ${pct}% confidence your plan comfortably clears the institutional 90% solvency target.`);
  if (d.ratchet_stats && d.ratchet_stats.tier1_pct > 50) goods.push(`${d.ratchet_stats.tier1_pct}% of futures trigger Tier 1 abundance ratchet (150%) — strong upside engagement.`);

  const alertsHTML = [
    ...warns.map(w => `<div class="pd-warn">\u26a0\ufe0f ${w}</div>`),
    ...notes.map(n => `<div class="pd-note">\u26a0\ufe0f ${n}</div>`),
    ...goods.map(g => `<div class="pd-good">\u2713 ${g}</div>`),
  ].join('') || '<div class="pd-good">\u2713 No critical warnings detected.</div>';

  const bridgeYrs = (s.median_ss_age || 67) - (d.milestones?.[0]?.age || 62);
  const narrativeProse = `
    Based on ${(d.trial_count||1000).toLocaleString()} simulations, your portfolio arrives at retirement with a median of ${fm(s.median_arrival)}.
    The SGOV bridge moat covers expenses for ${bridgeYrs > 0 ? bridgeYrs + ' years' : 'the full bridge'} until Social Security activates at a median claim age of ${s.median_ss_age || 67}.
    In the median outcome, your portfolio reaches ${fm(s.median_terminal)} at age 95 after ${fm(s.median_total_spend || 0)} in total discretionary lifetime spending.
    ${s.ripcord_rate > 0 ? `In ${s.ripcord_rate}% of adverse scenarios, the moat runs short and SS is claimed early to prevent ruin.` : ''}
    ${s.conv_tax_paid > 0 ? `Roth conversion activity costs ${fm(s.conv_tax_paid)} median in total taxes but repositions assets out of RMD exposure.` : ''}
  `.replace(/\s+/g, ' ').trim();

  const rasHTML = d.ratchet_stats ? `
    <div class="pd-section">
      <h3>Abundance Ratchet</h3>
      <div class="pd-row"><span class="pd-key">Tier 1 (150% of retirement portfolio)</span><span class="pd-val">${d.ratchet_stats.tier1_pct}% of trials ${d.ratchet_stats.median_tier1_age ? '— median age ' + d.ratchet_stats.median_tier1_age : ''}</span></div>
      <div class="pd-row"><span class="pd-key">Tier 2 (200%)</span><span class="pd-val">${d.ratchet_stats.tier2_pct}% of trials ${d.ratchet_stats.median_tier2_age ? '— median age ' + d.ratchet_stats.median_tier2_age : ''}</span></div>
      <div class="pd-row"><span class="pd-key">Tier 3 (250%)</span><span class="pd-val">${d.ratchet_stats.tier3_pct}% of trials ${d.ratchet_stats.median_tier3_age ? '— median age ' + d.ratchet_stats.median_tier3_age : ''}</span></div>
    </div>` : '';

  const phHTML = d.prime_harvest_stats ? `
    <div class="pd-section">
      <h3>Prime Harvest Buffer</h3>
      <div class="pd-row"><span class="pd-key">Median moat peak</span><span class="pd-val">${fm(d.prime_harvest_stats.median_peak)}</span></div>
      <div class="pd-row"><span class="pd-key">Fully funded by target</span><span class="pd-val">${d.prime_harvest_stats.funded_pct}%${d.prime_harvest_stats.median_funded_age ? ' (age ' + d.prime_harvest_stats.median_funded_age + ')' : ''}</span></div>
      <div class="pd-row"><span class="pd-key">Median total drawn</span><span class="pd-val">${fm(d.prime_harvest_stats.median_drawn)}</span></div>
      <div class="pd-row"><span class="pd-key">Refill cycles</span><span class="pd-val">${d.prime_harvest_stats.median_refills} median (${d.prime_harvest_stats.recycled_pct}% had \u22651)</span></div>
    </div>` : '';

  t2.innerHTML = `
    <div class="pd-section">
      <h3>Bridge Overview</h3>
      <div class="pd-row"><span class="pd-key">Median arrival wealth</span><span class="pd-val">${fm(s.median_arrival)}</span></div>
      <div class="pd-row"><span class="pd-key">Median SS claim age</span><span class="pd-val">${s.median_ss_age}</span></div>
      <div class="pd-row"><span class="pd-key">Early SS (ripcord) rate</span><span class="pd-val">${s.ripcord_rate}%</span></div>
      <div class="pd-row"><span class="pd-key">Moat breach rate</span><span class="pd-val">${s.moat_breach_rate}%</span></div>
      <div class="pd-row"><span class="pd-key">Median max drawdown</span><span class="pd-val">${s.median_drawdown}%</span></div>
    </div>
    <div class="pd-section">
      <h3>Conditional Warnings</h3>
      ${alertsHTML}
    </div>
    <div class="pd-section">
      <div class="ai-playbook-hdr" style="margin-bottom:10px;">
        <h3 style="margin-bottom:0;">Plan Narrative</h3>
        <span class="ai-badge">AI</span>
      </div>
      <p class="pd-prose" id="aiPlanNarrativeBody">
        <span class="ai-thinking">
          <span class="ai-thinking-dots"><span></span><span></span><span></span></span>
          <span>Generating AI narrative\u2026</span>
        </span>
      </p>
    </div>
    ${rasHTML}
    ${phHTML}
  `;
  // Async AI Plan Details narrative
  (async () => {
    const model = document.getElementById('aiModel')?.value || 'llama3.1:8b';
    try {
      const r = await fetch('/api/summarize', {
        method: 'POST',
        headers: {'Content-Type':'application/json'},
        body: JSON.stringify({sim_data: d, summary_type: 'plan_details', model})
      });
      const sd = await r.json();
      const el = document.getElementById('aiPlanNarrativeBody');
      if (el) el.innerHTML = sd.error ? `\u26a0 ${sd.error}` : mdLite(sd.summary);
    } catch(_) {
      const el = document.getElementById('aiPlanNarrativeBody');
      if (el) el.innerHTML = '\u26a0 AI unavailable \u2014 check Ollama connection.';
    }
  })();
}

function switchMCTab(n) {
  document.getElementById('mc_t1').style.display = n === 0 ? '' : 'none';
  document.getElementById('mc_t2').style.display = n === 1 ? '' : 'none';
  document.querySelectorAll('.mc-itab').forEach((b, i) => b.classList.toggle('active', i === n));
}

function pinScenario() {
  scenarioA = window._lastMCResult;
  const btn = document.getElementById('pinScenBtn');
  if (btn) btn.textContent = 'Scenario A pinned \u2713';
}

// ── Section toggle helper ──────────────────────────────────────────────────
function sectionToggle(label, id, innerHTML, defaultOpen) {
  const openCls = defaultOpen ? ' open' : '';
  const maxH    = defaultOpen ? '2000px' : '0px';
  return `
    <button class="section-toggle${openCls}" onclick="
      const b=this; const bd=document.getElementById('${id}');
      b.classList.toggle('open');
      bd.style.maxHeight = b.classList.contains('open') ? '2000px' : '0px';
    ">${label}<span class="st-arrow">&#9654;</span></button>
    <div class="section-body" id="${id}" style="max-height:${maxH};">${innerHTML}</div>
  `;
}

// ── Allocation Donut Chart ─────────────────────────────────────────────────
const DONUT_COLORS = ['#ec4899','#6366f1','#0ea5e9','#f59e0b','#14b8a6','#8b5cf6','#10b981','#f43f5e','#84cc16','#f97316'];

function renderAllocChart(alloc, canvasId, legendId) {
  canvasId  = canvasId  || 'allocChart2';
  legendId  = legendId  || 'allocLegend2';
  const canvas = document.getElementById(canvasId);
  const legend = document.getElementById(legendId);
  if (!canvas || !alloc) return;
  const entries = Object.entries(alloc).filter(([,v]) => v > 0);
  if (!entries.length) return;
  const labels = entries.map(([k]) => k);
  const values = entries.map(([,v]) => v);
  const total  = values.reduce((a,b) => a+b, 0);
  const colors = entries.map((_, i) => DONUT_COLORS[i % DONUT_COLORS.length]);

  const key = canvasId + '_chart';
  if (window[key]) window[key].destroy();
  window[key] = new Chart(canvas, {
    type: 'doughnut',
    data: { labels, datasets: [{ data: values, backgroundColor: colors, borderWidth: 3, borderColor: '#080d1a', hoverBorderColor: '#131e30', hoverOffset: 6 }] },
    options: {
      responsive: false,
      plugins: {
        legend: { display: false },
        tooltip: {
          callbacks: {
            label: ctx => `${ctx.label}: ${fm(ctx.raw)} (${(ctx.raw/total*100).toFixed(1)}%)`
          },
          backgroundColor: 'rgba(30,41,59,0.95)', borderColor: 'rgba(51,65,85,0.8)', borderWidth: 1,
          titleColor: '#94a3b8', bodyColor: '#f1f5f9'
        }
      }
    }
  });

  if (legend) {
    legend.innerHTML = entries.map(([k,v], i) =>
      `<div class="dl-row"><span class="dl-dot" style="background:${colors[i]};"></span><span class="dl-name">${k}</span><span class="dl-val">${fm(v)}</span></div>`
    ).join('');
  }
}

// ── Roth Conversion Ladder ─────────────────────────────────────────────────
function calcRothLadder() {
  const trad0    = parseFloat(document.getElementById('rl_trad').value)    || 0;
  const curAge   = parseInt(document.getElementById('rl_age').value)        || 45;
  const retAge   = parseInt(document.getElementById('rl_ret').value)        || 62;
  const filing   = document.getElementById('rl_filing').value;
  const ssMo     = parseFloat(document.getElementById('rl_ss').value)       || 0;
  const stateTax = parseFloat(document.getElementById('rl_state').value)/100|| 0;
  const bracket  = parseFloat(document.getElementById('rl_bracket').value)  || 0.22;
  const retRate  = parseFloat(document.getElementById('rl_ret_rate').value)/100 || 0.08;

  // 2026 brackets (MFJ) — inflated 3%/yr
  const INFL = 0.03;
  const BASE_YEAR = 2026;
  const STD_DED   = filing === 'mfj' ? 31500 : 15750;
  // Brackets: [top of bracket, rate] for ordinary income
  const BRACKETS_MFJ    = [[23850,0.10],[96950,0.12],[206700,0.22],[394600,0.24],[501050,0.32],[751600,0.35],[Infinity,0.37]];
  const BRACKETS_SINGLE = [[11925,0.10],[48475,0.12],[103350,0.22],[197300,0.24],[250525,0.32],[626350,0.35],[Infinity,0.37]];
  const rawBrackets = filing === 'mfj' ? BRACKETS_MFJ : BRACKETS_SINGLE;

  function fedTaxAndSpace(taxableIncome, yr) {
    const yrs = yr - BASE_YEAR;
    const adj = Math.pow(1 + INFL, yrs);
    const bkts = rawBrackets.map(([top, r]) => [top === Infinity ? Infinity : top * adj, r]);
    let tax = 0, prev = 0, space = 0;
    for (const [top, r] of bkts) {
      const sliceTop = Math.min(top, taxableIncome);
      if (sliceTop > prev) tax += (sliceTop - prev) * r;
      if (r === bracket && space === 0) space = Math.max(0, top - Math.max(prev, taxableIncome));
      prev = top;
      if (prev >= taxableIncome) break;
    }
    return { tax, space };
  }

  let tradBal = trad0;
  let rothBal = 0;
  const rows  = [];
  const currentYear = new Date().getFullYear();

  for (let age = curAge; age <= 90 && tradBal > 0; age++) {
    const yr   = currentYear + (age - curAge);
    const adjStd = STD_DED * Math.pow(1 + INFL, yr - BASE_YEAR);
    const isRetired = age >= retAge;
    const ssIncome = isRetired && age >= 67 ? ssMo * 12 * 0.85 : 0; // 85% of SS is taxable
    const otherTaxable = ssIncome;
    const roomBeforeConv = Math.max(0, otherTaxable - adjStd);
    const { space } = fedTaxAndSpace(roomBeforeConv, yr);

    let convert = Math.min(space, tradBal);
    convert = Math.max(0, convert);

    const taxableInc = roomBeforeConv + convert;
    const { tax: fedTax } = fedTaxAndSpace(taxableInc, yr);
    const priorFedTax = fedTaxAndSpace(roomBeforeConv, yr).tax;
    const convFedTax  = Math.max(0, fedTax - priorFedTax);
    const convStateTax = convert * stateTax;

    tradBal = (tradBal - convert) * (1 + retRate);
    rothBal = (rothBal + convert) * (1 + retRate);

    rows.push({ age, yr, convert, fedTax: convFedTax, stateTax: convStateTax, roth: rothBal, trad: tradBal });
  }

  const el = document.getElementById('rl_results');
  if (!rows.length) { el.innerHTML = '<div class="rl-result-hint">No Traditional balance to convert.</div>'; return; }

  const tblRows = rows.map(r => `
    <tr>
      <td>${r.age}</td><td>${r.yr}</td>
      <td class="h-r">${r.convert > 0 ? fm(r.convert) : '—'}</td>
      <td class="h-r">${r.fedTax > 0 ? fm(r.fedTax) : '—'}</td>
      <td class="h-r">${r.stateTax > 0 ? fm(r.stateTax) : '—'}</td>
      <td class="h-r clr-green">${fm(r.roth)}</td>
      <td class="h-r clr-red">${fm(r.trad)}</td>
    </tr>`).join('');

  el.innerHTML = `
    <table class="port-htable" style="width:100%;font-size:0.78rem;">
      <thead><tr>
        <th>Age</th><th>Year</th><th class="h-r">Convert</th><th class="h-r">Fed Tax</th><th class="h-r">State Tax</th><th class="h-r">Roth Bal</th><th class="h-r">Trad Rem.</th>
      </tr></thead>
      <tbody>${tblRows}</tbody>
    </table>
    <div style="margin-top:16px;"><canvas id="rothChart" style="max-height:220px;"></canvas></div>
  `;

  setTimeout(() => renderRothChart(rows), 60);
}

function renderRothChart(rows) {
  const canvas = document.getElementById('rothChart');
  if (!canvas) return;
  if (window._rothChart) window._rothChart.destroy();
  window._rothChart = new Chart(canvas, {
    type: 'bar',
    data: {
      labels: rows.map(r => r.age),
      datasets: [
        { label: 'Roth', data: rows.map(r => r.roth), backgroundColor: 'rgba(16,185,129,0.80)', hoverBackgroundColor: '#10b981', borderRadius: 3, stack: 'bal' },
        { label: 'Trad', data: rows.map(r => r.trad), backgroundColor: 'rgba(244,63,94,0.60)',  hoverBackgroundColor: '#f43f5e', borderRadius: 3, stack: 'bal' },
      ]
    },
    options: {
      responsive: true,
      plugins: {
        legend: { labels: { color: '#94a3b8', font: { family: 'system-ui', size: 11 } } },
        tooltip: {
          backgroundColor: 'rgba(30,41,59,0.95)', borderColor: 'rgba(51,65,85,0.8)', borderWidth: 1,
          titleColor: '#94a3b8', bodyColor: '#f1f5f9',
          callbacks: { label: ctx => `${ctx.dataset.label}: ${fm(ctx.raw)}` }
        }
      },
      scales: {
        x: { stacked: true, ticks: { color: '#64748b', font: { family: 'ui-monospace', size: 10 } }, grid: { color: 'rgba(51,65,85,0.4)' } },
        y: { stacked: true, ticks: { color: '#64748b', callback: v => '$'+(v/1000).toFixed(0)+'k', font: { family: 'ui-monospace', size: 10 } }, grid: { color: 'rgba(51,65,85,0.4)' } }
      }
    }
  });
}

// ── getParams — universal MC param reader ─────────────────────────────────
function getParams() {
  const trials = parseInt(document.querySelector('input[name="mc_trials"]:checked')?.value || 1000);
  const seq_yr  = parseInt(document.querySelector('input[name="mc_seq"]:checked')?.value  || 0);
  return {
    current_age:         +document.getElementById('mc_age').value,
    target_age:          +document.getElementById('mc_ret_age').value,
    filing_status:        document.getElementById('mc_filing').value,
    start_engine:        +document.getElementById('mc_engine').value,
    start_sgov:          +document.getElementById('mc_sgov').value,
    start_checking:      +document.getElementById('mc_checking').value,
    annual_contribution: +document.getElementById('mc_contrib').value,
    full_ss:             +document.getElementById('mc_full_ss').value,
    mean_return:         +document.getElementById('mc_mu').value / 100,
    volatility:          +document.getElementById('mc_sigma').value / 100,
    sgov_yield:          +document.getElementById('mc_syld').value / 100,
    inflation_rate:      +document.getElementById('mc_infl').value / 100,
    strict_moat_cost:    +document.getElementById('mc_floor').value,
    moat_target:         +document.getElementById('mc_moat_target').value,
    use_conversion:       document.getElementById('mc_use_conv').checked,
    trad_balance:        +document.getElementById('mc_trad').value,
    target_bracket:      +document.getElementById('mc_tgt_bkt').value,
    state_tax_rate:      +document.getElementById('mc_state_tx').value / 100,
    use_ss_haircut:       document.getElementById('mc_haircut').checked,
    use_tail_shock:       document.getElementById('mc_tail').checked,
    use_mortality_weighting: document.getElementById('mc_mort').checked,
    use_stochastic_inflation: document.getElementById('mc_use_si').checked,
    inflation_volatility: +(document.getElementById('mc_infl_vol')?.value || 1) / 100,
    inflation_min:        +(document.getElementById('mc_infl_min')?.value || 1) / 100,
    inflation_max:        +(document.getElementById('mc_infl_max')?.value || 8) / 100,
    stagflation_corr:     +(document.getElementById('mc_stag_corr')?.value || 30) / 100,
    use_aca_shock:        document.getElementById('mc_use_aca').checked,
    aca_shock_prob:       +(document.getElementById('mc_aca_prob')?.value || 30) / 100,
    aca_shock_mag:        +(document.getElementById('mc_aca_mag')?.value || 15000),
    use_tax_reversion:    document.getElementById('mc_use_taxrev').checked,
    tax_risk_near:        +(document.getElementById('mc_tax_near')?.value || 20) / 100,
    tax_risk_mid:         +(document.getElementById('mc_tax_mid')?.value || 40) / 100,
    tax_risk_late:        +(document.getElementById('mc_tax_late')?.value || 60) / 100,
    trials: trials, seq_shock_year: seq_yr,
    ss_age: 67, ss_haircut_pct: 0.21, phase3_moat_years: 2, euphoric_offset: 0.03,
    gogo_e:0.25, gogo_n:0.15, slowgo_e:0.20, slowgo_n:0.10, nogo_e:0.10, nogo_n:0.05,
    gk_trigger:0.20, gk_cut_rate:0.50, bear_streak_years:3, bear_streak_cut:0.25,
    portfolio_cap:5000000, cap_inflation:0.03, cap_gogo:0.10, cap_slowgo:0.05, cap_nogo:0.02,
    tent_skim_rate:0.50, dividend_yield:0.015, wage_growth:0.02,
    tail_shock_return:-0.25, tail_shock_count:1,
  };
}

// ── Net Worth History Chart ───────────────────────────────────────────────
function renderNWHistoryChart(hist) {
  if (!hist || hist.length < 1) return;
  const panel = document.getElementById('nwhPanel');
  if (panel) panel.style.display = '';
  const canvas = document.getElementById('nwhChart');
  if (!canvas) return;
  if (window._nwhChart) window._nwhChart.destroy();
  window._nwhChart = new Chart(canvas, {
    type: 'line',
    data: {
      labels: hist.map(h => h.label),
      datasets: [
        { label: 'Liquid NW', data: hist.map(h => h.lnw), borderColor: '#ec4899', backgroundColor: 'rgba(236,72,153,0.08)', fill: true, tension: 0.3, pointRadius: hist.length > 30 ? 0 : 3 },
        { label: 'Total NW',  data: hist.map(h => h.tnw), borderColor: '#6366f1', backgroundColor: 'rgba(99,102,241,0.06)', fill: true, tension: 0.3, pointRadius: hist.length > 30 ? 0 : 3 },
      ]
    },
    options: {
      responsive: true,
      plugins: {
        legend: { labels: { color: '#8899b8', font: { size: 11 } } },
        tooltip: {
          backgroundColor: 'rgba(9,13,26,0.92)', titleColor: '#8899b8', bodyColor: '#f0f4ff',
          padding: 10, cornerRadius: 8,
          callbacks: { label: ctx => `  ${ctx.dataset.label}: ${fm(ctx.raw)}` }
        }
      },
      scales: {
        x: { ticks: { color: '#526070', maxTicksLimit: 8, font: { size: 10 } }, grid: { color: 'rgba(255,255,255,0.04)' } },
        y: { ticks: { color: '#526070', callback: v => v >= 1e6 ? '$'+(v/1e6).toFixed(1)+'M' : '$'+(v/1000).toFixed(0)+'k', font: { size: 10 } }, grid: { color: 'rgba(255,255,255,0.04)' } }
      }
    }
  });
}
function clearNWHistory() {
  localStorage.removeItem('retAdv_nwHistory');
  const panel = document.getElementById('nwhPanel');
  if (panel) panel.style.display = 'none';
  if (window._nwhChart) { window._nwhChart.destroy(); window._nwhChart = null; }
}

// ── Time-to-FI Calculator ────────────────────────────────────────────────
function calcTimeToFI(lnw, fiTarget, annualContrib, returnRate) {
  if (lnw <= 0 || fiTarget <= 0 || lnw >= fiTarget) return null;
  const r = returnRate / 12;
  const monthly = (annualContrib || 0) / 12;
  let w = lnw, months = 0;
  while (w < fiTarget && months < 600) {
    w = w * (1 + r) + monthly;
    months++;
  }
  return months < 600 ? months : null;
}

function calcContribTFI(lnwNum, targetNum, rules) {
  const weeklyTaxable = parseFloat(document.getElementById('ci_taxable')?.value)  || 0;
  const weeklyIRA     = parseFloat(document.getElementById('ci_ira')?.value)       || 0;
  const emp401kPct    = parseFloat(document.getElementById('ci_401k_pct')?.value)  || 0;
  const matchPct      = parseFloat(document.getElementById('ci_match_pct')?.value) || 0;
  const annualGross   = parseFloat(document.getElementById('ci_gross')?.value)     || 0;

  localStorage.setItem('retAdv_contrib', JSON.stringify({weeklyTaxable, weeklyIRA, emp401kPct, matchPct, annualGross}));

  // IRS limits
  const iraLimit      = rules.ira_roth_limit || 7500;
  const k401EmpLimit  = rules.contrib_401k   || 24500;
  const k401Combined  = 70000; // 2026 combined employee + employer cap

  // Taxable — no cap
  const annualTaxable = weeklyTaxable * 52;

  // IRA — capped at annual limit; contributions stop when limit is hit mid-year
  const rawIRA    = weeklyIRA * 52;
  const cappedIRA = Math.min(rawIRA, iraLimit);
  const iraCapped = rawIRA > iraLimit;

  // 401k employee — capped at employee elective deferral limit
  const raw401kEmp    = annualGross * emp401kPct / 100;
  const capped401kEmp = Math.min(raw401kEmp, k401EmpLimit);
  const empCapped     = raw401kEmp > k401EmpLimit;

  // 401k employer match — applied to actual gross %, but capped by combined limit
  const raw401kMatch    = annualGross * matchPct / 100;
  const headroom        = Math.max(0, k401Combined - capped401kEmp);
  const capped401kMatch = Math.min(raw401kMatch, headroom);
  const matchCapped     = raw401kMatch > headroom;

  const total = annualTaxable + cappedIRA + capped401kEmp + capped401kMatch;

  const wrap = document.getElementById('tfiCardWrap');
  if (!wrap) return;

  const tfiMonths = calcTimeToFI(lnwNum, targetNum, total, 0.09);
  if (tfiMonths === null) { wrap.innerHTML = ''; return; }

  const tfiYr    = Math.floor(tfiMonths / 12);
  const tfiMo    = tfiMonths % 12;
  const tfiDate  = new Date();
  tfiDate.setMonth(tfiDate.getMonth() + tfiMonths);
  const tfiLabel = tfiDate.toLocaleDateString('en-US', {month:'short', year:'numeric'});

  const breakdown = [
    annualTaxable > 0 ? `Taxable $${Math.round(annualTaxable).toLocaleString()}` : null,
    cappedIRA     > 0 ? `IRA $${Math.round(cappedIRA).toLocaleString()}${iraCapped?' (limit)':''}` : null,
    capped401kEmp > 0 ? `401k $${Math.round(capped401kEmp).toLocaleString()}${empCapped?' (limit)':''}` : null,
    capped401kMatch>0 ? `Match $${Math.round(capped401kMatch).toLocaleString()}${matchCapped?' (limit)':''}` : null,
  ].filter(Boolean).join(' &bull; ');

  wrap.innerHTML = `<div class="stat-grid" style="margin-top:10px;">
    <div class="stat-card accent-pink">
      <div class="sc-label">Time to FI (w/ contributions)</div>
      <div class="sc-value">${tfiYr}y ${tfiMo}mo</div>
      <div class="sc-sub">Est. ${tfiLabel} &bull; +$${Math.round(total).toLocaleString()}/yr &bull; 9% return</div>
      ${breakdown ? `<div class="sc-sub" style="margin-top:3px;font-size:0.68rem;opacity:0.7;">${breakdown}</div>` : ''}
    </div>
  </div>`;
}

// ── Budget vs Actual ──────────────────────────────────────────────────────
let _lastSpending = null;
let _lastMonths   = null;
function renderBudgetTable(spending, months) {
  _lastSpending = spending;
  _lastMonths   = months;
  const budget = JSON.parse(localStorage.getItem('retAdv_budget') || '{}');
  const el = document.getElementById('budgetTable');
  if (!el) return;
  const cats = Object.entries(spending || {}).filter(([k]) =>
    k !== 'Income' && k !== 'TRUE SAVING RATE' && k !== 'SAVINGS RATE'
  );
  if (!cats.length) return;
  // Use most recent month value
  const rows = cats.map(([cat, vals]) => {
    const actual = typeof vals[0] === 'number' ? Math.abs(vals[0]) : 0;
    const bgt    = parseFloat(budget[cat] || 0);
    const delta  = bgt > 0 ? bgt - actual : null;
    const deltaCls = delta === null ? '' : delta >= 0 ? 'delta-under' : 'delta-over';
    const deltaStr = delta === null ? '—' : (delta >= 0 ? '&#9660; Under ' : '&#9650; Over ') + fm(Math.abs(delta));
    return `<tr>
      <td>${cat}</td>
      <td>${bgt > 0 ? fm(bgt) : '<span style="color:var(--muted2)">—</span>'}</td>
      <td>${actual > 0 ? fm(actual) : '—'}</td>
      <td class="${deltaCls}">${deltaStr}</td>
    </tr>`;
  }).join('');
  const recentMonth = (months || []).length ? months[0] : 'Recent';
  el.innerHTML = `<table class="budget-table">
    <thead><tr><th>Category</th><th>Budget/mo</th><th>Actual (${recentMonth})</th><th>Delta</th></tr></thead>
    <tbody>${rows}</tbody>
  </table>`;
}
function toggleBudgetEdit() {
  const form = document.getElementById('budgetEditForm');
  if (!form) return;
  if (form.style.display === 'none') {
    const budget = JSON.parse(localStorage.getItem('retAdv_budget') || '{}');
    const cats = Object.entries(_lastSpending || {}).filter(([k]) =>
      k !== 'Income' && k !== 'TRUE SAVING RATE' && k !== 'SAVINGS RATE'
    );
    const inp = document.getElementById('budgetInputs');
    if (inp) inp.innerHTML = cats.map(([cat]) =>
      `<div class="field">
        <label>${cat}</label>
        <input type="number" id="bgt_${cat.replace(/[^a-z0-9]/gi,'_')}" value="${budget[cat] || ''}" placeholder="0" min="0">
      </div>`
    ).join('');
    form.style.display = '';
  } else {
    form.style.display = 'none';
  }
}
function saveBudget() {
  const budget = {};
  const cats = Object.keys(_lastSpending || {}).filter(k =>
    k !== 'Income' && k !== 'TRUE SAVING RATE' && k !== 'SAVINGS RATE'
  );
  for (const cat of cats) {
    const el = document.getElementById('bgt_' + cat.replace(/[^a-z0-9]/gi,'_'));
    if (el && el.value) budget[cat] = parseFloat(el.value);
  }
  localStorage.setItem('retAdv_budget', JSON.stringify(budget));
  document.getElementById('budgetEditForm').style.display = 'none';
  renderBudgetTable(_lastSpending, _lastMonths);
}

// ── Save / Load MC Scenarios ──────────────────────────────────────────────
function saveScenario() {
  const name = (document.getElementById('scenNameInput')?.value || '').trim();
  if (!name) return;
  if (!window._lastMCResult) { alert('Run a simulation first.'); return; }
  const scenarios = JSON.parse(localStorage.getItem('retAdv_scenarios') || '[]');
  const existing  = scenarios.findIndex(s => s.name === name);
  const entry = { name, params: getParams(), result: window._lastMCResult, saved: new Date().toLocaleDateString() };
  if (existing >= 0) scenarios[existing] = entry;
  else scenarios.push(entry);
  localStorage.setItem('retAdv_scenarios', JSON.stringify(scenarios));
  renderSavedScenarios();
}
function renderSavedScenarios() {
  const el = document.getElementById('savedScenariosList');
  if (!el) return;
  const scenarios = JSON.parse(localStorage.getItem('retAdv_scenarios') || '[]');
  if (!scenarios.length) { el.innerHTML = '<div style="color:var(--muted2);font-size:0.72rem;">No saved scenarios yet.</div>'; return; }
  el.innerHTML = scenarios.map((s, i) => {
    const pct2 = s.result?.success_pct ?? '—';
    const cls = pct2 >= 90 ? 'clr-green' : pct2 >= 75 ? 'clr-amber' : 'clr-red';
    return `<div class="scenario-item">
      <span class="sn-pct ${cls}">${pct2}%</span>
      <span class="sn-name">${s.name}</span>
      <span class="sn-date">${s.saved}</span>
      <button class="btn-sm" onclick="loadSavedScenario(${i})">Load</button>
      <button class="btn-sm" onclick="deleteScenario(${i})" style="color:var(--red);">&#10005;</button>
    </div>`;
  }).join('');
}
function loadSavedScenario(idx) {
  const scenarios = JSON.parse(localStorage.getItem('retAdv_scenarios') || '[]');
  const entry = scenarios[idx];
  if (!entry) return;
  const p = entry.params;
  const setV = (id, v) => { const el = document.getElementById(id); if (el) el.value = v; };
  setV('mc_age',     p.current_age);
  setV('mc_ret_age', p.target_age);
  setV('mc_engine',  p.start_engine);
  setV('mc_sgov',    p.start_sgov);
  setV('mc_checking',p.start_checking);
  setV('mc_contrib', p.annual_contribution);
  setV('mc_full_ss', p.full_ss);
  setV('mc_mu',      Math.round((p.mean_return || 0.09) * 100));
  setV('mc_sigma',   Math.round((p.volatility  || 0.16) * 100));
  setV('mc_syld',    Math.round((p.sgov_yield  || 0.04) * 100));
  setV('mc_infl',    Math.round((p.inflation_rate || 0.03) * 100));
  setV('mc_floor',   p.strict_moat_cost);
  document.getElementById('mc_filing').value = p.filing_status || 'single';
  if (entry.result) renderMCResults(entry.result);
}
function deleteScenario(idx) {
  const scenarios = JSON.parse(localStorage.getItem('retAdv_scenarios') || '[]');
  scenarios.splice(idx, 1);
  localStorage.setItem('retAdv_scenarios', JSON.stringify(scenarios));
  renderSavedScenarios();
}

// ── Contribution Optimizer ────────────────────────────────────────────────
async function optimizeContribution() {
  const el = document.getElementById('contribResult');
  if (!el) return;
  el.innerHTML = '<div style="color:var(--muted2);font-size:0.72rem;margin-top:8px;">Optimizing&hellip; (~5s)</div>';
  try {
    const params = getParams();
    params.target_success_pct = 95;
    const r = await fetch('/api/optimize-contribution', {
      method: 'POST', headers: {'Content-Type':'application/json'}, body: JSON.stringify(params)
    });
    const d = await r.json();
    if (d.error) { el.innerHTML = `<div class="contrib-result" style="color:var(--red);">${d.error}</div>`; return; }
    el.innerHTML = `<div class="contrib-result">
      <span style="color:var(--muted);">Annual contribution for 95% success:</span>
      <span class="cr-val">${fm(d.optimal_contribution)}/yr</span>
      <span style="color:var(--muted2);font-size:0.68rem;">Achieved: ${d.achieved_success_pct}%</span>
    </div>`;
  } catch(e) {
    el.innerHTML = `<div class="contrib-result" style="color:var(--red);">Error: ${e.message}</div>`;
  }
}

// ── Grid Search ───────────────────────────────────────────────────────────
const GS_PARAMS = [
  {key:'retirement_age',    label:'Retire Age',        def:62,    min:55, max:70,  step:1},
  {key:'annual_contribution',label:'Annual Contrib ($)', def:19341, min:5000,max:50000,step:1000},
  {key:'moat_target',       label:'Moat Target ($)',   def:300000,min:100000,max:500000,step:25000},
  {key:'mean_return',       label:'Mean Return (%)',   def:9,     min:5,  max:14,  step:0.5},
  {key:'inflation_rate',    label:'Inflation (%)',     def:3,     min:2,  max:6,   step:0.5},
  {key:'strict_moat_cost',  label:'Annual Draw ($)',   def:54292, min:30000,max:80000,step:2000},
];

let gsAxisCount = 0;

function toggleGSPanel() {
  const p = document.getElementById('gsPanel');
  p.style.display = p.style.display === 'none' ? '' : 'none';
  if (p.style.display !== 'none' && gsAxisCount === 0) addGridAxis();
}

function addGridAxis() {
  if (gsAxisCount >= 3) return;
  gsAxisCount++;
  const id = `gsAxis_${gsAxisCount}`;
  const opts = GS_PARAMS.map(p => `<option value="${p.key}">${p.label}</option>`).join('');
  const def = GS_PARAMS[0];
  document.getElementById('gsAxes').insertAdjacentHTML('beforeend', `
    <div class="gs-axis-row" id="${id}">
      <select onchange="gsAxisChanged('${id}')">${opts}</select>
      <div>
        <div class="gs-axis-label">Min</div>
        <input type="number" class="gs-min" value="${def.min}" step="${def.step}" onchange="updateGSPreview()">
      </div>
      <div>
        <div class="gs-axis-label">Max</div>
        <input type="number" class="gs-max" value="${def.max}" step="${def.step}" onchange="updateGSPreview()">
      </div>
      <div>
        <div class="gs-axis-label">Step</div>
        <input type="number" class="gs-step" value="${def.step}" step="${def.step}" onchange="updateGSPreview()">
      </div>
      <button onclick="removeGridAxis('${id}')" style="background:none;border:none;color:var(--muted2);cursor:pointer;font-size:1.1rem;">&#215;</button>
    </div>
  `);
  updateGSPreview();
}

function gsAxisChanged(id) {
  const row = document.getElementById(id);
  const key = row.querySelector('select').value;
  const p = GS_PARAMS.find(x => x.key === key);
  if (!p) return;
  row.querySelector('.gs-min').value = p.min;
  row.querySelector('.gs-max').value = p.max;
  row.querySelector('.gs-step').value = p.step;
  updateGSPreview();
}

function removeGridAxis(id) {
  document.getElementById(id)?.remove();
  gsAxisCount = Math.max(0, gsAxisCount - 1);
  updateGSPreview();
}

function getGSAxes() {
  return [...document.querySelectorAll('#gsAxes .gs-axis-row')].map(row => {
    const key = row.querySelector('select').value;
    const mn = parseFloat(row.querySelector('.gs-min').value);
    const mx = parseFloat(row.querySelector('.gs-max').value);
    const st = parseFloat(row.querySelector('.gs-step').value) || 1;
    const vals = [];
    for (let v = mn; v <= mx + 1e-9; v = Math.round((v + st) * 1e9) / 1e9) vals.push(v);
    return {key, vals};
  }).filter(a => a.vals.length > 0);
}

function updateGSPreview() {
  const axes = getGSAxes();
  const counts = axes.map(a => a.vals.length);
  const total = counts.reduce((a, b) => a * b, 1);
  const parts = counts.map((c, i) => `${axes[i] ? axes[i].vals.length : 0}`).join(' × ');
  const el = document.getElementById('gsComboPreview');
  el.textContent = counts.length > 0
    ? `${parts} = ${total} combination${total !== 1 ? 's' : ''}${total > 300 ? ' ⚠ max 300' : ''}`
    : '';
}

async function runGridSearch() {
  const axes = getGSAxes();
  if (axes.length === 0) { alert('Add at least one axis.'); return; }
  const trials = parseInt(document.querySelector('input[name="gsTrials"]:checked')?.value || '200');
  const base_params = getParams();
  base_params.trials = trials;
  const grid_axes = {};
  axes.forEach(a => { grid_axes[a.key] = a.vals; });

  const el = document.getElementById('gsResults');
  el.innerHTML = '<div style="color:var(--muted2);font-size:0.72rem;padding:12px 0;">Running grid search&#8230;</div>';
  try {
    const r = await fetch('/api/grid-search', {
      method: 'POST', headers: {'Content-Type':'application/json'},
      body: JSON.stringify({base_params, grid_axes})
    });
    const d = await r.json();
    if (d.error) { el.innerHTML = `<div style="color:var(--red);font-size:0.72rem;">${d.error}</div>`; return; }
    renderGridResults(d, axes);
  } catch(e) {
    el.innerHTML = `<div style="color:var(--red);font-size:0.72rem;">Error: ${e.message}</div>`;
  }
}

function renderGridResults(d, axes) {
  const el = document.getElementById('gsResults');
  const res = d.results || [];
  if (!res.length) { el.innerHTML = '<div style="color:var(--muted2);font-size:0.72rem;padding:8px 0;">No results.</div>'; return; }

  const axisNames = d.axis_names || axes.map(a => a.key);
  const paramLabel = k => (GS_PARAMS.find(p => p.key === k) || {label: k}).label;

  // Table
  let thead = '<tr>' + axisNames.map(k => `<th>${paramLabel(k)}</th>`).join('') +
    '<th>Success%</th><th>Terminal</th><th>Ripcord%</th></tr>';
  let tbody = res.map((row, i) => {
    const cls = i < 3 ? ' class="gs-top"' : '';
    const pvals = axisNames.map(k => `<td>${row[k] != null ? row[k] : '—'}</td>`).join('');
    const sucColor = row.success_pct >= 90 ? 'var(--green)' : row.success_pct >= 75 ? 'var(--amber)' : 'var(--red)';
    return `<tr${cls}>${pvals}
      <td style="color:${sucColor};font-weight:700;">${row.success_pct}%</td>
      <td>${fm(row.median_terminal || 0)}</td>
      <td style="color:${(row.ripcord_rate||0)>20?'var(--red)':'var(--muted2)'};">${row.ripcord_rate||0}%</td>
    </tr>`;
  }).join('');

  let html = `<div style="font-size:0.65rem;color:var(--muted2);margin:10px 0 6px;">
    ${res.length} results · ${d.runtime_ms}ms · top 3 highlighted
  </div>
  <div style="overflow-x:auto;">
    <table class="gs-result-table"><thead>${thead}</thead><tbody>${tbody}</tbody></table>
  </div>`;

  // Heatmap for exactly 2 axes
  if (axisNames.length === 2) {
    html += `<div class="gs-heatmap-wrap">
      <div style="margin:12px 0 6px;display:flex;align-items:center;justify-content:space-between;">
        <span style="font-size:0.65rem;color:var(--muted2);">Heatmap — Success %</span>
        <button class="btn-sm" onclick="toggleGSHeatmap()" style="font-size:0.62rem;padding:4px 8px;">&#128200; Toggle Heatmap</button>
      </div>
      <div id="gsHeatmapWrap" style="display:none;">
        <canvas id="gsHeatmapCanvas" style="width:100%;max-height:240px;"></canvas>
      </div>
    </div>`;
    // Store data for heatmap rendering
    window._gsHeatData = {results: res, axes: axisNames, axisVals: axes.map(a => a.vals)};
  }

  el.innerHTML = html;
}

function toggleGSHeatmap() {
  const wrap = document.getElementById('gsHeatmapWrap');
  if (!wrap) return;
  const showing = wrap.style.display !== 'none';
  wrap.style.display = showing ? 'none' : '';
  if (!showing && window._gsHeatData) renderGSHeatmap();
}

function renderGSHeatmap() {
  const {results, axes, axisVals} = window._gsHeatData;
  const canvas = document.getElementById('gsHeatmapCanvas');
  if (!canvas || !window.Chart) return;
  const [ax0, ax1] = axes;
  const vals0 = [...new Set(results.map(r => r[ax0]))].sort((a,b)=>a-b);
  const vals1 = [...new Set(results.map(r => r[ax1]))].sort((a,b)=>a-b);
  const lookup = {};
  results.forEach(r => { lookup[`${r[ax0]}_${r[ax1]}`] = r.success_pct; });
  const data = [];
  vals0.forEach((v0, xi) => {
    vals1.forEach((v1, yi) => {
      data.push({x: xi, y: yi, v: lookup[`${v0}_${v1}`] ?? 0});
    });
  });
  if (canvas._chartInst) canvas._chartInst.destroy();
  canvas._chartInst = new Chart(canvas, {
    type: 'matrix',
    data: { datasets: [{
      data,
      backgroundColor(ctx) {
        const v = ctx.dataset.data[ctx.dataIndex]?.v ?? 0;
        const g = Math.round((v / 100) * 180);
        return `rgba(16,${g},129,${0.2 + v/100*0.6})`;
      },
      width(ctx) { return (ctx.chart.chartArea?.width / vals0.length) - 2 || 20; },
      height(ctx) { return (ctx.chart.chartArea?.height / vals1.length) - 2 || 20; },
    }]},
    options: {
      responsive: true, maintainAspectRatio: false,
      plugins: { legend: {display:false}, tooltip: { callbacks: { label(c) { const d=c.dataset.data[c.dataIndex]; return `${vals0[d.x]} × ${vals1[d.y]} → ${d.v}%`; }}}},
      scales: {
        x: { type:'linear', min:-0.5, max:vals0.length-0.5, ticks:{callback:(_,i)=>vals0[i]??''}, title:{display:true,text:(GS_PARAMS.find(p=>p.key===ax0)||{label:ax0}).label,color:'var(--muted2)',font:{size:10}}},
        y: { type:'linear', min:-0.5, max:vals1.length-0.5, ticks:{callback:(_,i)=>vals1[i]??''}, title:{display:true,text:(GS_PARAMS.find(p=>p.key===ax1)||{label:ax1}).label,color:'var(--muted2)',font:{size:10}}}
      }
    }
  });
}

// ── Sensitivity Tornado Chart ─────────────────────────────────────────────
async function runSensitivity() {
  const btn = document.getElementById('sensitivityBtn');
  const el  = document.getElementById('tornadoResult');
  if (!el) return;
  if (btn) btn.disabled = true;
  el.innerHTML = '<div style="color:var(--muted2);font-size:0.72rem;margin-top:8px;">Running sensitivity analysis (~2s)&hellip;</div>';
  try {
    const r = await fetch('/api/sensitivity', {
      method: 'POST', headers: {'Content-Type':'application/json'}, body: JSON.stringify(getParams())
    });
    const d = await r.json();
    if (d.error) { el.innerHTML = `<div style="color:var(--red);font-size:0.75rem;margin-top:8px;">${d.error}</div>`; return; }
    renderTornadoChart(d);
  } catch(e) {
    el.innerHTML = `<div style="color:var(--red);font-size:0.75rem;margin-top:8px;">Error: ${e.message}</div>`;
  } finally {
    if (btn) btn.disabled = false;
  }
}
async function runSSSensitivity() {
  const el = document.getElementById('ssSensResult');
  if (!el) return;
  el.innerHTML = '<div style="padding:10px;font-size:0.72rem;color:var(--muted);">Running SS sensitivity (5 simulations)…</div>';
  const params = getParams();
  params.trials = Math.min(params.trials || 500, 500);
  try {
    const r = await fetch('/api/ss-sensitivity', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(params) });
    const d = await r.json();
    if (d.error) { el.innerHTML = `<span class="err">${d.error}</span>`; return; }
    const rows = d.results.map(row => {
      const dcls = row.delta_from_67 > 0 ? 'clr-green' : row.delta_from_67 < 0 ? 'clr-red' : '';
      const ageCls = row.ss_age < 67 ? 'clr-red' : row.ss_age > 67 ? 'clr-green' : '';
      return `<tr>
        <td class="${ageCls}">Age ${row.ss_age}</td>
        <td>${row.success_pct}%</td>
        <td>${fm(row.median_terminal)}</td>
        <td class="${dcls}">${row.delta_from_67 >= 0 ? '+' : ''}${fm(row.delta_from_67)}</td>
      </tr>`;
    }).join('');
    el.innerHTML = `<div class="panel accent-indigo" style="margin-top:8px;padding:16px 18px;">
      <h2>SS Claim Age Sensitivity</h2>
      <table class="mc-table">
        <thead><tr><th>Claim Age</th><th>Success %</th><th>Terminal (P50)</th><th>vs Age 67</th></tr></thead>
        <tbody>${rows}</tbody>
      </table>
      <div style="font-size:0.65rem;color:var(--muted2);margin-top:6px;">500 trials each. Red = early claim, Green = delayed claim.</div>
    </div>`;
  } catch(e) {
    el.innerHTML = `<span class="err">${e.message}</span>`;
  }
}

function renderTornadoChart(data) {
  const el = document.getElementById('tornadoResult');
  el.innerHTML = `<div class="panel accent-teal tornado-wrap">
    <h2>Sensitivity Analysis <span style="font-weight:400;font-size:0.72rem;color:var(--muted);">(base: ${data.base_pct}%)</span></h2>
    <canvas id="tornadoCanvas"></canvas>
  </div>`;
  const canvas = document.getElementById('tornadoCanvas');
  const _pname = {mean_return:'Return', volatility:'Volatility', inflation_rate:'Inflation',
                  annual_contribution:'Contribution', moat_target:'Moat Size', strict_moat_cost:'Floor Draw'};
  const labels   = data.results.map(r => _pname[r.param] || r.param);
  const posDeltas = data.results.map(r => r.pos_delta);
  const negDeltas = data.results.map(r => r.neg_delta);
  const posLabels = data.results.map(r => r.pos_label);
  const negLabels = data.results.map(r => r.neg_label);
  if (window._tornadoChart) window._tornadoChart.destroy();
  window._tornadoChart = new Chart(canvas, {
    type: 'bar',
    data: {
      labels,
      datasets: [
        { label: 'Positive change', data: posDeltas, backgroundColor: posDeltas.map(v => v >= 0 ? 'rgba(16,185,129,0.75)' : 'rgba(244,63,94,0.75)'), borderWidth: 0 },
        { label: 'Negative change', data: negDeltas, backgroundColor: negDeltas.map(v => v >= 0 ? 'rgba(16,185,129,0.75)' : 'rgba(244,63,94,0.75)'), borderWidth: 0 },
      ]
    },
    options: {
      indexAxis: 'y',
      responsive: true,
      plugins: {
        legend: { display: false },
        tooltip: {
          backgroundColor: 'rgba(9,13,26,0.92)', titleColor: '#8899b8', bodyColor: '#f0f4ff',
          padding: 10, cornerRadius: 8,
          callbacks: {
            label: (ctx) => {
              const idx = ctx.dataIndex;
              const lbl = ctx.datasetIndex === 0 ? posLabels[idx] : negLabels[idx];
              const val = ctx.raw;
              return `  ${lbl}: ${val >= 0 ? '+' : ''}${val}pp`;
            }
          }
        }
      },
      scales: {
        x: {
          ticks: { color: '#526070', callback: v => (v >= 0 ? '+' : '') + v + 'pp', font: { size: 10 } },
          grid: { color: 'rgba(255,255,255,0.04)' }
        },
        y: { ticks: { color: '#8899b8', font: { size: 10 } }, grid: { display: false } }
      }
    }
  });
}

// ── Email Digest ──────────────────────────────────────────────────────────
function loadDigestCreds() {
  const creds = JSON.parse(localStorage.getItem('retAdv_smtp') || '{}');
  const setV = (id, v) => { const el = document.getElementById(id); if (el && v) el.value = v; };
  setV('dig_user', creds.user);
  setV('dig_to',   creds.to || creds.user);
}
async function sendDigest() {
  const user = (document.getElementById('dig_user')?.value || '').trim();
  const pass = (document.getElementById('dig_pass')?.value || '').trim();
  const to   = (document.getElementById('dig_to')?.value  || user).trim();
  const st   = document.getElementById('digestStatus');
  if (!user || !pass) { if (st) st.textContent = 'Enter SMTP user and app password.'; return; }
  localStorage.setItem('retAdv_smtp', JSON.stringify({user, to}));
  if (st) st.textContent = 'Sending\u2026';
  try {
    const r = await fetch('/api/send-digest', {
      method: 'POST', headers: {'Content-Type':'application/json'},
      body: JSON.stringify({ smtp_host: 'smtp.gmail.com', smtp_port: 587, smtp_user: user, smtp_pass: pass, to_email: to })
    });
    const d = await r.json();
    if (st) st.textContent = d.ok ? `\u2713 Sent to ${d.sent_to}` : `Error: ${d.error}`;
    if (st) st.style.color = d.ok ? 'var(--green)' : 'var(--red)';
  } catch(e) {
    if (st) { st.textContent = 'Network error: ' + e.message; st.style.color = 'var(--red)'; }
  }
}

// ── Init saved scenarios on MC tab ────────────────────────────────────────
document.addEventListener('DOMContentLoaded', () => renderSavedScenarios());

// ── Roadmap ────────────────────────────────────────────────────────────────
async function loadRoadmap() {
  roadmapLoaded = true;
  const el = document.getElementById('roadmapContent');
  el.innerHTML = '<div style="color:var(--muted2);font-size:0.8rem;">Loading&hellip;</div>';
  try {
    const r = await fetch('/api/roadmap');
    const d = await r.json();
    if (d.error) { el.innerHTML = `<span class="err">${d.error}</span>`; return; }
    const cfg = d.config || {};
    const configRows = Object.entries(cfg).map(([k, v]) => {
      const fv = typeof v === 'number' && v < 1 && v > -1 ? (v*100).toFixed(1)+'%' : typeof v === 'number' ? fm(v) : v;
      return `<div style="display:flex;justify-content:space-between;padding:4px 0;border-bottom:1px solid var(--border);font-size:0.72rem;">
        <span style="color:var(--muted)">${k}</span><span style="font-family:var(--font-mono);color:var(--text)">${fv}</span></div>`;
    }).join('');
    const phaseColor = p => p === 'Sprint' ? 'var(--pink)' : p === 'Coast' ? 'var(--amber)' : 'var(--green)';
    const rows = (d.rows || []).map(row => `<tr>
      <td style="font-family:var(--font-mono)">${row.year}</td>
      <td>${row.age}</td>
      <td style="color:${phaseColor(row.phase)};font-weight:600">${row.phase}</td>
      <td style="font-family:var(--font-mono);text-align:right">${fm(row.liquid_nw)}</td>
      <td style="font-family:var(--font-mono);text-align:right">${fm(row.total_nw)}</td>
      <td style="font-family:var(--font-mono);text-align:right;color:var(--muted2)">${fm(row.sgov)}</td>
      <td style="font-family:var(--font-mono);text-align:right;color:var(--muted2)">${fm(row.schwab)}</td>
      <td style="font-family:var(--font-mono);text-align:right;color:var(--muted2)">${fm(row.roth)}</td>
      <td style="font-family:var(--font-mono);text-align:right;color:var(--muted2)">${fm(row.k401)}</td>
    </tr>`).join('');
    el.innerHTML = `
      <div style="margin-bottom:14px;">
        <div style="font-size:0.68rem;font-weight:700;color:var(--muted2);letter-spacing:0.07em;text-transform:uppercase;margin-bottom:8px;">Control Panel</div>
        <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:0 20px;">${configRows}</div>
      </div>
      <div style="overflow-x:auto;">
        <table style="width:100%;border-collapse:collapse;font-size:0.72rem;">
          <thead><tr style="border-bottom:1px solid var(--border2);">
            <th style="text-align:left;padding:6px 8px;color:var(--muted2);font-size:0.63rem;text-transform:uppercase;">Year</th>
            <th style="text-align:left;padding:6px 8px;color:var(--muted2);font-size:0.63rem;text-transform:uppercase;">Age</th>
            <th style="text-align:left;padding:6px 8px;color:var(--muted2);font-size:0.63rem;text-transform:uppercase;">Phase</th>
            <th style="text-align:right;padding:6px 8px;color:var(--muted2);font-size:0.63rem;text-transform:uppercase;">Liquid NW</th>
            <th style="text-align:right;padding:6px 8px;color:var(--muted2);font-size:0.63rem;text-transform:uppercase;">Total NW</th>
            <th style="text-align:right;padding:6px 8px;color:var(--muted2);font-size:0.63rem;text-transform:uppercase;">SGOV</th>
            <th style="text-align:right;padding:6px 8px;color:var(--muted2);font-size:0.63rem;text-transform:uppercase;">Schwab</th>
            <th style="text-align:right;padding:6px 8px;color:var(--muted2);font-size:0.63rem;text-transform:uppercase;">Roth</th>
            <th style="text-align:right;padding:6px 8px;color:var(--muted2);font-size:0.63rem;text-transform:uppercase;">401k</th>
          </tr></thead>
          <tbody>${rows}</tbody>
        </table>
      </div>`;
  } catch(e) { el.innerHTML = `<span class="err">${e.message}</span>`; }
}

// ── Transactions ───────────────────────────────────────────────────────────
async function loadTransactions(page) {
  transactionsLoaded = true;
  const el = document.getElementById('txContent');
  const pagerEl = document.getElementById('txPager');
  const month = document.getElementById('txMonthFilter')?.value || '';
  const type  = document.getElementById('txTypeFilter')?.value || '';
  el.innerHTML = '<div style="color:var(--muted2);font-size:0.8rem;">Loading&hellip;</div>';
  try {
    const qs = new URLSearchParams({ page: page||1, limit:50, ...(month&&{month}), ...(type&&{type}) });
    const r = await fetch('/api/transactions?' + qs);
    const d = await r.json();
    if (d.error) { el.innerHTML = `<span class="err">${d.error}</span>`; return; }
    _txState = { page: d.page, pages: d.pages };

    // Populate filter dropdowns on first load
    if ((d.months||[]).length && !document.getElementById('txMonthFilter').children.length > 1) {
      const ms = document.getElementById('txMonthFilter');
      const ts = document.getElementById('txTypeFilter');
      if (ms.options.length === 1) d.months.forEach(m => { const o=document.createElement('option'); o.value=m; o.text=m; ms.add(o); });
      if (ts.options.length === 1) (d.types||[]).forEach(t => { const o=document.createElement('option'); o.value=t; o.text=t; ts.add(o); });
    }

    const typeColor = t => ({Income:'var(--green)',Investment:'var(--sky,#38bdf8)',Expense:'var(--red)',Fun:'var(--amber)'}[t]||'var(--muted)');
    const rows = (d.rows||[]).map(row => `<tr style="border-bottom:1px solid var(--border);">
      <td style="padding:6px 8px;font-size:0.68rem;color:var(--muted2);font-family:var(--font-mono)">${row.date}</td>
      <td style="padding:6px 8px;"><span style="font-size:0.62rem;font-weight:700;color:${typeColor(row.type)};text-transform:uppercase;">${row.type}</span></td>
      <td style="padding:6px 8px;font-size:0.70rem;color:var(--muted)">${row.category}</td>
      <td style="padding:6px 8px;font-size:0.70rem;color:var(--muted2);max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">${row.memo}</td>
      <td style="padding:6px 8px;font-family:var(--font-mono);font-size:0.72rem;text-align:right;color:${row.signed>=0?'var(--green)':'var(--red)'}">${row.signed>=0?'+':''}${fm(row.signed)}</td>
      <td style="padding:6px 8px;font-size:0.65rem;color:var(--muted2)">${row.account}</td>
    </tr>`).join('');
    el.innerHTML = `<div style="overflow-x:auto;"><table style="width:100%;border-collapse:collapse;">
      <thead><tr style="border-bottom:1px solid var(--border2);">
        <th style="text-align:left;padding:6px 8px;color:var(--muted2);font-size:0.60rem;text-transform:uppercase;">Date</th>
        <th style="text-align:left;padding:6px 8px;color:var(--muted2);font-size:0.60rem;text-transform:uppercase;">Type</th>
        <th style="text-align:left;padding:6px 8px;color:var(--muted2);font-size:0.60rem;text-transform:uppercase;">Category</th>
        <th style="text-align:left;padding:6px 8px;color:var(--muted2);font-size:0.60rem;text-transform:uppercase;">Memo</th>
        <th style="text-align:right;padding:6px 8px;color:var(--muted2);font-size:0.60rem;text-transform:uppercase;">Amount</th>
        <th style="text-align:left;padding:6px 8px;color:var(--muted2);font-size:0.60rem;text-transform:uppercase;">Account</th>
      </tr></thead>
      <tbody>${rows}</tbody>
    </table></div>`;
    pagerEl.innerHTML = d.pages > 1 ? `
      <button class="btn-sm" onclick="loadTransactions(${d.page-1})" ${d.page<=1?'disabled':''} style="padding:4px 10px;">&#8249; Prev</button>
      <span>Page ${d.page} of ${d.pages} &mdash; ${d.total} transactions</span>
      <button class="btn-sm" onclick="loadTransactions(${d.page+1})" ${d.page>=d.pages?'disabled':''} style="padding:4px 10px;">Next &#8250;</button>
    ` : `<span>${d.total} transactions</span>`;
  } catch(e) { el.innerHTML = `<span class="err">${e.message}</span>`; }
}

// ── Forecast ───────────────────────────────────────────────────────────────
async function loadForecast() {
  forecastLoaded = true;
  const el = document.getElementById('forecastContent');
  el.innerHTML = '<div style="color:var(--muted2);font-size:0.8rem;">Loading&hellip;</div>';
  try {
    const r = await fetch('/api/forecast');
    const d = await r.json();
    if (d.error) { el.innerHTML = `<span class="err">${d.error}</span>`; return; }
    const c = d.calibration || {};
    const calibHtml = c.current_date ? `
      <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:10px;margin-bottom:16px;">
        <div class="stat-card accent-sky"><div class="sc-label">As Of</div><div class="sc-value" style="font-size:0.9rem;">${c.current_date}</div></div>
        <div class="stat-card accent-emerald"><div class="sc-label">Projected Checking</div><div class="sc-value" style="font-size:0.9rem;">${fm(c.projected_checking)}</div></div>
        <div class="stat-card accent-teal"><div class="sc-label">Projected SGOV</div><div class="sc-value" style="font-size:0.9rem;">${fm(c.projected_savings)}</div></div>
        <div class="stat-card accent-indigo"><div class="sc-label">Projected Total</div><div class="sc-value" style="font-size:0.9rem;">${fm(c.projected_total)}</div></div>
      </div>` : '';
    const dayClr = day => ['Sat','Sun'].includes(day) ? 'var(--muted2)' : 'var(--text)';
    const rows = (d.rows||[]).map(row => {
      const net = row.income + row.expense + row.invest;
      return `<tr style="border-bottom:1px solid var(--border);">
        <td style="padding:5px 8px;font-family:var(--font-mono);font-size:0.68rem;color:var(--muted2)">${row.date}</td>
        <td style="padding:5px 8px;font-size:0.65rem;color:${dayClr(row.day)}">${row.day}</td>
        <td style="padding:5px 8px;font-family:var(--font-mono);font-size:0.68rem;text-align:right;color:${row.income>0?'var(--green)':'var(--muted2)'}">${row.income>0?'+'+fm(row.income):'—'}</td>
        <td style="padding:5px 8px;font-family:var(--font-mono);font-size:0.68rem;text-align:right;color:${row.expense<0?'var(--red)':'var(--muted2)'}">${row.expense<0?fm(row.expense):'—'}</td>
        <td style="padding:5px 8px;font-family:var(--font-mono);font-size:0.68rem;text-align:right;color:${row.invest<0?'var(--amber)':'var(--muted2)'}">${row.invest<0?fm(row.invest):'—'}</td>
        <td style="padding:5px 8px;font-family:var(--font-mono);font-size:0.70rem;text-align:right;font-weight:600">${fm(row.checking)}</td>
        <td style="padding:5px 8px;font-family:var(--font-mono);font-size:0.70rem;text-align:right;color:var(--muted)">${fm(row.savings)}</td>
        <td style="padding:5px 8px;font-family:var(--font-mono);font-size:0.70rem;text-align:right;color:var(--pink);font-weight:700">${fm(row.total)}</td>
      </tr>`;
    }).join('');
    el.innerHTML = calibHtml + `<div style="overflow-x:auto;"><table style="width:100%;border-collapse:collapse;">
      <thead><tr style="border-bottom:1px solid var(--border2);">
        <th style="text-align:left;padding:5px 8px;color:var(--muted2);font-size:0.60rem;text-transform:uppercase;">Date</th>
        <th style="text-align:left;padding:5px 8px;color:var(--muted2);font-size:0.60rem;text-transform:uppercase;">Day</th>
        <th style="text-align:right;padding:5px 8px;color:var(--green);font-size:0.60rem;text-transform:uppercase;">In</th>
        <th style="text-align:right;padding:5px 8px;color:var(--red);font-size:0.60rem;text-transform:uppercase;">Out</th>
        <th style="text-align:right;padding:5px 8px;color:var(--amber);font-size:0.60rem;text-transform:uppercase;">Invest</th>
        <th style="text-align:right;padding:5px 8px;color:var(--muted2);font-size:0.60rem;text-transform:uppercase;">Checking</th>
        <th style="text-align:right;padding:5px 8px;color:var(--muted2);font-size:0.60rem;text-transform:uppercase;">SGOV</th>
        <th style="text-align:right;padding:5px 8px;color:var(--pink);font-size:0.60rem;text-transform:uppercase;">Total</th>
      </tr></thead>
      <tbody>${rows}</tbody>
    </table></div>`;
  } catch(e) { el.innerHTML = `<span class="err">${e.message}</span>`; }
}

async function loadTaxLoss() {
  taxlossLoaded = true;
  const sumEl = document.getElementById('tlSummary');
  const conEl = document.getElementById('tlContent');
  try {
    const r = await fetch('/api/tax-loss');
    const d = await r.json();
    if (d.error) { conEl.innerHTML = `<span class="err">${d.error}</span>`; return; }

    const netClr = d.net_carryover < 0 ? 'var(--green)' : d.net_carryover > 0 ? 'var(--red)' : 'var(--muted2)';
    sumEl.innerHTML = `
      <div class="stat-card accent-rose">
        <div class="sc-label">Net Carryover</div>
        <div class="sc-value" style="font-size:1rem;color:${netClr}">${fm(d.net_carryover)}</div>
        <div class="sc-sub">${d.net_carryover < 0 ? 'Loss carryover' : d.net_carryover > 0 ? 'Net gain realized' : 'Neutral'}</div>
      </div>
      <div class="stat-card">
        <div class="sc-label">Harvested Losses</div>
        <div class="sc-value" style="font-size:1rem;color:var(--green)">${fm(Math.abs(d.total_harvested))}</div>
        <div class="sc-sub">total captured</div>
      </div>
      <div class="stat-card">
        <div class="sc-label">Realized Gains</div>
        <div class="sc-value" style="font-size:1rem;color:var(--red)">${fm(d.total_realized)}</div>
        <div class="sc-sub">${d.entry_count} entries</div>
      </div>`;

    if (!d.rows || d.rows.length === 0) {
      conEl.innerHTML = '<div class="tl-empty">No tax-loss entries recorded yet. Add entries to your TAX-LOSS sheet in Road To FI.xlsx.</div>';
      return;
    }
    let running = 0;
    const rows = d.rows.map(row => {
      running += row.signed;
      const signedClr = row.signed < 0 ? 'var(--green)' : row.signed > 0 ? 'var(--red)' : 'var(--muted2)';
      const runClr    = running < 0 ? 'var(--green)' : running > 0 ? 'var(--red)' : 'var(--muted2)';
      return `<tr>
        <td style="font-family:var(--font-mono);font-size:0.68rem;color:var(--muted2)">${row.date}</td>
        <td style="font-size:0.72rem">${row.action}</td>
        <td style="font-family:var(--font-mono);text-align:right">${fm(row.amount)}</td>
        <td style="font-family:var(--font-mono);text-align:right;color:${signedClr}">${row.signed >= 0 ? '+' : ''}${fm(row.signed)}</td>
        <td style="font-family:var(--font-mono);text-align:right;color:${runClr};font-weight:600">${fm(running)}</td>
        <td style="font-size:0.68rem;color:var(--muted2)">${row.notes}</td>
      </tr>`;
    }).join('');
    conEl.innerHTML = `<div style="overflow-x:auto;"><table class="tl-table">
      <thead><tr>
        <th>Date</th><th>Action</th><th style="text-align:right">Amount</th>
        <th style="text-align:right">Signed</th><th style="text-align:right">Running Net</th><th>Notes</th>
      </tr></thead>
      <tbody>${rows}</tbody>
    </table></div>`;
  } catch(e) { conEl.innerHTML = `<span class="err">${e.message}</span>`; }
}

loadHome();
</script>

<!-- ══ AI ADVISOR TAB ═══════════════════════════════════════════════════════ -->
<div id="tab-chat" class="tab-pane">
<main style="max-width:760px;margin:0 auto;display:flex;flex-direction:column;height:calc(100vh - var(--bottom-nav-h) - 48px);padding:16px 16px 8px;">
  <div style="display:flex;align-items:center;gap:10px;margin-bottom:12px;">
    <span style="font-size:1.4rem;">&#129302;</span>
    <h2 style="margin:0;font-size:1.1rem;color:var(--gold);">AI Advisor</h2>
    <select id="chatCtx" title="Context" style="margin-left:auto;background:var(--surface3);border:1px solid var(--border2);color:var(--text);border-radius:6px;padding:4px 8px;font-size:0.75rem;">
      <option value="all">All data</option>
      <option value="simulation">Last Sim</option>
      <option value="dashboard">Dashboard</option>
    </select>
    <select id="chatModel" style="background:var(--surface3);border:1px solid var(--border2);color:var(--text);border-radius:6px;padding:4px 8px;font-size:0.75rem;font-family:var(--font-mono);">
      <option value="llama3.1:8b">llama3.1:8b</option>
      <option value="llama3.2:3b">llama3.2:3b</option>
      <option value="mistral">mistral</option>
    </select>
    <button onclick="clearChat()" title="Clear history" style="background:none;border:1px solid var(--border2);color:var(--muted2);border-radius:6px;padding:4px 8px;cursor:pointer;font-size:0.85rem;">&#8635; Clear</button>
  </div>
  <div id="chatHistory" class="chat-hist" style="flex:1;overflow-y:auto;min-height:0;"></div>
  <div class="chat-input-row" style="margin-top:8px;">
    <textarea id="chatInput" placeholder="Ask about your finances\u2026" rows="2"
      style="flex:1;resize:none;background:var(--surface3);border:1px solid var(--border2);color:var(--text);border-radius:8px;padding:10px 12px;font-size:0.9rem;font-family:inherit;"
      onkeydown="if(event.key==='Enter'&&!event.shiftKey){event.preventDefault();sendChat();}"></textarea>
    <button class="chat-send" id="chatSendBtn" onclick="sendChat()">&#8593;</button>
  </div>
</main>
</div>

<script>
// ── Chat globals ──────────────────────────────────────────────────────────────
let chatMessages = [];
const CHAT_SUGGESTIONS = [
  "What\u2019s my plan\u2019s biggest tail risk?",
  "Is my SGOV moat sized correctly for my bridge period?",
  "What does sequence-of-returns research say about my setup?",
  "Did my spending spike compared to baseline?",
  "How long until I hit my FI number at current pace?"
];

function openChatTab() {
  switchTab('chat', document.querySelector('[data-tab="chat"]'));
  if (chatMessages.length === 0) showChatSuggestions();
  setTimeout(() => document.getElementById('chatInput')?.focus(), 80);
}

function showChatSuggestions() {
  const hist = document.getElementById('chatHistory');
  const sugDiv = document.createElement('div');
  sugDiv.className = 'chat-suggested';
  sugDiv.id = 'chatSuggestions';
  sugDiv.innerHTML = '<div style="font-size:0.62rem;color:var(--muted2);margin-bottom:4px;">Suggested questions:</div>' +
    CHAT_SUGGESTIONS.map(s =>
      `<button class="chat-sug-btn" onclick="useSuggestion(this,'${s.replace(/'/g,"\\'")}')">${s}</button>`
    ).join('');
  hist.appendChild(sugDiv);
}

function useSuggestion(btn, text) {
  document.getElementById('chatSuggestions')?.remove();
  document.getElementById('chatInput').value = text;
  sendChat();
}

function clearChat() {
  chatMessages = [];
  const hist = document.getElementById('chatHistory');
  hist.innerHTML = '';
  showChatSuggestions();
}

async function sendChat() {
  const input   = document.getElementById('chatInput');
  const sendBtn = document.getElementById('chatSendBtn');
  const hist    = document.getElementById('chatHistory');
  const msg     = input.value.trim();
  if (!msg) return;

  document.getElementById('chatSuggestions')?.remove();
  input.value = '';
  sendBtn.disabled = true;

  const userBubble = document.createElement('div');
  userBubble.className = 'chat-msg-user';
  userBubble.textContent = msg;
  hist.appendChild(userBubble);

  const thinkBubble = document.createElement('div');
  thinkBubble.className = 'chat-msg-ai ai-thinking';
  thinkBubble.innerHTML = '<div class="ai-thinking-dots"><span></span><span></span><span></span></div><span>Thinking\u2026</span>';
  hist.appendChild(thinkBubble);
  hist.scrollTop = 99999;

  chatMessages.push({role: 'user', content: msg});

  const ctx   = document.getElementById('chatCtx')?.value || 'all';
  const model = document.getElementById('chatModel')?.value || document.getElementById('aiModel')?.value || 'llama3.1:8b';
  const payload = {
    message:        msg,
    context_type:   ctx,
    model:          model,
    history:        chatMessages.slice(-8),
    sim_data:       window._lastMCResult  || null,
    dashboard_data: window._lastDashboard || null,
  };

  try {
    const r = await fetch('/api/chat', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(payload)
    });
    const d = await r.json();
    thinkBubble.remove();
    const aiBubble = document.createElement('div');
    aiBubble.className = 'chat-msg-ai';
    if (d.error) {
      aiBubble.style.color = 'var(--red)';
      aiBubble.textContent = '\u26a0 ' + d.error;
    } else {
      aiBubble.innerHTML = mdLite(d.reply || '');
      chatMessages.push({role: 'assistant', content: d.reply || ''});
    }
    hist.appendChild(aiBubble);
  } catch(e) {
    thinkBubble.remove();
    const errBubble = document.createElement('div');
    errBubble.className = 'chat-msg-ai';
    errBubble.style.color = 'var(--red)';
    errBubble.textContent = '\u26a0 Could not reach server: ' + e.message;
    hist.appendChild(errBubble);
  }
  hist.scrollTop = 99999;
  sendBtn.disabled = false;
  input.focus();
}
</script>

</body>
</html>"""

async def dashboard(request: Request):
    return HTMLResponse(DASHBOARD_HTML, headers={"Cache-Control": "no-store, no-cache, must-revalidate", "Pragma": "no-cache"})

# ── OAuth bypass (local LAN trust) ────────────────────────────────────────────

async def oauth_metadata(request: Request):
    base = str(request.base_url).rstrip("/")
    return JSONResponse({
        "issuer": base,
        "authorization_endpoint": f"{base}/oauth/authorize",
        "token_endpoint": f"{base}/oauth/token",
        "response_types_supported": ["code"],
        "grant_types_supported": ["authorization_code", "client_credentials"],
        "token_endpoint_auth_methods_supported": ["none"]
    })

async def oauth_token(request: Request):
    return JSONResponse({
        "access_token": "local-lan-bypass",
        "token_type": "bearer",
        "expires_in": 86400
    })

# ── PWA assets ────────────────────────────────────────────────────────────────

MANIFEST_JSON = """{
  "name": "Road To FI",
  "short_name": "Road To FI",
  "description": "Personal retirement & FI tracking dashboard",
  "start_url": "/",
  "display": "standalone",
  "background_color": "#080d1a",
  "theme_color": "#080d1a",
  "icons": [
    { "src": "/icon-192.svg", "sizes": "192x192", "type": "image/svg+xml", "purpose": "any maskable" },
    { "src": "/icon-512.svg", "sizes": "512x512", "type": "image/svg+xml", "purpose": "any maskable" }
  ]
}"""

ICON_SVG = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 192 192">
  <rect width="192" height="192" rx="32" fill="#080d1a"/>
  <text x="96" y="130" font-size="120" text-anchor="middle" font-family="system-ui" fill="#ec4899">$</text>
</svg>"""

SW_JS = """const CACHE = 'road-to-fi-v3';
self.addEventListener('install', e => { self.skipWaiting(); });
self.addEventListener('activate', e => {
  e.waitUntil(caches.keys().then(keys => Promise.all(keys.map(k => caches.delete(k)))).then(() => self.clients.claim()));
});
self.addEventListener('fetch', e => {
  if (e.request.method !== 'GET') return;
  if (e.request.url.includes('/api/')) return;
  if (e.request.mode === 'navigate') { e.respondWith(fetch(e.request)); return; }
});"""

async def manifest(request: Request):
    from starlette.responses import Response
    return Response(MANIFEST_JSON, media_type="application/manifest+json")

async def service_worker(request: Request):
    from starlette.responses import Response
    return Response(SW_JS, media_type="application/javascript")

async def icon_svg(request: Request):
    from starlette.responses import Response
    return Response(ICON_SVG, media_type="image/svg+xml")

# ── App assembly ──────────────────────────────────────────────────────────────

mcp_app = mcp.sse_app()
app = Starlette(routes=[
    Route("/", dashboard),
    Route("/manifest.json", manifest),
    Route("/sw.js", service_worker),
    Route("/icon-192.svg", icon_svg),
    Route("/icon-512.svg", icon_svg),
    Route("/api/rules", api_rules),
    Route("/api/ledger/dashboard", api_ledger_dashboard),
    Route("/api/monte-carlo", api_monte_carlo, methods=["POST"]),
    Route("/api/portfolio", api_portfolio),
    Route("/api/portfolio/refresh", api_portfolio_refresh, methods=["POST"]),
    Route("/api/stock-price", api_stock_price, methods=["POST"]),
    Route("/api/send-digest", api_send_digest, methods=["POST"]),
    Route("/api/optimize-contribution", api_optimize_contribution, methods=["POST"]),
    Route("/api/sensitivity", api_sensitivity, methods=["POST"]),
    Route("/api/ss-sensitivity", api_ss_sensitivity, methods=["POST"]),
    Route("/api/roadmap", api_roadmap),
    Route("/api/transactions", api_transactions),
    Route("/api/forecast", api_forecast),
    Route("/api/tax-loss", api_tax_loss),
    Route("/api/grid-search", api_grid_search, methods=["POST"]),
    Route("/api/chat", api_chat, methods=["POST"]),
    Route("/api/chat/stream", api_chat_stream, methods=["POST"]),
    Route("/api/summarize", api_summarize, methods=["POST"]),
    Route("/.well-known/oauth-authorization-server", oauth_metadata),
    Route("/oauth/token", oauth_token, methods=["GET", "POST"]),
    Mount("/", mcp_app),
])

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
